# dashboard_kpp_final_v5.py
# PT KALIMANTAN PRIMA PERSADA - TWB Analytics Dashboard
# FINAL VERSION v5.0 - Trend Analysis Diubah Menjadi Financial Analysis
# Run: streamlit run dashboard_kpp_final_v5.py

import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from PIL import Image
import os
import base64
import io
from datetime import datetime
from pathlib import Path
import openpyxl
import openpyxl.styles

st.set_page_config(
    page_title="Mining Volume Deviation Monitoring PT. KPP",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Path logo - sesuaikan dengan path logo Anda
BASE_DIR = Path(__file__).resolve().parent
LOGO_PATH = BASE_DIR / "assets" / "logo.png"
DEMO_DATA_DIR = BASE_DIR / "demo_data"

def get_available_demo_files():
    if not DEMO_DATA_DIR.exists():
        return {}

    excel_files = sorted(DEMO_DATA_DIR.glob("*.xlsx"))
    return {file.stem: file for file in excel_files}

# Fungsi untuk mengonversi gambar ke base64 dengan resize
def get_base64_image(image_path, size=(120, 120)):  # Diperbesar dari 60 ke 120
    try:
        # Baca gambar
        if Path(image_path).exists():
            img = Image.open(image_path)
            
            # Resize gambar ke ukuran yang diinginkan
            img = img.resize(size, Image.Resampling.LANCZOS)
            
            # Konversi ke base64
            import io
            buffered = io.BytesIO()
            img.save(buffered, format="PNG")
            return base64.b64encode(buffered.getvalue()).decode()
        else:
            st.warning(f"⚠️ Logo tidak ditemukan di path: {image_path}")
            return None
    except Exception as e:
        st.warning(f"⚠️ Error loading logo: {str(e)}")
        return None

# Fungsi untuk menampilkan logo dengan berbagai metode
def display_logo(logo_path, width=120):
    try:
        logo_path = Path(logo_path)
        if logo_path.exists():
            img = Image.open(logo_path)
            img = img.resize((width, width), Image.Resampling.LANCZOS)
            return img
        else:
            from PIL import ImageDraw, ImageFont
            img = Image.new('RGB', (width, width), color=(0, 40, 80))
            d = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("arial.ttf", 40)
            except Exception:
                font = ImageFont.load_default()
            d.text((width // 2 - 40, width // 2 - 20), "KPP", fill=(0, 100, 136), font=font)
            return img
    except Exception as e:
        st.error(f"Error loading logo: {e}")
        from PIL import ImageDraw, ImageFont
        img = Image.new('RGB', (width, width), color=(0, 40, 80))
        d = ImageDraw.Draw(img)
        try:
            font = ImageFont.truetype("arial.ttf", 40)
        except Exception:
            font = ImageFont.load_default()
        d.text((width // 2 - 40, width // 2 - 20), "KPP", fill=(0, 255, 136), font=font)
        return img
        
    except Exception as e:
        st.error(f"Error loading logo: {e}")
        # Return placeholder
        from PIL import ImageDraw, ImageFont
        img = Image.new('RGB', (width, width), color=(0, 40, 80))
        d = ImageDraw.Draw(img)
        try:
            font = ImageFont.truetype("arial.ttf", 40)  # Diperbesar font
        except:
            font = ImageFont.load_default()
        d.text((width//2-40, width//2-20), "KPP", fill=(0, 255, 136), font=font)
        return img

# Coba konversi logo ke base64
logo_base64 = get_base64_image(LOGO_PATH, size=(200, 200))

# COMPLETE CSS - DARK MODE PREMIUM (DENGAN WARNA YANG DIPERBAIKI)

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800;900&display=swap');

* {{
    font-family: 'Inter', sans-serif;
    margin: 0;
    padding: 0;
}}

.block-container {{
    background: linear-gradient(135deg, #0a0a0f 0%, #1a1a2e 50%, #16213e 100%);
    color: #e5e7eb;
    padding: 1rem 2rem 2rem 2rem !important;
    max-width: 100% !important;
}}

[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, #0f0f1a 0%, #0a0a0f 100%);
    border-right: 1px solid rgba(255,255,255,0.1);
}}

[data-testid="stSidebar"] * {{
    color: #e5e7eb !important;
}}

/* JANGAN SEMBUNYIKAN HEADER */
footer {{ visibility: hidden; }}

/* Tambahkan style untuk header Streamlit */
header {{
    background: transparent !important;
}}

/* HEADER - DENGAN LOGO BESAR DAN BORDER */
.main-header {{
    background: linear-gradient(135deg, #000000 0%, #1a1a2e 70%, #16213e 100%);
    padding: 1.8rem 2.5rem;  /* Diperbesar padding untuk logo besar */
    border-radius: 0 0 28px 28px;
    margin: -1rem -2rem 2rem -2rem;
    box-shadow: 0 20px 60px rgba(0,0,0,0.6);
    border: 1px solid rgba(255,255,255,0.12);
}}

.header-container {{
    display: flex;
    align-items: center;
    gap: 2rem;  /* Diperbesar gap */
}}

.logo-container {{
    flex-shrink: 0;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 3px;
    background: #ffffff;
    border-radius: 20px;
    border: 3px solid #22c55e;  /* Border hijau solid lebih tebal */
    box-shadow: 
        0 0 20px rgba(0,255,136,0.4),    /* Glow pertama */
        0 0 40px rgba(0,255,136,0.2),    /* Glow kedua */
        0 8px 30px rgba(0,0,0,0.1);      /* Shadow biasa */
}}

.logo-img {{
    width: 120;  /* DIPERBESAR dari 60px ke 120px */
    height: 120px; /* DIPERBESAR dari 60px ke 120px */
    border-radius: 14px;
    object-fit: contain;
    background: rgba(255,255,255,0.03);
    padding: 3px;
}}

/* TITLE SECTION */
.title-wrapper {{
    flex-grow: 1;
    text-align: left;
    padding-left: 0.5rem;
}}

.main-title {{
    font-size: 2.4rem;  /* Sedikit diperbesar */
    font-weight: 900;
    margin: 0 0 0.4rem 0;
    background: linear-gradient(135deg, #22c55e 0%, #00ffcc 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    line-height: 1.2;
    letter-spacing: -0.5px;
}}

.main-subtitle {{
    color: #d1d5db;
    font-size: 1.1rem;  /* Diperbesar */
    font-weight: 400;
    letter-spacing: 0.3px;
    margin: 0;
    padding-left: 0.2rem;
}}

/* STATS SECTION */
.stats-grid {{
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 0.8rem;
    flex-shrink: 0;
    min-width: 220px;
}}

.stat-card {{
    background: linear-gradient(135deg, rgba(255,255,255,0.1) 0%, rgba(255,255,255,0.05) 100%);
    backdrop-filter: blur(20px);
    border-radius: 14px;
    padding: 0.9rem;  /* Sedikit diperbesar */
    text-align: center;
    border: 1px solid rgba(255,255,255,0.15);
    transition: all 0.3s ease;
}}

.stat-card:hover {{
    transform: translateY(-3px);
    box-shadow: 0 10px 30px rgba(0,255,136,0.2);
}}

.stat-value {{
    font-size: 1.5rem;  /* Diperbesar */
    font-weight: 900;
    color: #00ff88;
    display: block;
    line-height: 1;
}}

.stat-label {{
    font-size: 0.75rem;  /* Diperbesar */
    color: #9ca3af;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-top: 0.4rem;
    display: block;
}}

/* PERFORMANCE CARDS */
.perf-card {{
    background: linear-gradient(135deg, rgba(20,20,35,0.95) 0%, rgba(15,15,25,0.95) 100%);
    backdrop-filter: blur(25px);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 22px;
    padding: 2rem 1.8rem;
    box-shadow: 0 18px 55px rgba(0,0,0,0.5);
    transition: all 0.35s cubic-bezier(0.4, 0, 0.2, 1);
    position: relative;
    overflow: hidden;
}}

.perf-card::before {{
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 4px;
    background: var(--accent-color);
}}

.perf-card:hover {{
    transform: translateY(-10px);
    box-shadow: 0 25px 70px rgba(0,0,0,0.6);
}}

.perf-card.normal {{ --accent-color: linear-gradient(90deg, #22c55e, #10b981); }}
.perf-card.caution {{ --accent-color: linear-gradient(90deg, #eab308, #f59e0b); }} /* KUNING */
.perf-card.critical {{ --accent-color: linear-gradient(90deg, #ef4444, #dc2626); }} /* MERAH */
.perf-card.info {{ --accent-color: linear-gradient(90deg, #3b82f6, #2563eb); }}

.card-icon {{
    font-size: 2.6rem;
    margin-bottom: 1rem;
    display: block;
    opacity: 0.95;
}}

.card-label {{
    font-size: 0.74rem;
    color: #9ca3af;
    text-transform: uppercase;
    letter-spacing: 1.1px;
    font-weight: 600;
    margin-bottom: 0.8rem;
}}

.card-value {{
    font-size: 2.6rem;
    font-weight: 900;
    color: #ffffff;
    line-height: 1;
    margin-bottom: 0.6rem;
}}

.card-subtitle {{
    font-size: 0.86rem;
    color: #d1d5db;
    margin-bottom: 1rem;
}}

/* STATUS BADGES - WARNA DIPERBAIKI (Caution=Kuning, Critical=Merah) */
.status-badge {{
    padding: 0.55rem 1.25rem;
    border-radius: 50px;
    font-weight: 700;
    font-size: 0.8rem;
    border: 2px solid;
    display: inline-flex;
    align-items: center;
    gap: 0.6rem;
    transition: all 0.3s ease;
}}

.status-badge::before {{
    content: '';
    width: 8px;
    height: 8px;
    border-radius: 50%;
    animation: pulseDot 2s ease-in-out infinite;
}}

@keyframes pulseDot {{
    0%, 100% {{ opacity: 1; transform: scale(1); }}
    50% {{ opacity: 0.6; transform: scale(1.3); }}
}}

/* WARNA BENAR: Normal=Hijau, Caution=Kuning, Critical=Merah */
.badge-normal {{ 
    background: rgba(34,197,94,0.15); 
    color: #22c55e; 
    border-color: #22c55e; 
}}
.badge-normal::before {{ background: #22c55e; }}

.badge-caution {{ 
    background: rgba(234,179,8,0.15); 
    color: #eab308; 
    border-color: #eab308; 
}}
.badge-caution::before {{ background: #eab308; }}

.badge-critical {{ 
    background: rgba(239,68,68,0.15); 
    color: #ef4444; 
    border-color: #ef4444; 
}}
.badge-critical::before {{ background: #ef4444; }}

/* SECTION HEADERS */
.section-header {{
    display: flex;
    align-items: center;
    margin: 2rem 0 1rem 0;
    padding: 0.65rem 1.2rem;
    background: linear-gradient(135deg, rgba(5,46,22,0.55), rgba(15,23,42,0.7));
    border-left: 3px solid #22c55e;
    border-radius: 0 10px 10px 0;
    gap: 0;
}}
.section-header .section-icon {{ display: none; }}
.section-title {{
    font-size: 1.15rem !important;
    font-weight: 700 !important;
    color: #e2e8f0 !important;
    margin: 0 !important;
    padding: 0 !important;
    letter-spacing: 0.01em;
    line-height: 1.3;
}}

/* MATERIAL FLOW */
.flow-container {{
    background: linear-gradient(135deg, rgba(15,15,30,0.95) 0%, rgba(10,10,20,0.95) 100%);
    backdrop-filter: blur(30px);
    border: 2px solid rgba(255,255,255,0.15);
    border-radius: 26px;
    padding: 2.5rem;
    margin: 1.5rem 0 2rem 0;
    box-shadow: 0 22px 65px rgba(0,0,0,0.5);
}}

.flow-title {{
    text-align: center;
    font-size: 1.3rem;
    font-weight: 800;
    color: #ffffff;
    margin-bottom: 2rem;
    letter-spacing: 0.3px;
}}

.flow-chain {{
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 1.8rem;
    flex-wrap: wrap;
}}

.flow-item {{
    background: linear-gradient(135deg, rgba(255,255,255,0.08) 0%, rgba(255,255,255,0.03) 100%);
    border: 2px solid rgba(255,255,255,0.15);
    border-radius: 20px;
    padding: 1.8rem 1.2rem;
    text-align: center;
    min-width: 145px;
    transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
    box-shadow: 0 10px 30px rgba(0,0,0,0.4);
}}

.flow-item:hover {{
    transform: translateY(-12px) scale(1.05);
    border-color: var(--flow-color);
    box-shadow: 0 20px 50px rgba(0,0,0,0.6);
}}

.flow-1 {{ --flow-color: #3b82f6; }}
.flow-2 {{ --flow-color: #22c55e; }}
.flow-3 {{ --flow-color: #eab308; }}
.flow-4 {{ --flow-color: #ef4444; }}
.flow-5 {{ --flow-color: #8b5cf6; }}

.flow-item-title {{
    font-weight: 900;
    font-size: 1rem;
    color: #ffffff;
    margin-bottom: 0.3rem;
    letter-spacing: 0.2px;
}}

.flow-item-subtitle {{
    font-size: 0.72rem;
    color: #9ca3af;
    margin-bottom: 0.8rem;
}}

.flow-item-value {{
    font-size: 1.4rem;
    font-weight: 800;
    color: var(--flow-color);
    margin-top: 0.5rem;
}}

.flow-arrow {{
    font-size: 2.2rem;
    color: #00ff88;
    animation: arrowSlide 2s ease-in-out infinite;
    filter: drop-shadow(0 0 15px rgba(0,255,136,0.4));
}}

@keyframes arrowSlide {{
    0%, 100% {{ transform: translateX(0); opacity: 0.8; }}
    50% {{ transform: translateX(12px); opacity: 1; }}
}}

.flow-efficiency {{
    text-align: center;
    margin-top: 1.5rem;
    padding-top: 1.5rem;
    border-top: 1px solid rgba(255,255,255,0.1);
}}

.flow-eff-label {{
    color: #9ca3af;
    font-size: 0.95rem;
    margin-bottom: 0.5rem;
}}

.flow-eff-value {{
    color: #00ff88;
    font-size: 1.8rem;
    font-weight: 900;
}}

/* PIE CHARTS */
.pie-grid {{
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 2rem;
    margin: 2rem 0;
}}

.pie-card {{
    background: linear-gradient(135deg, rgba(15,15,30,0.95) 0%, rgba(10,10,20,0.95) 100%);
    border-radius: 20px;
    padding: 1.8rem;
    backdrop-filter: blur(25px);
    border: 1px solid rgba(255,255,255,0.12);
    box-shadow: 0 15px 45px rgba(0,0,0,0.4);
    transition: all 0.3s ease;
}}

.pie-card:hover {{
    transform: translateY(-8px);
    box-shadow: 0 20px 60px rgba(0,0,0,0.5);
    border-color: rgba(0,255,136,0.3);
}}

.pie-title {{
    font-size: 1.1rem;
    font-weight: 800;
    color: #ffffff;
    margin-bottom: 1.2rem;
    text-align: center;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 0.6rem;
}}

/* MATRIX TABLE */
.matrix-container {{
    background: linear-gradient(135deg, rgba(15,15,30,0.95) 0%, rgba(10,10,20,0.95) 100%);
    backdrop-filter: blur(25px);
    border-radius: 22px;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.12);
    margin: 2rem 0;
    box-shadow: 0 20px 60px rgba(0,0,0,0.5);
}}

.matrix-header {{
    background: linear-gradient(135deg, rgba(0,255,136,0.15) 0%, rgba(0,255,136,0.05) 100%);
    padding: 1.3rem 2rem;
    border-bottom: 2px solid rgba(255,255,255,0.1);
}}

.matrix-title {{
    color: #ffffff;
    font-size: 1.2rem;
    font-weight: 800;
    margin: 0;
}}

.matrix-row {{
    display: grid;
    grid-template-columns: 2fr 1fr 1fr 1fr 1fr 1.2fr;
    gap: 1rem;
    padding: 1.4rem 2rem;
    border-bottom: 1px solid rgba(255,255,255,0.05);
    align-items: center;
    transition: all 0.3s ease;
}}

.matrix-row:hover {{
    background: rgba(0,255,136,0.05);
}}

.matrix-row:last-child {{
    border-bottom: none;
}}

.matrix-stage {{
    font-weight: 700;
    color: #ffffff;
    font-size: 1rem;
}}

.matrix-cell {{
    text-align: center;
}}

.matrix-number {{
    font-size: 1.2rem;
    font-weight: 800;
    display: block;
    line-height: 1;
}}

.matrix-label {{
    font-size: 0.7rem;
    color: #9ca3af;
    display: block;
    margin-top: 0.3rem;
    text-transform: uppercase;
}}

/* WARNA MATRIX: Normal=Hijau, Caution=Kuning, Critical=Merah */
.num-normal {{ color: #22c55e; }}
.num-caution {{ color: #eab308; }}
.num-critical {{ color: #ef4444; }}
.num-total {{ color: #ffffff; }}
.num-avg {{ color: #00ff88; }}

/* TABS */
.stTabs [data-baseweb="tab-list"] {{
    gap: 1rem;
    background: transparent;
    border-bottom: 2px solid rgba(255,255,255,0.1);
    padding: 0;
}}

.stTabs [data-baseweb="tab"] {{
    border-radius: 14px 14px 0 0;
    padding: 1rem 2.5rem;
    font-weight: 700;
    font-size: 1.05rem;
    border: none;
    background: rgba(255,255,255,0.05);
    color: #9ca3af;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
}}

.stTabs [data-baseweb="tab"]:hover {{
    background: rgba(255,255,255,0.1);
    color: #e5e7eb;
    transform: translateY(-2px);
}}

.stTabs [aria-selected="true"] {{
    background: linear-gradient(135deg, #00ff88 0%, #00cc6a 100%) !important;
    color: #000000 !important;
    font-weight: 900 !important;
    box-shadow: 0 8px 25px rgba(0,255,136,0.4);
    transform: translateY(-2px);
}}

/* GENERAL */
h1, h2, h3, h4, h5, h6 {{
    color: #ffffff !important;
}}

p, span, div, label {{
    color: #e5e7eb;
}}

[data-testid="stMetric"] {{
    background: rgba(15,15,30,0.9);
    padding: 1.3rem;
    border-radius: 16px;
    border: 1px solid rgba(255,255,255,0.1);
}}

[data-testid="stMetricLabel"] {{
    color: #9ca3af !important;
    font-size: 0.85rem !important;
    font-weight: 600 !important;
}}

[data-testid="stMetricValue"] {{
    color: #ffffff !important;
    font-weight: 800 !important;
}}

/* Data Table Styling */
.data-table-container {{
    background: linear-gradient(135deg, rgba(15,15,30,0.95) 0%, rgba(10,10,20,0.95) 100%);
    backdrop-filter: blur(25px);
    border-radius: 18px;
    padding: 1.5rem;
    border: 1px solid rgba(255,255,255,0.12);
    margin-top: 1.5rem;
    box-shadow: 0 15px 40px rgba(0,0,0,0.4);
}}

/* Financial Cards Styling */
.financial-card {{
    background: linear-gradient(135deg, rgba(20,20,35,0.95) 0%, rgba(15,15,25,0.95) 100%);
    backdrop-filter: blur(25px);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 16px;
    padding: 1.5rem;
    box-shadow: 0 12px 40px rgba(0,0,0,0.4);
    margin: 0.5rem 0;
    text-align: center;
    height: 100%;
}}

.financial-main-value {{
    font-size: 2rem;
    font-weight: 900;
    color: #00ff88;
    margin-bottom: 0.5rem;
}}

.financial-main-label {{
    font-size: 0.9rem;
    color: #9ca3af;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 1rem;
}}

.financial-sub-metric {{
    font-size: 1.2rem;
    font-weight: 700;
    color: #ffffff;
    margin-top: 0.5rem;
}}

.financial-sub-label {{
    font-size: 0.8rem;
    color: #9ca3af;
}}

.aging-container {{
    background: linear-gradient(135deg, rgba(15,15,30,0.95) 0%, rgba(10,10,20,0.95) 100%);
    border-radius: 16px;
    padding: 1.5rem;
    border: 1px solid rgba(255,255,255,0.12);
    margin-top: 1.5rem;
    height: 100%;
}}
</style>
""", unsafe_allow_html=True)

# HELPER FUNCTIONS
def get_kpi_status(dev_pct):
    """Status: Normal ≤2%, Caution 2-3%, Critical >3%"""
    if pd.isna(dev_pct): 
        return 'Unknown', 'info'
    abs_dev = abs(dev_pct)
    if abs_dev <= 2: 
        return 'Normal', 'normal'
    elif abs_dev <= 3: 
        return 'Caution', 'caution'
    else: 
        return 'Critical', 'critical'

def format_number(num, decimals=2):
    if pd.isna(num): 
        return '-'
    return f"{num:,.{decimals}f}"

def format_large(num):
    if pd.isna(num): 
        return '-'
    if abs(num) >= 1e6: 
        return f"{num/1e6:.2f}M"
    elif abs(num) >= 1e3: 
        return f"{num/1e3:.1f}K"
    return f"{num:.0f}"

def get_active_data_source():
    available_demo_files = get_available_demo_files()

    with st.sidebar:
        st.markdown("### DATA SOURCE")

        source_mode = st.radio(
            "Pilih sumber data:",
            ["Data Training", "Upload Excel"],
            index=0,
            key="data_source_mode"
        )

        selected_file = None
        selected_file_name = None
        selected_source_type = None

        if source_mode == "Data Training":
            if not available_demo_files:
                st.warning("Folder demo_data belum berisi file demo (.xlsx).")
            else:
                demo_choice = st.selectbox(
                    "Data Training:",
                    list(available_demo_files.keys()),
                    key="demo_file_choice"
                )
                selected_file = available_demo_files[demo_choice]
                selected_file_name = available_demo_files[demo_choice].name
                selected_source_type = "demo"
                st.success(f"File Aktif: {selected_file_name}")

        else:
            uploaded = st.file_uploader(
                "Upload Excel File",
                type=["xlsx", "xls"],
                key="excel_uploader"
            )
            if uploaded is not None:
                selected_file = uploaded
                selected_file_name = uploaded.name
                selected_source_type = "upload"
                st.success(f"File aktif: {uploaded.name}")

    return selected_file, selected_file_name, selected_source_type

# Tambahkan fungsi ini setelah fungsi format_large
# ========== HELPER FUNCTIONS ==========
def sort_months_chronologically(month_list):
    """Mengurutkan bulan secara kronologis dari Januari-Desember"""
    month_order = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 
                   'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    
    def month_key(month_str):
        # Format: "Februari-2024" atau "Februari"
        try:
            # Pisahkan bulan dan tahun
            if '-' in month_str:
                parts = month_str.split('-')
                month_part = parts[0]
                # Cari tahun jika ada
                if len(parts) > 1 and parts[1].isdigit():
                    year_part = int(parts[1])
                else:
                    year_part = 2024
            else:
                month_part = month_str
                year_part = 2024
            
            # Cari indeks bulan
            if month_part in month_order:
                month_idx = month_order.index(month_part)
            else:
                # Jika bulan tidak dikenal, coba parse sebagai angka
                try:
                    month_idx = int(month_part) - 1
                except:
                    month_idx = 0
            
            return (year_part, month_idx)
        except:
            return (2024, 0)
    
    return sorted(month_list, key=month_key)

def get_status_color(status):
    """WARNA DIPERBAIKI: Normal=Hijau, Caution=Kuning, Critical=Merah"""
    if status == 'Normal':
        return '#22c55e'  # HIJAU
    elif status == 'Caution':
        return '#eab308'  # KUNING
    elif status == 'Critical':
        return '#ef4444'  # MERAH
    else:
        return '#9ca3af'
def render_status_distribution(status_col, total_count, key_prefix=""):
    """Reusable status distribution component"""
    if total_count > 0:
        status_counts = status_col.value_counts()
        cols1, cols2, cols3 = st.columns(3)
        status_order = ['Normal', 'Caution', 'Critical']
        for idx, status in enumerate(status_order):
            count = status_counts.get(status, 0)
            percentage = (count / total_count * 100) if total_count > 0 else 0
            color = {'Normal': '#22c55e', 'Caution': '#eab308', 'Critical': '#ef4444'}.get(status, '#9ca3af')
            col_status = [cols1, cols2, cols3][idx]
            with col_status:
                st.markdown(f"""
                    <div style='text-align:center;margin:5px 0;padding:12px;
                                background:rgba(255,255,255,0.05);border-radius:10px;
                                border-left:4px solid {color}'>
                        <div style='font-size:0.85rem;font-weight:600;color:{color};
                                    margin-bottom:5px;text-transform:uppercase'>{status}</div>
                        <div style='font-size:1.2rem;font-weight:800;color:#ffffff;line-height:1'>{count}</div>
                        <div style='font-size:0.75rem;color:#9ca3af;margin-top:3px'>{percentage:.1f}% of total</div>
                    </div>
                """, unsafe_allow_html=True)

@st.cache_data
def load_data(file):
    try:
        excel_file = pd.ExcelFile(file)
        required_sheets = ["OB Monthly", "CH CM"]

        missing_sheets = [sheet for sheet in required_sheets if sheet not in excel_file.sheet_names]
        if missing_sheets:
            st.error(f"❌ Sheet tidak ditemukan: {', '.join(missing_sheets)}")
            return None, None

        # Load OB data
        df_ob = pd.read_excel(file, sheet_name="OB Monthly")
        df_ob["Dev_Absolut"] = df_ob["JS"] - df_ob["TC"]
        df_ob["Dev_Relatif_Pct"] = ((df_ob["JS"] - df_ob["TC"]) / df_ob["TC"]) * 100
        df_ob["Status"], df_ob["Status_Color"] = zip(*df_ob["Dev_Relatif_Pct"].apply(get_kpi_status))

        # Load CH/CM data
        df_ch_cm = pd.read_excel(file, sheet_name="CH CM", skiprows=1)
        df_ch_cm.columns = ["Date", "Port_Darat", "Port_Laut", "CPP_Raw", "CPP_Product", "Sales", "CH_WB", "CM_WB"]
        df_ch_cm = df_ch_cm.dropna(subset=["Date"])
        df_ch_cm["Date"] = pd.to_datetime(df_ch_cm["Date"])

        # Hitung total stok Port & CPP
        df_ch_cm["Port_Total"] = df_ch_cm["Port_Darat"] + df_ch_cm["Port_Laut"]
        df_ch_cm["CPP_Total"] = df_ch_cm["CPP_Raw"].fillna(0) + df_ch_cm["CPP_Product"].fillna(0)

        # Baseline dari baris pertama
        if len(df_ch_cm) > 0:
            port_baseline = df_ch_cm["Port_Total"].iloc[0]
            cpp_baseline = df_ch_cm["CPP_Total"].iloc[0]

            df_ch_cm["TWB_CH"] = df_ch_cm["Sales"].fillna(0) + df_ch_cm["Port_Total"] - port_baseline
            df_ch_cm["TWB_CM"] = df_ch_cm["TWB_CH"] + df_ch_cm["CPP_Total"] - cpp_baseline
        else:
            df_ch_cm["TWB_CH"] = 0
            df_ch_cm["TWB_CM"] = 0

        # CH deviation
        df_ch_cm["Dev_CH_Absolut"] = df_ch_cm["TWB_CH"] - df_ch_cm["CH_WB"]
        df_ch_cm["Dev_CH_Relatif_Pct"] = np.where(
            df_ch_cm["CH_WB"] != 0,
            ((df_ch_cm["TWB_CH"] - df_ch_cm["CH_WB"]) / df_ch_cm["CH_WB"]) * 100,
            np.nan
        )
        df_ch_cm["Status_CH"], df_ch_cm["Status_CH_Color"] = zip(*df_ch_cm["Dev_CH_Relatif_Pct"].apply(get_kpi_status))

        # CM deviation
        df_ch_cm["Dev_CM_Absolut"] = df_ch_cm["TWB_CM"] - df_ch_cm["CM_WB"]
        df_ch_cm["Dev_CM_Relatif_Pct"] = np.where(
            df_ch_cm["CM_WB"] != 0,
            ((df_ch_cm["TWB_CM"] - df_ch_cm["CM_WB"]) / df_ch_cm["CM_WB"]) * 100,
            np.nan
        )
        df_ch_cm["Status_CM"], df_ch_cm["Status_CM_Color"] = zip(*df_ch_cm["Dev_CM_Relatif_Pct"].apply(get_kpi_status))

        return df_ob, df_ch_cm

    except Exception as e:
        st.error(f"❌ Error loading data: {str(e)}")
        return None, None

def create_matrix(df_ob, df_ch_cm):
    # OB stats
    ob_normal = len(df_ob[df_ob['Status'] == 'Normal'])
    ob_caution = len(df_ob[df_ob['Status'] == 'Caution'])
    ob_critical = len(df_ob[df_ob['Status'] == 'Critical'])
    ob_avg = df_ob['Dev_Relatif_Pct'].abs().mean()

    # CH stats
    df_ch = df_ch_cm[df_ch_cm['CH_WB'].notna()]
    ch_normal = len(df_ch[df_ch['Status_CH'] == 'Normal'])
    ch_caution = len(df_ch[df_ch['Status_CH'] == 'Caution'])
    ch_critical = len(df_ch[df_ch['Status_CH'] == 'Critical'])
    ch_avg = df_ch['Dev_CH_Relatif_Pct'].abs().mean() if len(df_ch) > 0 else 0

    # CM stats
    df_cm = df_ch_cm[df_ch_cm['CM_WB'].notna()]
    cm_normal = len(df_cm[df_cm['Status_CM'] == 'Normal'])
    cm_caution = len(df_cm[df_cm['Status_CM'] == 'Caution'])
    cm_critical = len(df_cm[df_cm['Status_CM'] == 'Critical'])
    cm_avg = df_cm['Dev_CM_Relatif_Pct'].abs().mean() if len(df_cm) > 0 else 0

    return pd.DataFrame({
        'Tahapan': ['Overburden (OB)', 'Coal Hauling (CH/Port)', 'Coal Mining (CM/CPP33)'],
        'Normal': [ob_normal, ch_normal, cm_normal],
        'Caution': [ob_caution, ch_caution, cm_caution],
        'Critical': [ob_critical, ch_critical, cm_critical],
        'Total': [len(df_ob), len(df_ch), len(df_cm)],
        'Avg': [ob_avg, ch_avg, cm_avg]
    })

def create_flow(df_ch_cm):
    df = df_ch_cm.dropna(subset=["TWB_CM", "TWB_CH"]).copy()
    if len(df) == 0:
        return None

    # Hitung perubahan stok periodik
    df["Delta CPP Stock"] = df["CPP_Total"].diff().fillna(0)
    df["Delta Port Stock"] = df["Port_Total"].diff().fillna(0)

    records = []
    for _, row in df.iterrows():
        cm_twb = row["TWB_CM"]
        ch_twb = row["TWB_CH"]
        sales = row["Sales"]

        delta_cpp = row["Delta CPP Stock"]
        delta_port = row["Delta Port Stock"]

        # Deviation berbasis konservasi massa
        # TWB_CM ≈ TWB_CH + Delta CPP Stock
        # TWB_CH ≈ Sales + Delta Port Stock
        dev_cpp = cm_twb - (ch_twb + delta_cpp)
        dev_port = ch_twb - (sales + delta_port)

        # Rasio flow, bukan recovery
        ch_ratio = (ch_twb / cm_twb * 100) if cm_twb not in [0, None] and pd.notna(cm_twb) else 0
        sales_ratio = (sales / ch_twb * 100) if ch_twb not in [0, None] and pd.notna(ch_twb) else 0

        records.append({
            "Date": row["Date"],
            "CPP Stock": row["CPP_Total"],
            "Port Stock": row["Port_Total"],
            "Delta CPP Stock": delta_cpp,
            "Delta Port Stock": delta_port,
            "CM TWB": cm_twb,
            "CH TWB": ch_twb,
            "Sales": sales,
            "Deviation CPP": dev_cpp,
            "Deviation Port": dev_port,
            "CH Flow Ratio (%)": ch_ratio,
            "Sales Flow Ratio (%)": sales_ratio,
        })

    return pd.DataFrame(records)

# CHART THEME
THEME = {
    'layout': {
        'font': {'family': 'Inter, sans-serif', 'color': '#e5e7eb'},
        'plot_bgcolor': 'rgba(0,0,0,0)',
        'paper_bgcolor': 'rgba(0,0,0,0)',
        'colorway': ['#00ff88', '#3b82f6', '#eab308', '#ef4444', '#8b5cf6'],
        'xaxis': {'gridcolor': 'rgba(255,255,255,0.1)', 'zerolinecolor': 'rgba(255,255,255,0.15)'},
        'yaxis': {'gridcolor': 'rgba(255,255,255,0.1)', 'zerolinecolor': 'rgba(255,255,255,0.15)'}
    }
}

def main():
    st.markdown('<div class="main-header">', unsafe_allow_html=True)

    logo_base64_final = get_base64_image(LOGO_PATH, size=(100, 100))

    if logo_base64_final:
        st.markdown(f"""
        <div class="header-container">
            <div class="logo-container">
                <img src="data:image/png;base64,{logo_base64_final}" class="logo-img" alt="KPP Logo">
            </div>
            <div class="title-wrapper">
                <h1 class="main-title">Mining Volume Deviation Monitoring</h1>
                <p class="main-subtitle">PT KALIMANTAN PRIMA PERSADA</p>
            </div>
            <div class="stats-grid">
                <div class="stat-card">
                    <span class="stat-value">3</span>
                    <span class="stat-label">Stages</span>
                </div>
                <div class="stat-card">
                    <span class="stat-value">100%</span>
                    <span class="stat-label">Auto</span>
                </div>
                <div class="stat-card">
                    <span class="stat-value">2026</span>
                    <span class="stat-label">Year</span>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)

    selected_file, selected_file_name, selected_source_type = get_active_data_source()

    with st.sidebar:
        st.markdown("---")
        st.markdown("### KPI REFERENCE")
        st.markdown("""
        <div style='background:rgba(255,255,255,0.08);padding:1.3rem;border-radius:16px;border:1px solid rgba(255,255,255,0.1)'>
            <div style='margin:0.9rem 0'>
                <span class="status-badge badge-normal"> Normal</span>
                <p style='font-size:0.82rem;margin:0.4rem 0 0;color:#9ca3af'>|ΔV%| ≤ 2%</p>
            </div>
            <div style='margin:0.9rem 0'>
                <span class="status-badge badge-caution"> Caution</span>
                <p style='font-size:0.82rem;margin:0.4rem 0 0;color:#9ca3af'>2% < |ΔV%| ≤ 3%</p>
            </div>
            <div style='margin:0.9rem 0'>
                <span class="status-badge badge-critical"> Critical</span>
                <p style='font-size:0.82rem;margin:0.4rem 0 0;color:#9ca3af'>|ΔV%| > 3%</p>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("""
        <div style='text-align:center;color:#9ca3af;margin-top:2rem'>
            <p style='color:#ffffff;font-weight:700;font-size:0.95rem'>Mining Volume Deviation Monitoring</p>
            <p style='font-size:0.85rem'>by : Michael Aragorn Purba</p>
            <p style='font-size:0.85rem'>2026</p>
        </div>
        """, unsafe_allow_html=True)

    if selected_file is None:
        st.info("Silakan Data Training atau upload file Excel dari sidebar.")
        st.markdown("""
        <div style='margin-top:2rem;padding:2rem;background:rgba(255,255,255,0.05);border-radius:20px;border:1px solid rgba(255,255,255,0.1)'>
            <h3 style='color:#00ff88;margin-bottom:1rem'>Required Format:</h3>
            <p style='color:#d1d5db'>• Sheet 1: <strong>"OB (Overburden) Monthly"</strong></p>
            <p style='color:#d1d5db'>• Sheet 2: <strong>"Coal Hauling &amp; Coal Mining"</strong></p>
            <p style='color:#d1d5db'>• Atau pilih file dari folder <strong>demo_data/</strong></p>
        </div>
        """, unsafe_allow_html=True)
        return

    with st.spinner("Processing data..."):
        df_ob, df_ch_cm = load_data(selected_file)

    if df_ob is None or df_ch_cm is None:
        st.error("Failed to load data")
        return

    source_badge_color = "#3b82f6" if selected_source_type == "upload" else "#22c55e"
    source_badge_label = "Uploaded File" if selected_source_type == "upload" else "Data Training"

    st.markdown(f"""
    <div style='margin-bottom:1rem;padding:0.8rem 1rem;background:rgba(255,255,255,0.05);
                border:1px solid rgba(255,255,255,0.1);border-radius:12px;display:flex;
                align-items:center;justify-content:space-between;gap:1rem;flex-wrap:wrap;'>
        <div>
            <div style='font-size:0.75rem;color:#9ca3af;'>Current Data Source</div>
            <div style='font-size:1rem;font-weight:700;color:#ffffff;'>{selected_file_name}</div>
        </div>
        <div style='padding:0.35rem 0.8rem;border-radius:999px;background:{source_badge_color}20;
                    color:{source_badge_color};border:1px solid {source_badge_color}55;
                    font-size:0.75rem;font-weight:700;letter-spacing:0.05em;'>
            {source_badge_label}
        </div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs([
        " Overview",
        " OB Analysis",
        " CPP33 & Port",
        " Reports"
    ])

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 1: OVERVIEW
    # ════════════════════════════════════════════════════════════════════════════
    with tab1:
        matrix = create_matrix(df_ob, df_ch_cm)

        # ══════════════════════════════════════════════════════════
        # PERFORMANCE CARDS (4 kolom) — Redesigned
        # ══════════════════════════════════════════════════════════
        st.markdown("""
        <style>
        .perf-card {
            background: linear-gradient(145deg, rgba(22,28,45,0.9), rgba(15,20,35,0.95));
            border: 1px solid rgba(148,163,184,0.08);
            border-radius: 14px;
            padding: 20px 20px 16px;
            position: relative;
            overflow: hidden;
            min-height: 155px;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }
        .perf-card:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 24px rgba(0,0,0,0.3);
        }
        .perf-card::before {
            content: '';
            position: absolute;
            top: 0; left: 0; right: 0;
            height: 3px;
            border-radius: 14px 14px 0 0;
        }
        .perf-card.normal::before  { background: linear-gradient(90deg, #22c55e, #4ade80); }
        .perf-card.caution::before { background: linear-gradient(90deg, #f59e0b, #fbbf24); }
        .perf-card.critical::before { background: linear-gradient(90deg, #ef4444, #f87171); }
        .perf-card.info::before    { background: linear-gradient(90deg, #3b82f6, #60a5fa); }

        .card-icon {
            font-size: 1.4rem;
            margin-bottom: 10px;
            display: block;
            opacity: 0.9;
        }
        .card-label {
            font-size: 0.68rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            color: #94a3b8;
            margin-bottom: 4px;
        }
        .card-value {
            font-size: 2.1rem;
            font-weight: 800;
            color: #f1f5f9;
            line-height: 1.1;
            margin-bottom: 2px;
        }
        .card-subtitle {
            font-size: 0.7rem;
            color: #64748b;
            margin-bottom: 10px;
        }
        .status-badge {
            display: inline-flex;
            align-items: center;
            gap: 5px;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 0.7rem;
            font-weight: 600;
            width: fit-content;
        }
        .badge-normal  { background: rgba(34,197,94,0.1); color: #4ade80; border: 1px solid rgba(34,197,94,0.2); }
        .badge-caution { background: rgba(245,158,11,0.1); color: #fbbf24; border: 1px solid rgba(245,158,11,0.2); }
        .badge-critical { background: rgba(239,68,68,0.1); color: #f87171; border: 1px solid rgba(239,68,68,0.2); }
        .badge-info    { background: rgba(59,130,246,0.1); color: #60a5fa; border: 1px solid rgba(59,130,246,0.2); }
        .perf-card.info .card-value { color: #93bbfc; }
        </style>
        """, unsafe_allow_html=True)

        col1, col2, col3, col4 = st.columns(4, gap="large")

        with col1:
            ob = matrix.iloc[0]
            status = 'normal' if ob['Avg'] <= 2 else 'caution' if ob['Avg'] <= 3 else 'critical'
            status_name = 'Normal' if status == 'normal' else 'Caution' if status == 'caution' else 'Critical'
            st.markdown(f"""
            <div class="perf-card {status}">
                <div>
                    <span class="card-icon"></span>
                    <div class="card-label">Overburden (OB)</div>
                    <div class="card-value">{format_number(ob['Avg'], 2)}</div>
                    <div class="card-subtitle">Avg Deviation %</div>
                </div>
                <span class="status-badge badge-{status}">● {status_name}</span>
            </div>
            """, unsafe_allow_html=True)

        with col2:
            ch = matrix.iloc[1]
            status = 'normal' if ch['Avg'] <= 2 else 'caution' if ch['Avg'] <= 3 else 'critical'
            status_name = 'Normal' if status == 'normal' else 'Caution' if status == 'caution' else 'Critical'
            st.markdown(f"""
            <div class="perf-card {status}">
                <div>
                    <span class="card-icon"></span>
                    <div class="card-label">Coal Hauling (CH)</div>
                    <div class="card-value">{format_number(ch['Avg'], 2)}</div>
                    <div class="card-subtitle">Avg Deviation %</div>
                </div>
                <span class="status-badge badge-{status}">● {status_name}</span>
            </div>
            """, unsafe_allow_html=True)

        with col3:
            cm = matrix.iloc[2]
            status = 'normal' if cm['Avg'] <= 2 else 'caution' if cm['Avg'] <= 3 else 'critical'
            status_name = 'Normal' if status == 'normal' else 'Caution' if status == 'caution' else 'Critical'
            st.markdown(f"""
            <div class="perf-card {status}">
                <div>
                    <span class="card-icon"></span>
                    <div class="card-label">Coal Mining (CM)</div>
                    <div class="card-value">{format_number(cm['Avg'], 2)}</div>
                    <div class="card-subtitle">Avg Deviation %</div>
                </div>
                <span class="status-badge badge-{status}">● {status_name}</span>
            </div>
            """, unsafe_allow_html=True)

        with col4:
            total_normal = matrix['Normal'].sum()
            total = matrix['Total'].sum()
            perf = (total_normal / total * 100) if total > 0 else 0
            st.markdown(f"""
            <div class="perf-card info">
                <div>
                    <span class="card-icon"></span>
                    <div class="card-label">Overall Performance</div>
                    <div class="card-value">{format_number(perf, 1)}</div>
                    <div class="card-subtitle">{total_normal}/{total} Periods Normal</div>
                </div>
                <span class="status-badge badge-info">● Performance Index</span>
            </div>
            """, unsafe_allow_html=True)
        # ══════════════════════════════════════════════════════════
        # EXECUTIVE SUMMARY (inserted between cards and flow)
        # ══════════════════════════════════════════════════════════
        total_normal_es = int(matrix['Normal'].sum())
        total_caution_es = int(matrix['Caution'].sum())
        total_critical_es = int(matrix['Critical'].sum())
        total_all_es = int(matrix['Total'].sum())
        perf_es = total_normal_es / total_all_es * 100 if total_all_es > 0 else 0

        ob_avg_es = matrix.iloc[0]['Avg'] if len(matrix) > 0 else 0
        ch_avg_es = matrix.iloc[1]['Avg'] if len(matrix) > 1 else 0
        cm_avg_es = matrix.iloc[2]['Avg'] if len(matrix) > 2 else 0

        worst_stage_es = "Coal Hauling (CH)" if ch_avg_es >= cm_avg_es and ch_avg_es >= ob_avg_es else \
                         "Coal Mining (CM)" if cm_avg_es >= ch_avg_es and cm_avg_es >= ob_avg_es else \
                         "Overburden (OB)"
        worst_val_es = max(ch_avg_es, cm_avg_es, ob_avg_es)
        best_stage_es = "Coal Hauling (CH)" if ch_avg_es <= cm_avg_es and ch_avg_es <= ob_avg_es else \
                        "Coal Mining (CM)" if cm_avg_es <= ch_avg_es and cm_avg_es <= ob_avg_es else \
                        "Overburden (OB)"
        best_val_es = min(ch_avg_es, cm_avg_es, ob_avg_es)

        severity_color_es = "#ef4444" if worst_val_es > 3 else "#f59e0b" if worst_val_es > 2 else "#22c55e"
        severity_text_es = "KRITIS" if worst_val_es > 3 else "PERHATIAN" if worst_val_es > 2 else "NORMAL"
        perf_c_es = '#4ade80' if perf_es >= 70 else '#fbbf24' if perf_es >= 50 else '#f87171'

        st.markdown(f"""
        <div style="background:linear-gradient(135deg,rgba(30,41,59,0.8),rgba(15,23,42,0.95));
            border:1px solid rgba(148,163,184,0.1);border-radius:14px;
            padding:20px 24px;margin:20px 0 8px 0;">
            <div style="display:flex;align-items:center;gap:10px;margin-bottom:14px;">
                <span style="background:{severity_color_es};width:10px;height:10px;border-radius:50%;display:inline-block;"></span>
                <span style="font-size:0.75rem;font-weight:700;text-transform:uppercase;
                    letter-spacing:0.1em;color:{severity_color_es};">Executive Summary — Status: {severity_text_es}</span>
            </div>
            <div style="font-size:0.88rem;color:#e2e8f0;line-height:1.7;">
                Dari <b style="color:#f1f5f9;">{total_all_es} periode</b> yang dianalisis,
                <b style="color:#4ade80;">{total_normal_es}</b> normal,
                <b style="color:#fbbf24;">{total_caution_es}</b> caution, dan
                <b style="color:#f87171;">{total_critical_es}</b> critical.
                <b style="color:{severity_color_es};">{worst_stage_es}</b> menjadi stage dengan deviasi tertinggi
                (<b style="color:{severity_color_es};">{worst_val_es:.2f}%</b>),
                sementara <b style="color:#4ade80;">{best_stage_es}</b> paling stabil
                (<b style="color:#4ade80;">{best_val_es:.2f}%</b>).
                Overall performance berada di <b style="color:{perf_c_es};">{perf_es:.1f}%</b>
                {"— perlu evaluasi menyeluruh." if perf_es < 50 else "— perlu peningkatan." if perf_es < 70 else "— dalam kondisi baik."}
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════
        # PRODUCTION FLOW SUMMARY — Animated Pipeline
        # ══════════════════════════════════════════════════════════
        st.markdown("""
        <div class="section-header">
            <div class="section-icon"></div>
            <h3 class="section-title">Production Flow Summary</h3>
        </div>
        """, unsafe_allow_html=True)

        flow = create_flow(df_ch_cm)

        if flow is not None and len(flow) > 0:
            avg_cm = flow["CM TWB"].mean()
            avg_ch = flow["CH TWB"].mean()
            avg_sales = flow["Sales"].mean()

            avg_delta_cpp = flow["Delta CPP Stock"].mean()
            avg_delta_port = flow["Delta Port Stock"].mean()

            avg_dev_cpp = flow["Deviation CPP"].mean()
            avg_dev_port = flow["Deviation Port"].mean()

            avg_ch_ratio = flow["CH Flow Ratio (%)"].mean()
            avg_sales_ratio = flow["Sales Flow Ratio (%)"].mean()

            # Skor overview sederhana berbasis deviation
            total_ref = max(abs(avg_cm) + abs(avg_ch) + abs(avg_sales), 1)
            overall_dev_score = max(0, 100 - ((abs(avg_dev_cpp) + abs(avg_dev_port)) / total_ref * 100))

            _r = 30
            _circ = 2 * 3.14159 * _r
            _score_pct = min(max(overall_dev_score, 0), 100)
            _offset = _circ * (1 - _score_pct / 100)
            _ring_color = "#22c55e" if _score_pct >= 98 else "#f59e0b" if _score_pct >= 95 else "#ef4444"

            import streamlit.components.v1 as components

            pipeline_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
            <meta charset="utf-8">
            <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{ background: transparent; font-family: 'Inter', 'Segoe UI', system-ui, sans-serif; overflow: hidden; }}
            .pipeline-wrap {{
                background: linear-gradient(135deg, rgba(30,41,59,0.85), rgba(15,23,42,0.95));
                border: 1px solid rgba(255,255,255,0.08);
                border-radius: 16px;
                padding: 32px 28px 24px;
                width: 100%;
            }}
            .pipeline-title {{
                text-align: center; font-size: 1.1rem;
                font-weight: 700; color: #e2e8f0;
                margin-bottom: 28px; letter-spacing: 0.02em;
            }}
            .pipe-chain {{
                display: flex; align-items: center;
                justify-content: center; gap: 0;
                width: 100%; padding: 0 8px;
            }}
            .pipe-node {{
                flex: 1; max-width: 220px; min-width: 160px;
                padding: 20px 18px 16px;
                border-radius: 14px; text-align: center;
                border: 1.5px solid rgba(255,255,255,0.1);
                transition: transform 0.25s ease, box-shadow 0.25s ease;
                cursor: default;
            }}
            .pipe-node:hover {{
                transform: translateY(-4px) scale(1.03);
                box-shadow: 0 8px 32px rgba(0,0,0,0.35); z-index: 2;
            }}
            .pipe-node.node-cm   {{ background: linear-gradient(145deg, #064e3b, #065f46); border-color: rgba(34,197,94,0.25); }}
            .pipe-node.node-ch   {{ background: linear-gradient(145deg, #1e3a5f, #1d4ed8); border-color: rgba(59,130,246,0.25); }}
            .pipe-node.node-sales {{ background: linear-gradient(145deg, #4c1d95, #6d28d9); border-color: rgba(167,139,250,0.25); }}

            .pipe-node-label {{ font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.1em; color: #94a3b8; margin-bottom: 2px; }}
            .pipe-node-title {{ font-size: 1.05rem; font-weight: 700; color: #f1f5f9; }}
            .pipe-node-value {{ font-size: 1.5rem; font-weight: 800; margin-top: 8px; }}
            .pipe-node.node-cm .pipe-node-value    {{ color: #4ade80; }}
            .pipe-node.node-ch .pipe-node-value    {{ color: #60a5fa; }}
            .pipe-node.node-sales .pipe-node-value {{ color: #c4b5fd; }}

            .pipe-connector {{
                display: flex; flex-direction: column; align-items: center;
                justify-content: center; flex: 0 0 130px; padding: 0 6px;
                position: relative;
            }}
            .pipe-arrow-wrap {{ position: relative; width: 100%; height: 16px; display: flex; align-items: center; }}
            .pipe-arrow-line {{
                width: 100%; height: 3px;
                background: linear-gradient(90deg, transparent 0%, #475569 12%, #475569 88%, transparent 100%);
                position: relative; overflow: hidden; border-radius: 2px;
            }}
            .pipe-arrow-line::after {{
                content: ''; position: absolute; top: -2px; left: -20px;
                width: 20px; height: 7px; border-radius: 4px;
                background: linear-gradient(90deg, transparent, #22c55e, transparent);
                animation: flowPulse 2.2s ease-in-out infinite;
            }}
            .pipe-connector.loss-connector .pipe-arrow-line::after {{
                background: linear-gradient(90deg, transparent, #3b82f6, transparent);
                animation-delay: 1.1s;
            }}
            @keyframes flowPulse {{
                0%   {{ left: -20px; opacity: 0; }}
                20%  {{ opacity: 1; }}
                80%  {{ opacity: 1; }}
                100% {{ left: calc(100% + 20px); opacity: 0; }}
            }}
            .pipe-arrow-tip {{
                width: 0; height: 0;
                border-top: 7px solid transparent; border-bottom: 7px solid transparent;
                border-left: 11px solid #475569;
                flex-shrink: 0; margin-left: -1px;
            }}
            .pipe-conn-stats {{
                display: flex; flex-direction: column; align-items: center;
                gap: 3px; margin-top: 10px;
            }}
            .pipe-loss-badge {{
                font-size: 0.72rem; padding: 3px 10px;
                border-radius: 10px; font-weight: 600;
                white-space: nowrap;
            }}
            .pipe-loss-badge.loss-val {{ background: rgba(239,68,68,0.15); color: #f87171; }}
            .pipe-loss-badge.eff-val  {{ background: rgba(34,197,94,0.12); color: #4ade80; }}

            .pipe-footer {{
                display: flex; align-items: center; justify-content: center;
                gap: 20px; margin-top: 26px; padding-top: 18px;
                border-top: 1px solid rgba(255,255,255,0.06);
            }}
            .pipe-eff-ring      {{ width: 76px; height: 76px; position: relative; flex-shrink: 0; }}
            .pipe-eff-ring svg  {{ transform: rotate(-90deg); }}
            .ring-bg            {{ fill: none; stroke: rgba(255,255,255,0.08); stroke-width: 6; }}
            .ring-fg            {{ fill: none; stroke-width: 6; stroke-linecap: round; transition: stroke-dashoffset 1s ease; }}
            .ring-label         {{ position: absolute; inset: 0; display: flex; align-items: center; justify-content: center; font-size: 0.9rem; font-weight: 800; color: #f1f5f9; }}

            .pipe-footer-text   {{ display: flex; flex-direction: column; gap: 5px; }}
            .pipe-footer-title  {{ font-size: 0.85rem; color: #94a3b8; font-weight: 600; }}
            .pipe-footer-detail {{ font-size: 0.78rem; color: #64748b; }}
            .pipe-footer-detail .hl-green {{ color: #4ade80; font-weight: 700; }}
            .pipe-footer-detail .hl-red   {{ color: #f87171; font-weight: 700; }}

            @media (max-width: 700px) {{
                .pipe-chain {{ flex-direction: column; gap: 8px; }}
                .pipe-connector {{ transform: rotate(90deg); flex: 0 0 60px; }}
            }}
            </style>
            </head>
            <body>
            <div class="pipeline-wrap">
                <div class="pipeline-title">Production Flow Summary</div>
                <div class="pipe-chain">
                    <div class="pipe-node node-cm">
                        <div class="pipe-node-label">From PIT To CPP</div>
                        <div class="pipe-node-title">Coal Mining</div>
                        <div class="pipe-node-value">{format_large(avg_cm)}</div>
                    </div>
                    <div class="pipe-connector">
                        <div class="pipe-arrow-wrap">
                            <div class="pipe-arrow-line"></div>
                            <div class="pipe-arrow-tip"></div>
                        </div>
                        <div class="pipe-conn-stats">
                            <span class="pipe-loss-badge loss-val">Δ {format_large(avg_delta_cpp)}</span>
                            <span class="pipe-loss-badge eff-val">Dev {format_large(avg_dev_cpp)}</span>
                        </div>
                    </div>
                    <div class="pipe-node node-ch">
                        <div class="pipe-node-label">From CPP To Port</div>
                        <div class="pipe-node-title">Coal Hauling</div>
                        <div class="pipe-node-value">{format_large(avg_ch)}</div>
                    </div>
                    <div class="pipe-connector loss-connector">
                        <div class="pipe-arrow-wrap">
                            <div class="pipe-arrow-line"></div>
                            <div class="pipe-arrow-tip"></div>
                        </div>
                        <div class="pipe-conn-stats">
                            <span class="pipe-loss-badge loss-val">Δ {format_large(avg_delta_port)}</span>
                            <span class="pipe-loss-badge eff-val">Dev {format_large(avg_dev_port)}</span>
                        </div>
                    </div>
                    <div class="pipe-node node-sales">
                        <div class="pipe-node-label">From Port To Customer</div>
                        <div class="pipe-node-title">Sales</div>
                        <div class="pipe-node-value">{format_large(avg_sales)}</div>
                    </div>
                </div>

                <div class="pipe-footer">
                    <div class="pipe-eff-ring">
                        <svg width="76" height="76" viewBox="0 0 76 76">
                            <circle class="ring-bg" cx="38" cy="38" r="{_r}"/>
                            <circle class="ring-fg" cx="38" cy="38" r="{_r}"
                                stroke="{_ring_color}"
                                stroke-dasharray="{_circ:.1f}"
                                stroke-dashoffset="{_offset:.1f}"/>
                        </svg>
                        <div class="ring-label">{format_number(_score_pct, 1)}%</div>
                    </div>
                    <div class="pipe-footer-text">
                        <div class="pipe-footer-title">Reconciliation Deviation Overview</div>
                        <div class="pipe-footer-detail">
                            CPP Dev <span class="hl-red">{format_large(avg_dev_cpp)}</span>
                            · Port Dev <span class="hl-red">{format_large(avg_dev_port)}</span>
                        </div>
                        <div class="pipe-footer-detail">
                            CH Ratio <span class="hl-green">{format_number(avg_ch_ratio, 1)}%</span>
                            · Sales Ratio <span class="hl-green">{format_number(avg_sales_ratio, 1)}%</span>
                        </div>
                    </div>
                </div>
            </div>
            </body>
            </html>
            """

            components.html(pipeline_html, height=340, scrolling=False)

            # ══════════════════════════════════════════════════════════
            # STATUS DISTRIBUTION BY STAGE — Fluid Modern Design
            # ══════════════════════════════════════════════════════════
            st.markdown("""
            <div class="section-header">
                <div class="section-icon"></div>
                <h3 class="section-title">Status Distribution by Stage</h3>
            </div>
            """, unsafe_allow_html=True)


            def make_status_bar(status_series, title, emoji):
                order = ['Normal', 'Caution', 'Critical']
                color_map = {'Normal': '#22C55E', 'Caution': '#F59E0B', 'Critical': '#EF4444'}
                bg_map = {'Normal': 'rgba(34,197,94,0.12)', 'Caution': 'rgba(245,158,11,0.12)', 'Critical': 'rgba(239,68,68,0.12)'}
                counts = status_series.value_counts()
                total = counts.sum()
                if total == 0:
                    return

                segments = []
                for s in order:
                    c = int(counts.get(s, 0))
                    pct = round(c / total * 100, 1) if total > 0 else 0
                    segments.append({'status': s, 'count': c, 'pct': pct, 'color': color_map[s], 'bg': bg_map[s]})

                bar_parts = ""
                for seg in segments:
                    if seg['count'] > 0:
                        bar_parts += (
                            f"<div style='flex:{seg['pct']};background:{seg['color']};"
                            f"height:100%;display:flex;align-items:center;justify-content:center;"
                            f"color:white;font-weight:700;font-size:0.78rem;min-width:32px;"
                            f"transition:flex 0.5s ease;'>"
                            f"{seg['pct']:.0f}%</div>"
                        )

                chips = ""
                for seg in segments:
                    if seg['count'] > 0:
                        chips += (
                            f"<span style='display:inline-flex;align-items:center;gap:5px;"
                            f"background:{seg['bg']};padding:4px 10px;border-radius:12px;"
                            f"font-size:0.75rem;color:#e5e7eb;margin-right:6px;font-weight:500;'>"
                            f"<span style='width:8px;height:8px;border-radius:50%;"
                            f"background:{seg['color']};display:inline-block;'></span>"
                            f"{seg['status']} <b>{seg['count']}</b> ({seg['pct']}%)"
                            f"</span>"
                        )

                html_block = f"""
                <div style="background:linear-gradient(135deg,rgba(20,20,40,0.6),rgba(15,23,42,0.8));
                    border:1px solid rgba(148,163,184,0.1);border-radius:14px;
                    padding:14px 16px;margin-bottom:8px;">
                    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px;">
                        <span style="font-size:0.85rem;font-weight:700;color:#e5e7eb;letter-spacing:0.01em;">
                            {emoji} {title}
                        </span>
                        <span style="font-size:0.72rem;color:#64748b;font-weight:500;">{total} data</span>
                    </div>
                    <div style="display:flex;height:28px;border-radius:8px;overflow:hidden;
                        gap:2px;margin-bottom:10px;">
                        {bar_parts}
                    </div>
                    <div style="display:flex;flex-wrap:wrap;gap:4px;">
                        {chips}
                    </div>
                </div>
                """
                st.markdown(html_block, unsafe_allow_html=True)


            col_b1, col_b2, col_b3 = st.columns(3)

            with col_b1:
                make_status_bar(df_ob['Status'], 'Overburden', '🟢')
            with col_b2:
                df_ch = df_ch_cm[df_ch_cm['CH_WB'].notna()]
                if len(df_ch) > 0:
                    make_status_bar(df_ch['Status_CH'], 'Coal Handling', '🔵')
            with col_b3:
                df_cm = df_ch_cm[df_ch_cm['CM_WB'].notna()]
                if len(df_cm) > 0:
                    make_status_bar(df_cm['Status_CM'], 'Coal Mining', '🟡')

            # ══════════════════════════════════════════════════════════
            # MATERIAL FLOW ANALYSIS SECTION
            # ══════════════════════════════════════════════════════════
            st.markdown("""
            <div class="section-header">
                <div class="section-icon"></div>
                <h3 class="section-title">Material Flow Analysis</h3>
            </div>
            """, unsafe_allow_html=True)


            if len(flow) > 0:


                # ══════════════════════════════════════════════════════════
                # PRODUCTION VOLUME + EFFICIENCY (Dual Y-Axis)
                # ══════════════════════════════════════════════════════════
                st.markdown("### Production Volume & Efficiency")


                from plotly.subplots import make_subplots


                flow['CM TWB Plot'] = flow['CM TWB'].replace(0, None)
                flow['CH TWB Plot'] = flow['CH TWB'].replace(0, None)
                flow['Sales Plot'] = flow['Sales'].replace(0, None)
                flow['Eff Plot'] = flow['CH Flow Ratio (%)'].replace(0, None)


                avg_eff_val = flow['CH Flow Ratio (%)'].mean()
                peak_cm = flow['CM TWB'].max()
                below_target = (flow['CH Flow Ratio (%)'] < 100).sum()
                total_p = len(flow)


                vc1, vc2, vc3, vc4 = st.columns(4)
                with vc1:
                    st.markdown(f"""
                    <div style="background:rgba(15,15,30,0.8);border-radius:10px;padding:0.8rem;
                    border:1px solid rgba(148,163,184,0.1);text-align:center">
                        <div style="color:#9ca3af;font-size:0.7rem;margin-bottom:0.2rem">Avg Volume</div>
                        <div style="color:#3b82f6;font-size:1.1rem;font-weight:700">{avg_cm:,.0f} t</div>
                    </div>""", unsafe_allow_html=True)
                with vc2:
                    st.markdown(f"""
                    <div style="background:rgba(15,15,30,0.8);border-radius:10px;padding:0.8rem;
                    border:1px solid rgba(148,163,184,0.1);text-align:center">
                        <div style="color:#9ca3af;font-size:0.7rem;margin-bottom:0.2rem">Peak Volume</div>
                        <div style="color:#22c55e;font-size:1.1rem;font-weight:700">{peak_cm:,.0f} t</div>
                    </div>""", unsafe_allow_html=True)
                with vc3:
                    eff_color = '#22c55e' if avg_eff_val >= 100 else '#f59e0b' if avg_eff_val >= 90 else '#ef4444'
                    st.markdown(f"""
                    <div style="background:rgba(15,15,30,0.8);border-radius:10px;padding:0.8rem;
                    border:1px solid rgba(148,163,184,0.1);text-align:center">
                        <div style="color:#9ca3af;font-size:0.7rem;margin-bottom:0.2rem">Avg Efficiency</div>
                        <div style="color:{eff_color};font-size:1.1rem;font-weight:700">{avg_eff_val:.1f}%</div>
                    </div>""", unsafe_allow_html=True)
                with vc4:
                    bt_pct = below_target / total_p * 100 if total_p > 0 else 0
                    bt_color = '#22c55e' if bt_pct < 30 else '#f59e0b' if bt_pct < 60 else '#ef4444'
                    st.markdown(f"""
                    <div style="background:rgba(15,15,30,0.8);border-radius:10px;padding:0.8rem;
                    border:1px solid rgba(148,163,184,0.1);text-align:center">
                        <div style="color:#9ca3af;font-size:0.7rem;margin-bottom:0.2rem">Below Target</div>
                        <div style="color:{bt_color};font-size:1.1rem;font-weight:700">{below_target}/{total_p} ({bt_pct:.0f}%)</div>
                    </div>""", unsafe_allow_html=True)


                fig_combined = make_subplots(specs=[[{"secondary_y": True}]])


                cm_max_idx = flow['CM TWB'].idxmax()
                cm_min_idx = flow['CM TWB'].idxmin()


                fig_combined.add_trace(go.Scatter(
                    x=flow['Date'], y=flow['CM TWB Plot'],
                    mode='lines', name='CM TWB',
                    line=dict(color='#3B82F6', width=2),
                    fill='tozeroy',
                    fillcolor='rgba(59,130,246,0.12)',
                    connectgaps=True,
                    hovertemplate='<b>CM TWB</b><br>%{y:,.0f} ton<extra></extra>'
                ), secondary_y=False)


                fig_combined.add_trace(go.Scatter(
                    x=flow['Date'], y=flow['CH TWB Plot'],
                    mode='lines', name='CH TWB',
                    line=dict(color='#22C55E', width=2),
                    fill='tozeroy',
                    fillcolor='rgba(34,197,94,0.10)',
                    connectgaps=True,
                    hovertemplate='<b>CH TWB</b><br>%{y:,.0f} ton<extra></extra>'
                ), secondary_y=False)


                fig_combined.add_trace(go.Scatter(
                    x=flow['Date'], y=flow['Sales Plot'],
                    mode='lines', name='Sales',
                    line=dict(color='#A78BFA', width=2),
                    fill='tozeroy',
                    fillcolor='rgba(167,139,250,0.08)',
                    connectgaps=True,
                    hovertemplate='<b>Sales</b><br>%{y:,.0f} ton<extra></extra>'
                ), secondary_y=False)


                eff_colors = []
                for e in flow['CH Flow Ratio (%)'].values:
                    if e >= 100:
                        eff_colors.append('#22C55E')
                    elif e >= 90:
                        eff_colors.append('#F59E0B')
                    else:
                        eff_colors.append('#EF4444')


                fig_combined.add_trace(go.Scatter(
                    x=flow['Date'], y=flow['Eff Plot'],
                    mode='lines+markers', name='Efficiency (%)',
                    line=dict(color='#F59E0B', width=2.5, dash='dot'),
                    marker=dict(size=9, color=eff_colors, symbol='diamond',
                                line=dict(width=1.5, color='rgba(255,255,255,0.7)')),
                    connectgaps=True,
                    hovertemplate='<b>Efficiency</b><br>%{y:.1f}%<extra></extra>'
                ), secondary_y=True)


                fig_combined.add_hline(
                    y=avg_cm, line_dash='dot',
                    line_color='rgba(59,130,246,0.4)', line_width=1,
                    annotation_text=f'Avg CM {avg_cm:,.0f}',
                    annotation_position='top left',
                    annotation_font=dict(size=9, color='#60A5FA'))

                fig_combined.add_hline(
                    y=100, line_dash='dash',
                    line_color='rgba(245,158,11,0.4)', line_width=1,
                    secondary_y=True,
                    annotation_text='Target 100%',
                    annotation_position='top right',
                    annotation_font=dict(size=9, color='#F59E0B'))

                fig_combined.add_hline(
                    y=avg_eff_val, line_dash='dot',
                    line_color='rgba(245,158,11,0.25)', line_width=1,
                    secondary_y=True,
                    annotation_text=f'Avg Eff {avg_eff_val:.1f}%',
                    annotation_position='bottom right',
                    annotation_font=dict(size=8, color='#FBBF24'))


                fig_combined.add_annotation(
                    x=flow.loc[cm_max_idx, 'Date'],
                    y=flow.loc[cm_max_idx, 'CM TWB'],
                    text=f"▲ Peak {flow.loc[cm_max_idx, 'CM TWB']:,.0f} t",
                    showarrow=True, arrowhead=2,
                    arrowwidth=1.5, arrowcolor='#60A5FA',
                    ax=0, ay=-30,
                    font=dict(size=10, color='#FFF'),
                    bgcolor='rgba(59,130,246,0.75)', borderpad=4)

                fig_combined.add_annotation(
                    x=flow.loc[cm_min_idx, 'Date'],
                    y=flow.loc[cm_min_idx, 'CM TWB'],
                    text=f"▼ Low {flow.loc[cm_min_idx, 'CM TWB']:,.0f} t",
                    showarrow=True, arrowhead=2,
                    arrowwidth=1.5, arrowcolor='#EF4444',
                    ax=0, ay=30,
                    font=dict(size=10, color='#FFF'),
                    bgcolor='rgba(239,68,68,0.75)', borderpad=4)

                eff_max_idx = flow['CH Flow Ratio (%)'].idxmax()
                fig_combined.add_annotation(
                    x=flow.loc[eff_max_idx, 'Date'],
                    y=flow.loc[eff_max_idx, 'CH Flow Ratio (%)'],
                    text=f"⚡ {flow.loc[eff_max_idx, 'CH Flow Ratio (%)']:.1f}%",
                    showarrow=True, arrowhead=2,
                    arrowwidth=1, arrowcolor='#F59E0B',
                    ax=20, ay=-25, yref='y2',
                    font=dict(size=9, color='#FFF'),
                    bgcolor='rgba(245,158,11,0.75)', borderpad=3)


                fig_combined.update_layout(
                    height=500,
                    font=dict(family='Inter, sans-serif', color='#CBD5E1', size=12),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(
                        gridcolor='rgba(148,163,184,0.08)',
                        title='', tickformat='%d %b',
                        tickfont=dict(size=10, color='#94A3B8'),
                        showline=True,
                        linecolor='rgba(148,163,184,0.2)',
                        rangeselector=dict(
                            buttons=list([
                                dict(count=1, label="1M", step="month", stepmode="backward"),
                                dict(count=3, label="3M", step="month", stepmode="backward"),
                                dict(step="all", label="ALL"),
                            ]),
                            bgcolor='rgba(30,30,50,0.8)',
                            activecolor='rgba(59,130,246,0.4)',
                            font=dict(size=10, color='#CBD5E1'),
                        ),
                        rangeslider=dict(visible=True, thickness=0.06),
                        type="date",
                    ),
                    showlegend=True,
                    legend=dict(
                        orientation='h', yanchor='bottom', y=1.08,
                        xanchor='center', x=0.5,
                        font=dict(size=11, color='#E2E8F0'),
                        bgcolor='rgba(0,0,0,0)',
                    ),
                    hovermode='x unified',
                    margin=dict(l=65, r=65, t=70, b=10),
                )

                fig_combined.update_yaxes(
                    title=dict(text='Volume (ton)', font=dict(color='#94A3B8', size=12)),
                    tickformat=',.0f',
                    gridcolor='rgba(148,163,184,0.08)',
                    tickfont=dict(size=10, color='#94A3B8'),
                    showline=False, zeroline=False,
                    secondary_y=False,
                )

                fig_combined.update_yaxes(
                    title=dict(text='Efficiency (%)', font=dict(color='#F59E0B', size=12)),
                    tickformat='.0f',
                    ticksuffix='%',
                    gridcolor='rgba(0,0,0,0)',
                    tickfont=dict(size=10, color='#F59E0B'),
                    showline=False,
                    zeroline=False,
                    range=[0, max(130, flow['CH Flow Ratio (%)'].max() + 15)],
                    secondary_y=True,
                )
                st.plotly_chart(fig_combined, use_container_width=True, key='production_vol_eff')


                # ══════════════════════════════════════════════════════════
                # MATERIAL LOSS ANALYSIS — Enhanced Stacked + Cumulative
                # ══════════════════════════════════════════════════════════
                st.markdown("###  Material Loss Analysis")


                flow['Net Loss'] = flow['CM Loss'] + flow['CH Loss']
                flow['Cum Net Loss'] = flow['Net Loss'].cumsum()

                avg_net = flow['Net Loss'].mean()
                std_loss = flow['Net Loss'].std() if len(flow) > 1 else 0
                max_loss_val = flow['Net Loss'].min()
                cum_total = flow['Cum Net Loss'].iloc[-1]
                critical_count = (flow['Net Loss'].abs() > avg_cm * 0.02).sum()
                total_periods = len(flow)


                lc1, lc2, lc3, lc4 = st.columns(4)
                with lc1:
                    net_color = '#22c55e' if avg_net >= 0 else '#ef4444'
                    st.markdown(f"""
                    <div style="background:rgba(15,15,30,0.8);border-radius:10px;padding:0.8rem;
                    border:1px solid rgba(148,163,184,0.1);text-align:center">
                        <div style="color:#9ca3af;font-size:0.7rem;margin-bottom:0.2rem">Avg Net Loss</div>
                        <div style="color:{net_color};font-size:1.1rem;font-weight:700">{avg_net:,.0f} t</div>
                    </div>""", unsafe_allow_html=True)
                with lc2:
                    st.markdown(f"""
                    <div style="background:rgba(15,15,30,0.8);border-radius:10px;padding:0.8rem;
                    border:1px solid rgba(148,163,184,0.1);text-align:center">
                        <div style="color:#9ca3af;font-size:0.7rem;margin-bottom:0.2rem">Max Single Loss</div>
                        <div style="color:#ef4444;font-size:1.1rem;font-weight:700">{max_loss_val:,.0f} t</div>
                    </div>""", unsafe_allow_html=True)
                with lc3:
                    cum_color = '#22c55e' if cum_total >= 0 else '#ef4444'
                    st.markdown(f"""
                    <div style="background:rgba(15,15,30,0.8);border-radius:10px;padding:0.8rem;
                    border:1px solid rgba(148,163,184,0.1);text-align:center">
                        <div style="color:#9ca3af;font-size:0.7rem;margin-bottom:0.2rem">Cumulative Loss</div>
                        <div style="color:{cum_color};font-size:1.1rem;font-weight:700">{cum_total:,.0f} t</div>
                    </div>""", unsafe_allow_html=True)
                with lc4:
                    crit_pct = critical_count / total_periods * 100 if total_periods > 0 else 0
                    crit_color = '#22c55e' if crit_pct < 20 else '#f59e0b' if crit_pct < 40 else '#ef4444'
                    st.markdown(f"""
                    <div style="background:rgba(15,15,30,0.8);border-radius:10px;padding:0.8rem;
                    border:1px solid rgba(148,163,184,0.1);text-align:center">
                        <div style="color:#9ca3af;font-size:0.7rem;margin-bottom:0.2rem">Critical Periods</div>
                        <div style="color:{crit_color};font-size:1.1rem;font-weight:700">{critical_count}/{total_periods} ({crit_pct:.0f}%)</div>
                    </div>""", unsafe_allow_html=True)


                from plotly.subplots import make_subplots

                fig_loss = make_subplots(specs=[[{"secondary_y": True}]])

                y_min_val = flow['Net Loss'].min()
                y_max_val = flow['Net Loss'].max()
                y_pad = max(abs(y_min_val), abs(y_max_val)) * 0.15
                for y0, y1, color in [
                    (avg_net - 1 * std_loss, avg_net + 1 * std_loss, 'rgba(34,197,94,0.06)'),
                    (avg_net - 2 * std_loss, avg_net - 1 * std_loss, 'rgba(245,158,11,0.06)'),
                    (avg_net + 1 * std_loss, avg_net + 2 * std_loss, 'rgba(245,158,11,0.06)'),
                ]:
                    fig_loss.add_shape(
                        type="rect", xref="paper", yref="y",
                        x0=0, x1=1, y0=y0, y1=y1,
                        fillcolor=color, line_width=0, layer="below"
                    )

                fig_loss.add_trace(go.Bar(
                    x=flow['Date'], y=flow['CH Loss'],
                    name='CH Loss',
                    marker=dict(color='rgba(239,68,68,0.7)',
                           line=dict(width=0.5, color='rgba(239,68,68,0.9)')),
                    hovertemplate='<b>CH Loss</b><br>%{x|%d %b %Y}<br>%{y:,.0f} ton<extra></extra>',
                ), secondary_y=False)

                fig_loss.add_trace(go.Bar(
                    x=flow['Date'], y=flow['CM Loss'],
                    name='CM Loss',
                    marker=dict(color='rgba(234,179,8,0.7)',
                           line=dict(width=0.5, color='rgba(234,179,8,0.9)')),
                    hovertemplate='<b>CM Loss</b><br>%{x|%d %b %Y}<br>%{y:,.0f} ton<extra></extra>',
                ), secondary_y=False)

                net_colors = []
                for v in flow['Net Loss'].values:
                    abs_v = abs(v)
                    if abs_v <= abs(avg_net) + 1 * std_loss:
                        net_colors.append('#22C55E')
                    elif abs_v <= abs(avg_net) + 2 * std_loss:
                        net_colors.append('#F59E0B')
                    else:
                        net_colors.append('#EF4444')

                fig_loss.add_trace(go.Scatter(
                    x=flow['Date'], y=flow['Net Loss'],
                    mode='lines+markers',
                    name='Net Loss',
                    line=dict(color='rgba(255,255,255,0.4)', width=1.5),
                    marker=dict(size=8, color=net_colors,
                            line=dict(width=1.5, color='rgba(255,255,255,0.6)')),
                    connectgaps=True,
                    hovertemplate='<b>Net Loss</b><br>%{x|%d %b %Y}<br>%{y:,.0f} ton<extra></extra>',
                ), secondary_y=False)

                fig_loss.add_trace(go.Scatter(
                    x=flow['Date'], y=flow['Cum Net Loss'],
                    mode='lines',
                    name='Cumulative Loss',
                    line=dict(color='#06b6d4', width=2.5, dash='dot'),
                    fill='tozeroy',
                    fillcolor='rgba(6,182,212,0.06)',
                    connectgaps=True,
                    hovertemplate='<b>Cumulative</b><br>%{y:,.0f} ton<extra></extra>',
                ), secondary_y=True)

                fig_loss.add_hline(y=0, line_color='rgba(148,163,184,0.5)', line_width=1.5)

                fig_loss.add_hline(
                    y=avg_net, line_dash='dot',
                    line_color='rgba(239,68,68,0.4)', line_width=1,
                    annotation_text=f'Avg {avg_net:,.0f} t',
                    annotation_position='bottom right',
                    annotation_font=dict(size=9, color='#F87171'))

                fig_loss.add_hline(
                    y=avg_net + 2 * std_loss,
                    line_color='#eab308', line_width=1, line_dash='dash',
                    annotation_text='+2σ', annotation_position='top right',
                    annotation_font=dict(size=8, color='#eab308'))

                fig_loss.add_hline(
                    y=avg_net - 2 * std_loss,
                    line_color='#eab308', line_width=1, line_dash='dash',
                    annotation_text='-2σ', annotation_position='bottom right',
                    annotation_font=dict(size=8, color='#eab308'))

                worst_idx = flow['Net Loss'].idxmin()
                fig_loss.add_annotation(
                    x=flow.loc[worst_idx, 'Date'],
                    y=flow.loc[worst_idx, 'Net Loss'],
                    text=f"▼ Worst: {flow.loc[worst_idx, 'Net Loss']:,.0f} t",
                    showarrow=True, arrowhead=2,
                    arrowwidth=1.5, arrowcolor='#EF4444',
                    ax=40, ay=-30,
                    font=dict(size=10, color='#FFF'),
                    bgcolor='rgba(239,68,68,0.85)', borderpad=4)

                best_idx = flow['Net Loss'].abs().idxmin()
                fig_loss.add_annotation(
                    x=flow.loc[best_idx, 'Date'],
                    y=flow.loc[best_idx, 'Net Loss'],
                    text=f"✓ Best: {flow.loc[best_idx, 'Net Loss']:,.0f} t",
                    showarrow=True, arrowhead=2,
                    arrowwidth=1.5, arrowcolor='#22C55E',
                    ax=-40, ay=-25,
                    font=dict(size=10, color='#FFF'),
                    bgcolor='rgba(34,197,94,0.85)', borderpad=4)

                fig_loss.update_layout(
                    height=460,
                    barmode='relative',
                    font=dict(family='Inter, sans-serif', color='#CBD5E1', size=11),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(
                        gridcolor='rgba(148,163,184,0.08)',
                        title='', tickformat='%d %b',
                        tickfont=dict(size=10, color='#94A3B8'),
                        showline=True,
                        linecolor='rgba(148,163,184,0.2)',
                        rangeselector=dict(
                            buttons=list([
                                dict(count=1, label="1M", step="month", stepmode="backward"),
                                dict(count=3, label="3M", step="month", stepmode="backward"),
                                dict(step="all", label="ALL"),
                            ]),
                            bgcolor='rgba(30,30,50,0.8)',
                            activecolor='rgba(59,130,246,0.4)',
                            font=dict(size=10, color='#CBD5E1'),
                        ),
                        rangeslider=dict(visible=True, thickness=0.06),
                        type="date",
                    ),
                    showlegend=True,
                    legend=dict(
                        orientation='h', yanchor='bottom', y=1.08,
                        xanchor='center', x=0.5,
                        font=dict(size=10, color='#E2E8F0'),
                        bgcolor='rgba(0,0,0,0)',
                    ),
                    hovermode='x unified',
                    margin=dict(l=60, r=60, t=70, b=10),
                )

                fig_loss.update_yaxes(
                    title=dict(text='Loss (ton)', font=dict(color='#94A3B8', size=11)),
                    tickformat=',.0f',
                    gridcolor='rgba(148,163,184,0.08)',
                    tickfont=dict(size=10, color='#94A3B8'),
                    showline=False, zeroline=False,
                    secondary_y=False,
                )

                fig_loss.update_yaxes(
                    title=dict(text='Cumulative (ton)', font=dict(color='#06b6d4', size=11)),
                    tickformat=',.0f',
                    gridcolor='rgba(0,0,0,0)',
                    tickfont=dict(size=10, color='#06b6d4'),
                    showline=False, zeroline=False,
                    secondary_y=True,
                )

                st.plotly_chart(fig_loss, use_container_width=True, key='loss_analysis')

    with tab2:
        # ══════════════════════════════════════════════════════════
        # TAB 2: OVERBURDEN ANALYSIS — v6 (Alert di bawah KPI)
        # ══════════════════════════════════════════════════════════
        st.markdown("""
        <div class="section-header">
            <h3 class="section-title">Overburden Analysis</h3>
        </div>
        """, unsafe_allow_html=True)

        # ── FILTER PERIODE ──
        if 'Bulan' in df_ob.columns and len(df_ob) > 0:
            bulan_list = df_ob['Bulan'].unique().tolist()
            bulan_sorted = sort_months_chronologically(bulan_list)

            st.markdown("### Filter Periode")
            if len(bulan_sorted) > 1:
                total = len(bulan_sorted)
                selected_range = st.slider(
                    "Pilih Range Bulan:",
                    min_value=1, max_value=total,
                    value=(1, total),
                    key='ob_month_slider_v2'
                )
                start_idx = selected_range[0] - 1
                end_idx = selected_range[1]
                selected_months = bulan_sorted[start_idx:end_idx]
                df_ob_filtered = df_ob[df_ob['Bulan'].isin(selected_months)].copy()
                month_order = {month: i for i, month in enumerate(bulan_sorted)}
                df_ob_filtered['Sort_Order'] = df_ob_filtered['Bulan'].map(month_order)
                df_ob_filtered = df_ob_filtered.sort_values('Sort_Order')
                if selected_months:
                    st.info(f"**Bulan yang dipilih:** {len(selected_months)} bulan ({selected_months[0]} - {selected_months[-1]})")
                df_ob_filtered = df_ob_filtered.drop(columns=['Sort_Order'])
            else:
                df_ob_filtered = df_ob.copy()
                st.info(f"Data hanya untuk 1 bulan: {bulan_sorted[0]}")
        else:
            df_ob_filtered = df_ob.copy()
            st.warning("Kolom 'Bulan' tidak ditemukan dalam data")

        # ── KPI CARDS (DULU, sebelum alert) ──
        n_periods = len(df_ob_filtered)
        total_tc = df_ob_filtered['TC'].sum()
        total_js = df_ob_filtered['JS'].sum()
        total_dev_val = df_ob_filtered['Dev_Absolut'].sum()
        avg_dev_pct = df_ob_filtered['Dev_Relatif_Pct'].abs().mean()
        dev_color = '#4ade80' if avg_dev_pct <= 2 else '#fbbf24' if avg_dev_pct <= 3 else '#f87171'

        kpi_data = [
            ("Periods", f"{n_periods}", "#60a5fa"),
            ("Total TC", format_large(total_tc), "#60a5fa"),
            ("Total JS", format_large(total_js), "#4ade80"),
            ("Total Deviation", format_large(total_dev_val), "#fbbf24"),
            ("Avg |Dev%|", f"{format_number(avg_dev_pct, 2)}%", dev_color),
        ]

        cols_kpi = st.columns(5, gap="medium")
        for i, (label, value, color) in enumerate(kpi_data):
            with cols_kpi[i]:
                st.markdown(f"""
                <div style="background:linear-gradient(145deg,rgba(22,28,45,0.9),rgba(15,20,35,0.95));
                    border:1px solid rgba(148,163,184,0.08);border-radius:12px;
                    padding:14px 12px;text-align:center;position:relative;overflow:hidden;">
                    <div style="position:absolute;top:0;left:0;right:0;height:2px;
                        background:{color};border-radius:12px 12px 0 0;"></div>
                    <div style="font-size:0.65rem;font-weight:600;text-transform:uppercase;
                        letter-spacing:0.08em;color:#94a3b8;margin-bottom:4px;">{label}</div>
                    <div style="font-size:1.5rem;font-weight:800;color:{color};line-height:1.1;">{value}</div>
                </div>
                """, unsafe_allow_html=True)

        # ── ALERT BANNER (sekarang DI BAWAH KPI, hanya jika ada Critical) ──
        critical_rows = df_ob_filtered[df_ob_filtered['Status'] == 'Critical']
        if len(critical_rows) > 0:
            alert_chips = " ".join([
                f'<span style="background:rgba(239,68,68,0.15);color:#f87171;'
                f'padding:4px 12px;border-radius:8px;font-size:0.78rem;font-weight:600;">'
                f'{row["Bulan"]}: {row["Dev_Relatif_Pct"]:.1f}%</span>'
                for _, row in critical_rows.iterrows()
            ])
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,rgba(239,68,68,0.08),rgba(15,23,42,0.9));
                border:1px solid rgba(239,68,68,0.2);border-radius:12px;
                padding:14px 20px;margin-top:12px;">
                <div style="color:#f87171;font-size:0.72rem;font-weight:700;
                    text-transform:uppercase;letter-spacing:0.1em;margin-bottom:8px;">
                    Critical Alert — {len(critical_rows)} Period(s) Detected</div>
                <div style="display:flex;flex-wrap:wrap;gap:8px;">{alert_chips}</div>
            </div>
            """, unsafe_allow_html=True)

        # ══ GRAFIK VOLUME & DEVIATION ══
        st.markdown("""
        <div class="section-header">
            <h3 class="section-title">Volume & Deviation Analysis</h3>
        </div>
        """, unsafe_allow_html=True)

        col_left, col_right = st.columns(2)

        with col_left:
            fig_comp = go.Figure()
            fig_comp.add_trace(go.Scatter(
                x=df_ob_filtered['Bulan'], y=df_ob_filtered['TC'],
                mode='lines+markers', name='TC (Target)',
                line=dict(color='#3B82F6', width=2.5),
                marker=dict(size=7, line=dict(width=1.5, color='rgba(255,255,255,0.6)')),
                hovertemplate='<b>TC</b><br>%{y:,.0f}<extra></extra>'
            ))
            fig_comp.add_trace(go.Scatter(
                x=df_ob_filtered['Bulan'], y=df_ob_filtered['JS'],
                mode='lines+markers', name='JS (Survey)',
                line=dict(color='#22C55E', width=2.5),
                marker=dict(size=7, line=dict(width=1.5, color='rgba(255,255,255,0.6)')),
                fill='tonexty', fillcolor='rgba(59,130,246,0.10)',
                hovertemplate='<b>JS</b><br>%{y:,.0f}<extra></extra>'
            ))

            avg_tc_val = df_ob_filtered['TC'].mean()
            fig_comp.add_hline(
                y=avg_tc_val, line_dash='dot',
                line_color='rgba(59,130,246,0.35)', line_width=1,
                annotation_text=f'Avg TC {avg_tc_val:,.0f}',
                annotation_position='top left',
                annotation_font=dict(size=9, color='#60A5FA')
            )

            if len(df_ob_filtered) > 1:
                pk_idx = df_ob_filtered['TC'].idxmax()
                lo_idx = df_ob_filtered['TC'].idxmin()
                fig_comp.add_annotation(
                    x=df_ob_filtered.loc[pk_idx, 'Bulan'],
                    y=df_ob_filtered.loc[pk_idx, 'TC'],
                    text=f"Peak {df_ob_filtered.loc[pk_idx, 'TC']:,.0f}",
                    showarrow=True, arrowhead=2, arrowwidth=1.5,
                    arrowcolor='#60A5FA', ax=0, ay=-30,
                    font=dict(size=10, color='#FFF'),
                    bgcolor='rgba(59,130,246,0.75)', borderpad=4
                )
                fig_comp.add_annotation(
                    x=df_ob_filtered.loc[lo_idx, 'Bulan'],
                    y=df_ob_filtered.loc[lo_idx, 'TC'],
                    text=f"Low {df_ob_filtered.loc[lo_idx, 'TC']:,.0f}",
                    showarrow=True, arrowhead=2, arrowwidth=1.5,
                    arrowcolor='#EF4444', ax=0, ay=30,
                    font=dict(size=10, color='#FFF'),
                    bgcolor='rgba(239,68,68,0.75)', borderpad=4
                )

            fig_comp.update_layout(
                height=420,
                font=dict(family='Inter, sans-serif', color='#CBD5E1', size=11),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                xaxis=dict(
                    gridcolor='rgba(148,163,184,0.08)', title='',
                    tickfont=dict(size=10, color='#94A3B8'),
                    showline=True, linecolor='rgba(148,163,184,0.2)'
                ),
                yaxis=dict(
                    title=dict(text='Volume', font=dict(color='#94A3B8', size=12)),
                    tickformat=',.0f', gridcolor='rgba(148,163,184,0.08)',
                    tickfont=dict(size=10, color='#94A3B8'),
                    showline=False, zeroline=False
                ),
                showlegend=True,
                legend=dict(
                    orientation='h', yanchor='bottom', y=1.05,
                    xanchor='center', x=0.5,
                    font=dict(size=10, color='#E2E8F0'),
                    bgcolor='rgba(0,0,0,0)'
                ),
                hovermode='x unified',
                margin=dict(l=65, r=20, t=50, b=40)
            )
            st.plotly_chart(fig_comp, use_container_width=True, key='ob_comp')

        with col_right:
            colors_bar = [get_status_color(s) for s in df_ob_filtered['Status']]
            fig_dev = go.Figure()

            fig_dev.add_shape(type='rect', xref='paper', yref='y',
                x0=0, x1=1, y0=-2, y1=2,
                fillcolor='rgba(34,197,94,0.04)', line_width=0, layer='below')
            fig_dev.add_shape(type='rect', xref='paper', yref='y',
                x0=0, x1=1, y0=2, y1=3,
                fillcolor='rgba(245,158,11,0.04)', line_width=0, layer='below')
            fig_dev.add_shape(type='rect', xref='paper', yref='y',
                x0=0, x1=1, y0=-3, y1=-2,
                fillcolor='rgba(245,158,11,0.04)', line_width=0, layer='below')

            fig_dev.add_trace(go.Bar(
                x=df_ob_filtered['Bulan'], y=df_ob_filtered['Dev_Relatif_Pct'],
                marker_color=colors_bar, showlegend=False,
                text=df_ob_filtered['Dev_Relatif_Pct'].apply(lambda x: f"{x:.1f}%"),
                textposition='outside',
                textfont=dict(size=10, color='#e2e8f0'),
                hovertemplate='<b>%{x}</b><br>Deviation: %{y:.2f}%<extra></extra>'
            ))

            fig_dev.add_hline(y=0, line_color='rgba(148,163,184,0.5)', line_width=1.5)
            fig_dev.add_hline(y=2, line_color='#22c55e', line_width=1, line_dash='dash')
            fig_dev.add_hline(y=-2, line_color='#22c55e', line_width=1, line_dash='dash')
            fig_dev.add_hline(y=3, line_color='#ef4444', line_width=1, line_dash='dash')
            fig_dev.add_hline(y=-3, line_color='#ef4444', line_width=1, line_dash='dash')

            fig_dev.add_annotation(xref='paper', yref='y', x=1.02, y=2,
                text='2%', showarrow=False, font=dict(size=8, color='#22c55e'))
            fig_dev.add_annotation(xref='paper', yref='y', x=1.02, y=3,
                text='3%', showarrow=False, font=dict(size=8, color='#ef4444'))

            if len(df_ob_filtered) > 0:
                w_idx = df_ob_filtered['Dev_Relatif_Pct'].abs().idxmax()
                w_val = df_ob_filtered.loc[w_idx, 'Dev_Relatif_Pct']
                fig_dev.add_annotation(
                    x=df_ob_filtered.loc[w_idx, 'Bulan'], y=w_val,
                    text=f"Worst {w_val:.1f}%",
                    showarrow=True, arrowhead=2, arrowwidth=1.5,
                    arrowcolor='#EF4444',
                    ax=35, ay=-25 if w_val > 0 else 25,
                    font=dict(size=10, color='#FFF'),
                    bgcolor='rgba(239,68,68,0.85)', borderpad=4
                )

            fig_dev.update_layout(
                height=420,
                font=dict(family='Inter, sans-serif', color='#CBD5E1', size=11),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                xaxis=dict(
                    gridcolor='rgba(148,163,184,0.08)', title='',
                    tickfont=dict(size=10, color='#94A3B8'),
                    showline=True, linecolor='rgba(148,163,184,0.2)'
                ),
                yaxis=dict(
                    title=dict(text='Deviation %', font=dict(color='#94A3B8', size=12)),
                    gridcolor='rgba(148,163,184,0.08)',
                    tickfont=dict(size=10, color='#94A3B8'),
                    showline=False, zeroline=False
                ),
                hovermode='x unified',
                margin=dict(l=60, r=40, t=40, b=40)
            )
            st.plotly_chart(fig_dev, use_container_width=True, key='ob_dev')

        # ══ DEVIATION HEATMAP (langsung tanpa header) ══
        if len(df_ob_filtered) > 0:
            hm_cells = ""
            for _, r in df_ob_filtered.iterrows():
                ad = abs(r['Dev_Relatif_Pct'])
                if ad <= 2:
                    bg, tc = "rgba(34,197,94,0.2)", "#4ade80"
                elif ad <= 3:
                    bg, tc = "rgba(245,158,11,0.2)", "#fbbf24"
                else:
                    bg, tc = "rgba(239,68,68,0.2)", "#f87171"
                hm_cells += f"""<div style="flex:1;background:{bg};border-radius:8px;
                    padding:10px 4px;text-align:center;min-width:50px;">
                    <div style="font-size:0.6rem;color:#94a3b8;margin-bottom:3px;">{r['Bulan'][:3]}</div>
                    <div style="font-size:0.82rem;font-weight:700;color:{tc};">{r['Dev_Relatif_Pct']:.1f}%</div>
                </div>"""
            st.markdown(f"""<div style="display:flex;gap:5px;margin:0 0 20px 0;">{hm_cells}</div>""",
                unsafe_allow_html=True)

        # ══ OB DATA TABLE — styled HTML (replace st.dataframe) ══
        st.markdown("""
        <div class="section-header" style="margin-top:2.5rem;">
            <h3 class="section-title">OB Data Table</h3>
        </div>
        """, unsafe_allow_html=True)

        if not df_ob_filtered.empty:
            ob_cols = ['Bulan', 'TC', 'JS', 'Dev_Absolut', 'Dev_Relatif_Pct', 'Status']
            ob_header = ''.join(
                f'<th style="padding:10px 10px;text-align:center;color:#e2e8f0;font-weight:700;'
                f'font-size:0.68rem;text-transform:uppercase;letter-spacing:0.06em;'
                f'position:sticky;top:0;background:rgba(22,101,52,0.95);z-index:1;">{c}</th>'
                for c in ob_cols
            )
            ob_rows = ""
            for _, row in df_ob_filtered.iterrows():
                status = row['Status']
                if status == 'Critical':
                    row_bg, bl, sc, sbg = 'rgba(239,68,68,0.08)', '3px solid #ef4444', '#f87171', 'rgba(239,68,68,0.15)'
                elif status == 'Caution':
                    row_bg, bl, sc, sbg = 'rgba(245,158,11,0.08)', '3px solid #f59e0b', '#fbbf24', 'rgba(245,158,11,0.15)'
                else:
                    row_bg, bl, sc, sbg = 'rgba(34,197,94,0.04)', '3px solid #22c55e', '#4ade80', 'rgba(34,197,94,0.15)'

                dev_val = row['Dev_Relatif_Pct']
                dev_color = '#f87171' if abs(dev_val) > 3 else '#fbbf24' if abs(dev_val) > 2 else '#4ade80'

                ob_rows += f"""
                <tr style="background:{row_bg};border-left:{bl};border-bottom:1px solid rgba(148,163,184,0.06);">
                    <td style="padding:9px 10px;text-align:center;color:#e2e8f0;font-weight:600;font-size:0.8rem;">{row['Bulan']}</td>
                    <td style="padding:9px 10px;text-align:center;color:#cbd5e1;font-size:0.8rem;">{row['TC']:,.2f}</td>
                    <td style="padding:9px 10px;text-align:center;color:#cbd5e1;font-size:0.8rem;">{row['JS']:,.2f}</td>
                    <td style="padding:9px 10px;text-align:center;color:#cbd5e1;font-size:0.8rem;">{row['Dev_Absolut']:,.2f}</td>
                    <td style="padding:9px 10px;text-align:center;color:{dev_color};font-weight:700;font-size:0.8rem;">{dev_val:+.2f}%</td>
                    <td style="padding:9px 10px;text-align:center;">
                        <span style="background:{sbg};color:{sc};padding:3px 10px;border-radius:20px;font-size:0.7rem;font-weight:700;">{status}</span>
                    </td>
                </tr>"""

            st.markdown(f"""
            <div style="background:linear-gradient(135deg,rgba(20,20,40,0.6),rgba(15,23,42,0.8));
                border:1px solid rgba(148,163,184,0.1);border-radius:10px;overflow:hidden;
                max-height:400px;overflow-y:auto;">
            <table style="width:100%;border-collapse:collapse;">
            <thead><tr style="background:rgba(22,101,52,0.95);">{ob_header}</tr></thead>
            <tbody>{ob_rows}</tbody>
            </table></div>
            """, unsafe_allow_html=True)
        else:
            st.info("Tidak ada data untuk ditampilkan")

    with tab3:
        # ══════════════════════════════════════════════════════════
        # TAB 3: CPP33 & PORT ANALYSIS — v6 (Samakan dengan Tab 2)
        # Perubahan: 1) Status Distribution dihilangkan
        #            2) Insight dihilangkan
        #            3) Alert di bawah KPI
        #            4) Data table diberi warna row sesuai status
        # ══════════════════════════════════════════════════════════
        st.markdown("""
        <div class="section-header">
            <h3 class="section-title">CPP33 & Port Analysis</h3>
        </div>
        """, unsafe_allow_html=True)

        df_valid = df_ch_cm.dropna(subset=['CH_WB', 'CM_WB'])

        if len(df_valid) > 0:
            df_valid_copy = df_valid.copy()
            df_valid_copy['YearMonth'] = df_valid_copy['Date'].dt.strftime('%b-%Y')
            available_months = sorted(df_valid_copy['YearMonth'].unique(),
                                      key=lambda x: pd.to_datetime(x, format='%b-%Y'))

            st.markdown("### Filter Periode")
            if len(available_months) > 1:
                total_months = len(available_months)
                selected_range_cpp = st.slider(
                    "Pilih Range Periode",
                    min_value=1, max_value=total_months,
                    value=(1, total_months),
                    key='cpp_port_slider')
                start_idx = selected_range_cpp[0] - 1
                end_idx = selected_range_cpp[1]
                selected_cpp_months = available_months[start_idx:end_idx]
                df_valid_filtered = df_valid_copy[df_valid_copy['YearMonth'].isin(selected_cpp_months)].copy()
                if selected_cpp_months:
                    st.info(f"Periode: **{len(selected_cpp_months)} bulan** ({selected_cpp_months[0]} - {selected_cpp_months[-1]})")
            else:
                df_valid_filtered = df_valid_copy.copy()
                st.info(f"Data hanya untuk 1 bulan: {available_months[0]}")
            df_valid_filtered = df_valid_filtered.drop(columns=['YearMonth'], errors='ignore')

            # ── KPI CARDS (dulu, sebelum alert) ──
            ch_avg = df_valid_filtered['Dev_CH_Relatif_Pct'].abs().mean()
            cm_avg = df_valid_filtered['Dev_CM_Relatif_Pct'].abs().mean()
            twb_ch = df_valid_filtered['TWB_CH'].mean()
            twb_cm = df_valid_filtered['TWB_CM'].mean()
            ch_color = '#4ade80' if ch_avg <= 2 else '#fbbf24' if ch_avg <= 3 else '#f87171'
            cm_color = '#4ade80' if cm_avg <= 2 else '#fbbf24' if cm_avg <= 3 else '#f87171'

            kpi_cpp = [
                ("CH Avg |Dev%|", f"{format_number(ch_avg, 2)}%", ch_color),
                ("CM Avg |Dev%|", f"{format_number(cm_avg, 2)}%", cm_color),
                ("Avg TWB CH", format_large(twb_ch), "#60a5fa"),
                ("Avg TWB CM", format_large(twb_cm), "#eab308"),
            ]
            cols_cpp = st.columns(4, gap="medium")
            for i, (label, value, color) in enumerate(kpi_cpp):
                with cols_cpp[i]:
                    st.markdown(f"""
                    <div style="background:linear-gradient(145deg,rgba(22,28,45,0.9),rgba(15,20,35,0.95));
                        border:1px solid rgba(148,163,184,0.08);border-radius:12px;
                        padding:14px 12px;text-align:center;position:relative;overflow:hidden;">
                        <div style="position:absolute;top:0;left:0;right:0;height:2px;
                            background:{color};border-radius:12px 12px 0 0;"></div>
                        <div style="font-size:0.65rem;font-weight:600;text-transform:uppercase;
                            letter-spacing:0.08em;color:#94a3b8;margin-bottom:4px;">{label}</div>
                        <div style="font-size:1.5rem;font-weight:800;color:{color};line-height:1.1;">{value}</div>
                    </div>
                    """, unsafe_allow_html=True)

            # ── ALERT BANNER (di bawah KPI, hanya jika ada Critical) ──
            crit_ch = df_valid_filtered[df_valid_filtered['Status_CH'] == 'Critical']
            crit_cm = df_valid_filtered[df_valid_filtered['Status_CM'] == 'Critical']
            n_crit = len(crit_ch) + len(crit_cm)

            if n_crit > 0:
                parts = []
                if len(crit_ch) > 0:
                    parts.append(f'<span style="background:rgba(59,130,246,0.15);color:#60a5fa;'
                        f'padding:4px 12px;border-radius:8px;font-size:0.78rem;font-weight:600;">'
                        f'CH: {len(crit_ch)} period(s)</span>')
                if len(crit_cm) > 0:
                    parts.append(f'<span style="background:rgba(245,158,11,0.15);color:#fbbf24;'
                        f'padding:4px 12px;border-radius:8px;font-size:0.78rem;font-weight:600;">'
                        f'CM: {len(crit_cm)} period(s)</span>')
                st.markdown(f"""
                <div style="background:linear-gradient(135deg,rgba(239,68,68,0.08),rgba(15,23,42,0.9));
                    border:1px solid rgba(239,68,68,0.2);border-radius:12px;
                    padding:14px 20px;margin-top:12px;">
                    <div style="color:#f87171;font-size:0.72rem;font-weight:700;
                        text-transform:uppercase;letter-spacing:0.1em;margin-bottom:8px;">
                        Critical Alert — {n_crit} Period(s) Detected</div>
                    <div style="display:flex;flex-wrap:wrap;gap:8px;">{" ".join(parts)}</div>
                </div>
                """, unsafe_allow_html=True)

            # ── CHARTS (langsung setelah alert, tanpa status distribution & insight) ──
            col_left, col_right = st.columns([1, 1])

            with col_left:
                st.markdown("""
                <div class="section-header">
                    <h3 class="section-title">Total Weight Bridge vs Weight Bridge Trend</h3>
                </div>
                """, unsafe_allow_html=True)

                # Coal Hauling
                fig_ch = go.Figure()
                fig_ch.add_trace(go.Scatter(
                    x=df_valid_filtered['Date'], y=df_valid_filtered['TWB_CH'],
                    name='TWB (Actual)',
                    line=dict(color='#3b82f6', width=2.5),
                    mode='lines+markers',
                    marker=dict(size=6, line=dict(width=1, color='rgba(255,255,255,0.5)')),
                    hovertemplate='<b>TWB CH</b><br>%{y:,.0f}<extra></extra>'
                ))
                fig_ch.add_trace(go.Scatter(
                    x=df_valid_filtered['Date'], y=df_valid_filtered['CH_WB'],
                    name='WB (Target)',
                    line=dict(color='#22c55e', width=2.5, dash='dash'),
                    mode='lines+markers', marker=dict(size=6),
                    fill='tonexty', fillcolor='rgba(59,130,246,0.08)',
                    hovertemplate='<b>WB Target</b><br>%{y:,.0f}<extra></extra>'
                ))
                if len(df_valid_filtered) > 1:
                    ch_pk = df_valid_filtered['TWB_CH'].idxmax()
                    fig_ch.add_annotation(
                        x=df_valid_filtered.loc[ch_pk, 'Date'],
                        y=df_valid_filtered.loc[ch_pk, 'TWB_CH'],
                        text=f"Peak {df_valid_filtered.loc[ch_pk, 'TWB_CH']:,.0f}",
                        showarrow=True, arrowhead=2, arrowwidth=1.5,
                        arrowcolor='#60a5fa', ax=0, ay=-28,
                        font=dict(size=9, color='#FFF'),
                        bgcolor='rgba(59,130,246,0.75)', borderpad=3)
                    ch_mean = df_valid_filtered['TWB_CH'].mean()
                    fig_ch.add_hline(y=ch_mean, line_dash='dot',
                        line_color='rgba(59,130,246,0.3)', line_width=1,
                        annotation_text=f'Avg {ch_mean:,.0f}',
                        annotation_position='top left',
                        annotation_font=dict(size=9, color='#60a5fa'))
                fig_ch.update_layout(
                    title=dict(text='Coal Hauling: TWB vs WB (ton)', font=dict(size=13, color='#ffffff')),
                    height=320,
                    font=dict(family='Inter, sans-serif', color='#e5e7eb'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(gridcolor='rgba(255,255,255,0.08)', title=''),
                    yaxis=dict(gridcolor='rgba(255,255,255,0.08)', title='Ton', tickformat=',.0f'),
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                    hovermode='x unified',
                    margin=dict(l=60, r=20, t=50, b=40)
                )
                st.plotly_chart(fig_ch, use_container_width=True, key='cpp_ch_comp')

                # Coal Mining
                fig_cm = go.Figure()
                fig_cm.add_trace(go.Scatter(
                    x=df_valid_filtered['Date'], y=df_valid_filtered['TWB_CM'],
                    name='TWB (Actual)',
                    line=dict(color='#eab308', width=2.5),
                    mode='lines+markers',
                    marker=dict(size=6, line=dict(width=1, color='rgba(255,255,255,0.5)')),
                    hovertemplate='<b>TWB CM</b><br>%{y:,.0f}<extra></extra>'
                ))
                fig_cm.add_trace(go.Scatter(
                    x=df_valid_filtered['Date'], y=df_valid_filtered['CM_WB'],
                    name='WB (Target)',
                    line=dict(color='#ef4444', width=2.5, dash='dash'),
                    mode='lines+markers', marker=dict(size=6),
                    fill='tonexty', fillcolor='rgba(234,179,8,0.08)',
                    hovertemplate='<b>WB Target</b><br>%{y:,.0f}<extra></extra>'
                ))
                if len(df_valid_filtered) > 1:
                    cm_pk = df_valid_filtered['TWB_CM'].idxmax()
                    fig_cm.add_annotation(
                        x=df_valid_filtered.loc[cm_pk, 'Date'],
                        y=df_valid_filtered.loc[cm_pk, 'TWB_CM'],
                        text=f"Peak {df_valid_filtered.loc[cm_pk, 'TWB_CM']:,.0f}",
                        showarrow=True, arrowhead=2, arrowwidth=1.5,
                        arrowcolor='#eab308', ax=0, ay=-28,
                        font=dict(size=9, color='#FFF'),
                        bgcolor='rgba(234,179,8,0.75)', borderpad=3)
                    cm_mean = df_valid_filtered['TWB_CM'].mean()
                    fig_cm.add_hline(y=cm_mean, line_dash='dot',
                        line_color='rgba(234,179,8,0.3)', line_width=1,
                        annotation_text=f'Avg {cm_mean:,.0f}',
                        annotation_position='top left',
                        annotation_font=dict(size=9, color='#eab308'))
                fig_cm.update_layout(
                    title=dict(text='Coal Mining: TWB vs WB (ton)', font=dict(size=13, color='#ffffff')),
                    height=320,
                    font=dict(family='Inter, sans-serif', color='#e5e7eb'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(gridcolor='rgba(255,255,255,0.08)', title=''),
                    yaxis=dict(gridcolor='rgba(255,255,255,0.08)', title='Ton', tickformat=',.0f'),
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                    hovermode='x unified',
                    margin=dict(l=60, r=20, t=50, b=40)
                )
                st.plotly_chart(fig_cm, use_container_width=True, key='cpp_cm_comp')

            with col_right:
                st.markdown("""
                <div class="section-header">
                    <h3 class="section-title">Deviation Pattern</h3>
                </div>
                """, unsafe_allow_html=True)

                # CH Deviation
                c_ch = [get_status_color(s) for s in df_valid_filtered['Status_CH']]
                fig_chd = go.Figure()
                fig_chd.add_shape(type='rect', xref='paper', yref='y',
                    x0=0, x1=1, y0=-2, y1=2,
                    fillcolor='rgba(34,197,94,0.04)', line_width=0, layer='below')
                fig_chd.add_trace(go.Bar(
                    x=df_valid_filtered['Date'],
                    y=df_valid_filtered['Dev_CH_Relatif_Pct'],
                    marker_color=c_ch, showlegend=False,
                    hovertemplate='<b>CH Dev</b><br>%{y:.2f}%<extra></extra>'
                ))
                fig_chd.add_hline(y=0, line_color="#9ca3af", line_width=2)
                fig_chd.add_hline(y=2, line_color="#22c55e", line_width=1, line_dash="dash",
                    annotation_text="2%", annotation_position="top right",
                    annotation_font=dict(size=8, color='#22c55e'))
                fig_chd.add_hline(y=-2, line_color="#22c55e", line_width=1, line_dash="dash")
                fig_chd.add_hline(y=3, line_color="#eab308", line_width=1, line_dash="dash",
                    annotation_text="3%", annotation_position="top right",
                    annotation_font=dict(size=8, color='#eab308'))
                fig_chd.add_hline(y=-3, line_color="#eab308", line_width=1, line_dash="dash")
                if len(df_valid_filtered) > 0:
                    cw = df_valid_filtered['Dev_CH_Relatif_Pct'].abs().idxmax()
                    cwv = df_valid_filtered.loc[cw, 'Dev_CH_Relatif_Pct']
                    fig_chd.add_annotation(
                        x=df_valid_filtered.loc[cw, 'Date'], y=cwv,
                        text=f"Worst {cwv:.1f}%",
                        showarrow=True, arrowhead=2, arrowwidth=1.5,
                        arrowcolor='#ef4444',
                        ax=30, ay=-25 if cwv > 0 else 25,
                        font=dict(size=9, color='#FFF'),
                        bgcolor='rgba(239,68,68,0.8)', borderpad=3)
                fig_chd.update_layout(
                    title=dict(text='CH Daily Deviation (%)', font=dict(size=13, color='#ffffff')),
                    height=320,
                    font=dict(family='Inter, sans-serif', color='#e5e7eb'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(gridcolor='rgba(255,255,255,0.08)', title=''),
                    yaxis=dict(gridcolor='rgba(255,255,255,0.08)', title='Deviation (%)'),
                    hovermode='x unified',
                    margin=dict(l=60, r=20, t=50, b=40)
                )
                st.plotly_chart(fig_chd, use_container_width=True, key='cpp_ch_dev')

                # CM Deviation
                c_cm = [get_status_color(s) for s in df_valid_filtered['Status_CM']]
                fig_cmd = go.Figure()
                fig_cmd.add_shape(type='rect', xref='paper', yref='y',
                    x0=0, x1=1, y0=-2, y1=2,
                    fillcolor='rgba(34,197,94,0.04)', line_width=0, layer='below')
                fig_cmd.add_trace(go.Bar(
                    x=df_valid_filtered['Date'],
                    y=df_valid_filtered['Dev_CM_Relatif_Pct'],
                    marker_color=c_cm, showlegend=False,
                    hovertemplate='<b>CM Dev</b><br>%{y:.2f}%<extra></extra>'
                ))
                fig_cmd.add_hline(y=0, line_color="#9ca3af", line_width=2)
                fig_cmd.add_hline(y=2, line_color="#22c55e", line_width=1, line_dash="dash",
                    annotation_text="2%", annotation_position="top right",
                    annotation_font=dict(size=8, color='#22c55e'))
                fig_cmd.add_hline(y=-2, line_color="#22c55e", line_width=1, line_dash="dash")
                fig_cmd.add_hline(y=3, line_color="#eab308", line_width=1, line_dash="dash",
                    annotation_text="3%", annotation_position="top right",
                    annotation_font=dict(size=8, color='#eab308'))
                fig_cmd.add_hline(y=-3, line_color="#eab308", line_width=1, line_dash="dash")
                if len(df_valid_filtered) > 0:
                    mw = df_valid_filtered['Dev_CM_Relatif_Pct'].abs().idxmax()
                    mwv = df_valid_filtered.loc[mw, 'Dev_CM_Relatif_Pct']
                    fig_cmd.add_annotation(
                        x=df_valid_filtered.loc[mw, 'Date'], y=mwv,
                        text=f"Worst {mwv:.1f}%",
                        showarrow=True, arrowhead=2, arrowwidth=1.5,
                        arrowcolor='#ef4444',
                        ax=30, ay=-25 if mwv > 0 else 25,
                        font=dict(size=9, color='#FFF'),
                        bgcolor='rgba(239,68,68,0.8)', borderpad=3)
                fig_cmd.update_layout(
                    title=dict(text='CM Daily Deviation (%)', font=dict(size=13, color='#ffffff')),
                    height=320,
                    font=dict(family='Inter, sans-serif', color='#e5e7eb'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(gridcolor='rgba(255,255,255,0.08)', title=''),
                    yaxis=dict(gridcolor='rgba(255,255,255,0.08)', title='Deviation (%)'),
                    hovermode='x unified',
                    margin=dict(l=60, r=20, t=50, b=40)
                )
                st.plotly_chart(fig_cmd, use_container_width=True, key='cpp_cm_dev')

            # ── CH-CM DATA TABLE — styled HTML (replace st.dataframe) ──
            st.markdown("""
            <div class="section-header" style="margin-top:2.5rem;">
                <h3 class="section-title">CH-CM Data Table</h3>
            </div>
            """, unsafe_allow_html=True)

            if not df_valid_filtered.empty:
                chcm_cols_def = [
                    ('Date', 'Date'), ('Port_Darat', 'Port Darat'), ('Port_Laut', 'Port Laut'),
                    ('Port_Total', 'Port Total'), ('CPP_Raw', 'CPP Raw'), ('CPP_Product', 'CPP Product'),
                    ('CPP_Total', 'CPP Total'), ('Sales', 'Sales'),
                    ('CH_WB', 'CH WB'), ('TWB_CH', 'TWB CH'), ('Dev_CH_Relatif_Pct', 'CH Dev%'), ('Status_CH', 'Status CH'),
                    ('CM_WB', 'CM WB'), ('TWB_CM', 'TWB CM'), ('Dev_CM_Relatif_Pct', 'CM Dev%'), ('Status_CM', 'Status CM'),
                ]
                chcm_cols = [(k, v) for k, v in chcm_cols_def if k in df_valid_filtered.columns]

                chcm_header = ''.join(
                    f'<th style="padding:9px 8px;text-align:center;color:#e2e8f0;font-weight:700;'
                    f'font-size:0.62rem;text-transform:uppercase;letter-spacing:0.04em;'
                    f'position:sticky;top:0;background:rgba(22,101,52,0.95);z-index:1;white-space:nowrap;">{label}</th>'
                    for _, label in chcm_cols
                )

                chcm_rows = ""
                for _, row in df_valid_filtered.iterrows():
                    ch_st = row.get('Status_CH', 'Normal')
                    cm_st = row.get('Status_CM', 'Normal')
                    worst = 'Critical' if 'Critical' in (ch_st, cm_st) else 'Caution' if 'Caution' in (ch_st, cm_st) else 'Normal'

                    if worst == 'Critical':
                        row_bg, bl = 'rgba(239,68,68,0.06)', '3px solid #ef4444'
                    elif worst == 'Caution':
                        row_bg, bl = 'rgba(245,158,11,0.06)', '3px solid #f59e0b'
                    else:
                        row_bg, bl = 'rgba(34,197,94,0.03)', '3px solid #22c55e'

                    cells = ""
                    for col_key, _ in chcm_cols:
                        val = row.get(col_key, '')

                        if col_key in ('Status_CH', 'Status_CM'):
                            s = str(val)
                            if s == 'Critical':
                                sc, sbg = '#f87171', 'rgba(239,68,68,0.15)'
                            elif s == 'Caution':
                                sc, sbg = '#fbbf24', 'rgba(245,158,11,0.15)'
                            else:
                                sc, sbg = '#4ade80', 'rgba(34,197,94,0.15)'
                            cells += (f'<td style="padding:8px 6px;text-align:center;">'
                                      f'<span style="background:{sbg};color:{sc};padding:2px 8px;'
                                      f'border-radius:20px;font-size:0.65rem;font-weight:700;">{s}</span></td>')

                        elif col_key in ('Dev_CH_Relatif_Pct', 'Dev_CM_Relatif_Pct'):
                            dv = float(val) if pd.notnull(val) else 0
                            dc = '#f87171' if abs(dv) > 3 else '#fbbf24' if abs(dv) > 2 else '#4ade80'
                            cells += (f'<td style="padding:8px 6px;text-align:center;color:{dc};'
                                      f'font-weight:700;font-size:0.75rem;">{dv:+.2f}%</td>')

                        elif col_key == 'Date':
                            cells += (f'<td style="padding:8px 6px;text-align:center;color:#cbd5e1;'
                                      f'font-size:0.72rem;white-space:nowrap;">{str(val)[:10]}</td>')

                        elif isinstance(val, (int, float)) and pd.notnull(val):
                            cells += (f'<td style="padding:8px 6px;text-align:center;color:#cbd5e1;'
                                      f'font-size:0.72rem;">{val:,.2f}</td>')
                        else:
                            cells += (f'<td style="padding:8px 6px;text-align:center;color:#cbd5e1;'
                                      f'font-size:0.72rem;">{val}</td>')

                    chcm_rows += (f'<tr style="background:{row_bg};border-left:{bl};'
                                  f'border-bottom:1px solid rgba(148,163,184,0.06);">{cells}</tr>')

                st.markdown(f"""
                <div style="background:linear-gradient(135deg,rgba(20,20,40,0.6),rgba(15,23,42,0.8));
                    border:1px solid rgba(148,163,184,0.1);border-radius:10px;overflow:hidden;
                    max-height:420px;overflow-y:auto;overflow-x:auto;">
                <table style="width:100%;border-collapse:collapse;min-width:1200px;">
                <thead><tr style="background:rgba(22,101,52,0.95);">{chcm_header}</tr></thead>
                <tbody>{chcm_rows}</tbody>
                </table></div>
                """, unsafe_allow_html=True)
            else:
                st.info("Tidak ada data untuk ditampilkan")

    with tab4:
        # ══════════════════════════════════════════════════════════
        # TAB 4: PERFORMANCE REPORTS
        # ══════════════════════════════════════════════════════════

        matrix = create_matrix(df_ob, df_ch_cm)
        flow = create_flow(df_ch_cm)
        df_ch = df_ch_cm[df_ch_cm['CH_WB'].notna()].copy()
        df_cm_data = df_ch_cm[df_ch_cm['CM_WB'].notna()].copy()

        ob_avg = matrix.iloc[0]['Avg'] if len(matrix) > 0 else 0
        ch_avg = matrix.iloc[1]['Avg'] if len(matrix) > 1 else 0
        cm_avg = matrix.iloc[2]['Avg'] if len(matrix) > 2 else 0

        total_normal = int(matrix['Normal'].sum())
        total_caution = int(matrix['Caution'].sum())
        total_critical = int(matrix['Critical'].sum())
        total_all = int(matrix['Total'].sum())
        perf = total_normal / total_all * 100 if total_all > 0 else 0

        ob_avg_dev = df_ob['Dev_Relatif_Pct'].abs().mean()
        ch_avg_dev = df_ch['Dev_CH_Relatif_Pct'].abs().mean() if len(df_ch) > 0 else 0
        cm_avg_dev = df_cm_data['Dev_CM_Relatif_Pct'].abs().mean() if len(df_cm_data) > 0 else 0

        avg_eff = flow['Overall Efficiency (%)'].mean() if (flow is not None and len(flow) > 0) else 0
        avg_cm = flow['CM TWB'].mean() if (flow is not None and len(flow) > 0) else 0
        avg_ch = flow['CH TWB'].mean() if (flow is not None and len(flow) > 0) else 0
        avg_sales = flow['Sales'].mean() if (flow is not None and len(flow) > 0) else 0
        cm_loss = flow['CM Loss'].mean() if (flow is not None and len(flow) > 0) else 0
        ch_loss = flow['CH Loss'].mean() if (flow is not None and len(flow) > 0) else 0
        ch_eff = flow['CH Efficiency (%)'].mean() if (flow is not None and len(flow) > 0) else 0
        sales_eff = flow['Sales Efficiency (%)'].mean() if (flow is not None and len(flow) > 0) else 0

        report_date = datetime.now().strftime("%d %B %Y")

        def get_status_label(val):
            if val <= 2: return "Normal", "#4ade80"
            elif val <= 3: return "Caution", "#fbbf24"
            else: return "Critical", "#f87171"

        def scolor(status):
            if status == 'Normal': return '#22C55E'
            elif status == 'Caution': return '#F59E0B'
            return '#EF4444'

        def sbadge(val):
            if val <= 2: return '#22C55E', 'Normal'
            elif val <= 3: return '#F59E0B', 'Caution'
            return '#EF4444', 'Critical'

        ob_lbl, ob_c = get_status_label(ob_avg)
        ch_lbl, ch_c = get_status_label(ch_avg)
        cm_lbl, cm_c = get_status_label(cm_avg)
        perf_c = '#4ade80' if perf >= 70 else '#fbbf24' if perf >= 50 else '#f87171'
        perf_lbl = "Good" if perf >= 70 else "Fair" if perf >= 50 else "Poor"

        ob_clr, ob_lbl2 = sbadge(ob_avg_dev)
        ch_clr, ch_lbl2 = sbadge(ch_avg_dev)
        cm_clr, cm_lbl2 = sbadge(cm_avg_dev)

        # ════════════════════════════════════════════════════
        # 2) PERFORMANCE MATRIX TABLE
        # ════════════════════════════════════════════════════
        st.markdown("""
        <div class="section-header" style="margin-top:2.5rem;">
            <h3 class="section-title">Performance Matrix</h3>
        </div>
        """, unsafe_allow_html=True)

        if not matrix.empty:
            perf_html = """
            <div style="background:linear-gradient(135deg,rgba(20,20,40,0.6),rgba(15,23,42,0.8));
                border:1px solid rgba(148,163,184,0.1);border-radius:12px;overflow:hidden;">
            <table style="width:100%;border-collapse:collapse;font-size:0.82rem;">
            <thead>
            <tr style="background:rgba(22,101,52,0.4);">
                <th style="padding:12px 16px;text-align:left;color:#e2e8f0;font-weight:700;font-size:0.72rem;text-transform:uppercase;letter-spacing:0.08em;">Stage</th>
                <th style="padding:12px 14px;text-align:center;color:#e2e8f0;font-weight:700;font-size:0.72rem;text-transform:uppercase;">Avg |Dev%|</th>
                <th style="padding:12px 14px;text-align:center;color:#4ade80;font-weight:700;font-size:0.72rem;text-transform:uppercase;">Normal</th>
                <th style="padding:12px 14px;text-align:center;color:#fbbf24;font-weight:700;font-size:0.72rem;text-transform:uppercase;">Caution</th>
                <th style="padding:12px 14px;text-align:center;color:#f87171;font-weight:700;font-size:0.72rem;text-transform:uppercase;">Critical</th>
                <th style="padding:12px 14px;text-align:center;color:#e2e8f0;font-weight:700;font-size:0.72rem;text-transform:uppercase;">Total</th>
                <th style="padding:12px 14px;text-align:center;color:#e2e8f0;font-weight:700;font-size:0.72rem;text-transform:uppercase;">Performance</th>
            </tr>
            </thead><tbody>"""

            stage_names = ['Overburden (OB)', 'Coal Hauling (CH/Port)', 'Coal Mining (CM/CPP33)']
            for idx, row in matrix.iterrows():
                avg_d = row['Avg']
                dc = '#4ade80' if avg_d <= 2 else '#fbbf24' if avg_d <= 3 else '#f87171'
                norm = int(row['Normal'])
                caut = int(row['Caution'])
                crit = int(row['Critical'])
                tot = int(row['Total'])
                p = norm / tot * 100 if tot > 0 else 0
                pc = '#4ade80' if p >= 70 else '#fbbf24' if p >= 50 else '#f87171'
                name = stage_names[idx] if idx < len(stage_names) else f"Stage {idx}"
                perf_html += f"""
                <tr style="border-bottom:1px solid rgba(148,163,184,0.06);">
                    <td style="padding:11px 16px;color:#e2e8f0;font-weight:600;">{name}</td>
                    <td style="padding:11px 14px;text-align:center;color:{dc};font-weight:700;">{avg_d:.2f}%</td>
                    <td style="padding:11px 14px;text-align:center;color:#4ade80;font-weight:700;">{norm}</td>
                    <td style="padding:11px 14px;text-align:center;color:#fbbf24;font-weight:700;">{caut}</td>
                    <td style="padding:11px 14px;text-align:center;color:#f87171;font-weight:700;">{crit}</td>
                    <td style="padding:11px 14px;text-align:center;color:#94a3b8;">{tot}</td>
                    <td style="padding:11px 14px;text-align:center;color:{pc};font-weight:700;">{p:.1f}%</td>
                </tr>"""

            perf_html += f"""
            <tr style="border-top:2px solid #22c55e;background:rgba(34,197,94,0.06);">
                <td style="padding:12px 16px;color:#22c55e;font-weight:800;">OVERALL</td>
                <td style="padding:12px 14px;text-align:center;color:#64748b;">—</td>
                <td style="padding:12px 14px;text-align:center;color:#22c55e;font-weight:800;">{total_normal}</td>
                <td style="padding:12px 14px;text-align:center;color:#fbbf24;font-weight:800;">{total_caution}</td>
                <td style="padding:12px 14px;text-align:center;color:#f87171;font-weight:800;">{total_critical}</td>
                <td style="padding:12px 14px;text-align:center;color:#e2e8f0;font-weight:800;">{total_all}</td>
                <td style="padding:12px 14px;text-align:center;color:{perf_c};font-weight:800;">{perf:.1f}%</td>
            </tr>"""

            perf_html += "</tbody></table></div>"
            st.markdown(perf_html, unsafe_allow_html=True)

        # ════════════════════════════════════════════════════
        # 3) CRITICAL / CAUTION PERIODS — styled HTML tables (scrollable)
        # ════════════════════════════════════════════════════
        st.markdown("""
        <div class="section-header" style="margin-top:2.5rem;">
            <h3 class="section-title">Critical / Caution Periods</h3>
        </div>
        """, unsafe_allow_html=True)

        def render_alert_table(df_alert, cols, dev_col, status_col):
            """Render styled HTML table with KPI-colored rows and scroll"""
            if df_alert.empty:
                return '<div style="color:#4ade80;font-size:0.8rem;padding:12px;">Semua periode Normal</div>'

            header_html = ''.join(
                f'<th style="padding:10px 8px;text-align:center;color:#e2e8f0;font-weight:700;'
                f'font-size:0.65rem;text-transform:uppercase;letter-spacing:0.06em;'
                f'position:sticky;top:0;background:rgba(22,101,52,0.95);z-index:1;">{c}</th>'
                for c in cols
            )

            rows_html = ""
            for _, row in df_alert.iterrows():
                status = row[status_col]
                if status == 'Critical':
                    row_bg = 'rgba(239,68,68,0.08)'
                    border_left = '3px solid #ef4444'
                    status_color = '#f87171'
                    status_bg = 'rgba(239,68,68,0.15)'
                elif status == 'Caution':
                    row_bg = 'rgba(245,158,11,0.08)'
                    border_left = '3px solid #f59e0b'
                    status_color = '#fbbf24'
                    status_bg = 'rgba(245,158,11,0.15)'
                else:
                    row_bg = 'transparent'
                    border_left = '3px solid #22c55e'
                    status_color = '#4ade80'
                    status_bg = 'rgba(34,197,94,0.15)'

                cells = ""
                for c in cols:
                    val = row[c]
                    if c == status_col:
                        cells += (
                            f'<td style="padding:8px 6px;text-align:center;">'
                            f'<span style="background:{status_bg};color:{status_color};'
                            f'padding:3px 8px;border-radius:20px;font-size:0.65rem;font-weight:700;">'
                            f'{val}</span></td>'
                        )
                    elif c == dev_col:
                        dev_val = float(val) if not pd.isna(val) else 0
                        dev_color = '#f87171' if abs(dev_val) > 3 else '#fbbf24' if abs(dev_val) > 2 else '#4ade80'
                        cells += (
                            f'<td style="padding:8px 6px;text-align:center;color:{dev_color};'
                            f'font-weight:700;font-size:0.75rem;">{dev_val:+.2f}%</td>'
                        )
                    elif isinstance(val, (int, float)) and not pd.isna(val):
                        cells += (
                            f'<td style="padding:8px 6px;text-align:center;color:#cbd5e1;'
                            f'font-size:0.72rem;">{val:,.2f}</td>'
                        )
                    else:
                        display_val = str(val)[:10] if 'Date' in c or 'date' in str(c).lower() else str(val)
                        cells += (
                            f'<td style="padding:8px 6px;text-align:center;color:#cbd5e1;'
                            f'font-size:0.72rem;">{display_val}</td>'
                        )

                rows_html += (
                    f'<tr style="background:{row_bg};border-left:{border_left};'
                    f'border-bottom:1px solid rgba(148,163,184,0.06);">{cells}</tr>'
                )

            return f"""
            <div style="background:linear-gradient(135deg,rgba(20,20,40,0.6),rgba(15,23,42,0.8));
                border:1px solid rgba(148,163,184,0.1);border-radius:10px;overflow:hidden;
                max-height:320px;overflow-y:auto;">
            <table style="width:100%;border-collapse:collapse;">
            <thead><tr style="background:rgba(22,101,52,0.95);">{header_html}</tr></thead>
            <tbody>{rows_html}</tbody>
            </table></div>"""

        col_cp1, col_cp2, col_cp3 = st.columns(3)

        with col_cp1:
            st.markdown('<div style="font-size:0.82rem;font-weight:700;color:#60a5fa;margin-bottom:10px;">Overburden (OB)</div>', unsafe_allow_html=True)
            ob_alert = df_ob[df_ob['Status'].isin(['Critical', 'Caution'])][['Bulan', 'TC', 'JS', 'Dev_Relatif_Pct', 'Status']].copy()
            st.markdown(render_alert_table(ob_alert, ['Bulan', 'TC', 'JS', 'Dev_Relatif_Pct', 'Status'], 'Dev_Relatif_Pct', 'Status'), unsafe_allow_html=True)

        with col_cp2:
            st.markdown('<div style="font-size:0.82rem;font-weight:700;color:#60a5fa;margin-bottom:10px;">Coal Hauling (CH)</div>', unsafe_allow_html=True)
            ch_alert = df_ch[df_ch['Status_CH'].isin(['Critical', 'Caution'])][['Date', 'TWB_CH', 'CH_WB', 'Dev_CH_Relatif_Pct', 'Status_CH']].copy() if len(df_ch) > 0 else pd.DataFrame()
            st.markdown(render_alert_table(ch_alert, ['Date', 'TWB_CH', 'CH_WB', 'Dev_CH_Relatif_Pct', 'Status_CH'], 'Dev_CH_Relatif_Pct', 'Status_CH'), unsafe_allow_html=True)

        with col_cp3:
            st.markdown('<div style="font-size:0.82rem;font-weight:700;color:#60a5fa;margin-bottom:10px;">Coal Mining (CM)</div>', unsafe_allow_html=True)
            cm_alert = df_cm_data[df_cm_data['Status_CM'].isin(['Critical', 'Caution'])][['Date', 'TWB_CM', 'CM_WB', 'Dev_CM_Relatif_Pct', 'Status_CM']].copy() if len(df_cm_data) > 0 else pd.DataFrame()
            st.markdown(render_alert_table(cm_alert, ['Date', 'TWB_CM', 'CM_WB', 'Dev_CM_Relatif_Pct', 'Status_CM'], 'Dev_CM_Relatif_Pct', 'Status_CM'), unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════
        # EXPORT
        # ══════════════════════════════════════════════════════════
        st.markdown("""
        <div class="section-header">
            <h3 class="section-title"> Export Reports</h3>
        </div>
        """, unsafe_allow_html=True)

        # ── Chart themes for export ──
        chart_theme_export = dict(
            font=dict(family='Inter, Arial, sans-serif', color='#E2E8F0', size=12),
            plot_bgcolor='#1E293B', paper_bgcolor='#1E293B',
            margin=dict(l=60, r=40, t=55, b=45), hovermode='x unified'
        )
        chart_theme_st = dict(
            font=dict(family='Inter, Arial, sans-serif', color='#E2E8F0', size=11),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=50, r=30, t=50, b=40), hovermode='x unified'
        )
        axis_style = dict(
            gridcolor='rgba(148,163,184,0.15)',
            tickfont=dict(size=11, color='#CBD5E1'),
            linecolor='rgba(148,163,184,0.25)', showline=True,
            title_font=dict(size=12, color='#CBD5E1')
        )
        legend_cfg = dict(
            orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1,
            font=dict(size=10, color='#CBD5E1'), bgcolor='rgba(0,0,0,0)'
        )

        # ── 8 Chart functions ──
        def make_fig1(theme):
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=df_ob['Bulan'], y=df_ob['TC'], mode='lines+markers', name='TC (Target)',
                line=dict(color='#60A5FA', width=2.5), marker=dict(size=7, color='#60A5FA'),
                hovertemplate='TC: %{y:,.0f} BCM<extra></extra>'))
            fig.add_trace(go.Scatter(x=df_ob['Bulan'], y=df_ob['JS'], mode='lines+markers', name='JS (Realisasi)',
                line=dict(color='#34D399', width=2.5), marker=dict(size=7, symbol='diamond', color='#34D399'),
                hovertemplate='JS: %{y:,.0f} BCM<extra></extra>'))
            fig.update_layout(**theme, height=350,
                title=dict(text='OB Monthly: TC vs JS Volume (BCM)', font=dict(size=14, color='#F1F5F9')),
                xaxis=dict(**axis_style, title='Bulan'), yaxis=dict(**axis_style, title='Volume (BCM)', zeroline=False),
                legend=legend_cfg)
            return fig

        def make_fig2(theme):
            fig = go.Figure()
            fig.add_trace(go.Bar(x=df_ob['Bulan'], y=df_ob['Dev_Relatif_Pct'],
                marker_color=[scolor(s) for s in df_ob['Status']],
                text=df_ob['Dev_Relatif_Pct'].apply(lambda x: f"{x:+.1f}%"), textposition='outside',
                textfont=dict(size=9, color='#CBD5E1'), hovertemplate='Deviasi: %{y:.2f}%<extra></extra>'))
            for yv, clr, dsh in [(0,'#64748B','solid'),(2,'#F59E0B','dash'),(-2,'#F59E0B','dash'),(3,'#EF4444','dash'),(-3,'#EF4444','dash')]:
                ann = '' if yv == 0 else (f'{"Caution" if abs(yv)==2 else "Critical"} ±{abs(yv)}%' if yv > 0 else '')
                fig.add_hline(y=yv, line_color=clr, line_dash=dsh, line_width=1, annotation_text=ann,
                    annotation_position='right', annotation_font_size=9, annotation_font_color=clr)
            avg_line = df_ob['Dev_Relatif_Pct'].abs().mean()
            fig.add_hline(y=avg_line, line_color='#A78BFA', line_dash='dot', line_width=1,
                annotation_text=f'Avg: {avg_line:.2f}%', annotation_position='top right',
                annotation_font_size=9, annotation_font_color='#A78BFA')
            fig.update_layout(**theme, height=350, showlegend=False,
                title=dict(text='OB Deviation Trend (%)', font=dict(size=14, color='#F1F5F9')),
                xaxis=dict(**axis_style, title='Bulan'), yaxis=dict(**axis_style, title='Deviasi Relatif (%)', zeroline=False))
            return fig

        def make_fig3(theme):
            fig = go.Figure()
            if len(df_ch) > 0:
                fig.add_trace(go.Scatter(x=df_ch['Date'], y=df_ch['TWB_CH'], mode='lines+markers', name='TWB (Actual)',
                    line=dict(color='#60A5FA', width=2), marker=dict(size=5), hovertemplate='TWB: %{y:,.0f}<extra></extra>'))
                fig.add_trace(go.Scatter(x=df_ch['Date'], y=df_ch['CH_WB'], mode='lines', name='WB (Target)',
                    line=dict(color='#34D399', width=2, dash='dash'), hovertemplate='WB: %{y:,.0f}<extra></extra>'))
            fig.update_layout(**theme, height=320,
                title=dict(text='Coal Hauling: TWB vs WB (ton)', font=dict(size=14, color='#F1F5F9')),
                xaxis=dict(**axis_style, tickformat='%d %b'), yaxis=dict(**axis_style, title='Ton', zeroline=False), legend=legend_cfg)
            return fig

        def make_fig4(theme):
            fig = go.Figure()
            if len(df_cm_data) > 0:
                fig.add_trace(go.Scatter(x=df_cm_data['Date'], y=df_cm_data['TWB_CM'], mode='lines+markers', name='TWB (Actual)',
                    line=dict(color='#60A5FA', width=2), marker=dict(size=5), hovertemplate='TWB: %{y:,.0f}<extra></extra>'))
                fig.add_trace(go.Scatter(x=df_cm_data['Date'], y=df_cm_data['CM_WB'], mode='lines', name='WB (Target)',
                    line=dict(color='#34D399', width=2, dash='dash'), hovertemplate='WB: %{y:,.0f}<extra></extra>'))
            fig.update_layout(**theme, height=320,
                title=dict(text='Coal Mining: TWB vs WB (ton)', font=dict(size=14, color='#F1F5F9')),
                xaxis=dict(**axis_style, tickformat='%d %b'), yaxis=dict(**axis_style, title='Ton', zeroline=False), legend=legend_cfg)
            return fig

        def make_fig5(theme):
            fig = go.Figure()
            if len(df_ch) > 0:
                fig.add_trace(go.Bar(x=df_ch['Date'], y=df_ch['Dev_CH_Relatif_Pct'],
                    marker_color=[scolor(s) for s in df_ch['Status_CH']], hovertemplate='CH Dev: %{y:.1f}%<extra></extra>'))
                fig.add_hline(y=0, line_color='#64748B', line_width=0.5)
                fig.add_hline(y=3, line_color='#EF4444', line_dash='dash', line_width=1, annotation_text='Critical +3%',
                    annotation_position='top right', annotation_font_size=8, annotation_font_color='#EF4444')
                fig.add_hline(y=-3, line_color='#EF4444', line_dash='dash', line_width=1, annotation_text='Critical -3%',
                    annotation_position='bottom right', annotation_font_size=8, annotation_font_color='#EF4444')
            fig.update_layout(**theme, height=280, showlegend=False,
                title=dict(text='CH Daily Deviation (%)', font=dict(size=14, color='#F1F5F9')),
                xaxis=dict(**axis_style, tickformat='%d %b'), yaxis=dict(**axis_style, title='Deviasi (%)', zeroline=False))
            return fig

        def make_fig6(theme):
            fig = go.Figure()
            if len(df_cm_data) > 0:
                fig.add_trace(go.Bar(x=df_cm_data['Date'], y=df_cm_data['Dev_CM_Relatif_Pct'],
                    marker_color=[scolor(s) for s in df_cm_data['Status_CM']], hovertemplate='CM Dev: %{y:.1f}%<extra></extra>'))
                fig.add_hline(y=0, line_color='#64748B', line_width=0.5)
                fig.add_hline(y=3, line_color='#EF4444', line_dash='dash', line_width=1, annotation_text='Critical +3%',
                    annotation_position='top right', annotation_font_size=8, annotation_font_color='#EF4444')
                fig.add_hline(y=-3, line_color='#EF4444', line_dash='dash', line_width=1, annotation_text='Critical -3%',
                    annotation_position='bottom right', annotation_font_size=8, annotation_font_color='#EF4444')
            fig.update_layout(**theme, height=280, showlegend=False,
                title=dict(text='CM Daily Deviation (%)', font=dict(size=14, color='#F1F5F9')),
                xaxis=dict(**axis_style, tickformat='%d %b'), yaxis=dict(**axis_style, title='Deviasi (%)', zeroline=False))
            return fig

        def make_fig7(theme):
            fig = go.Figure()
            if flow is not None and len(flow) > 0:
                fig.add_trace(go.Scatter(x=flow['Date'], y=flow['CM TWB'], mode='lines+markers', name='CM TWB',
                    line=dict(color='#60A5FA', width=2.5), marker=dict(size=6, color='#60A5FA'),
                    hovertemplate='CM: %{y:,.0f}<extra></extra>'))
                fig.add_trace(go.Scatter(x=flow['Date'], y=flow['CH TWB'], mode='lines+markers', name='CH TWB',
                    line=dict(color='#34D399', width=2.5), marker=dict(size=6, color='#34D399'),
                    hovertemplate='CH: %{y:,.0f}<extra></extra>'))
                fig.add_trace(go.Scatter(x=flow['Date'], y=flow['Sales'], mode='lines+markers', name='Sales',
                    line=dict(color='#A78BFA', width=2.5), marker=dict(size=6, color='#A78BFA'),
                    hovertemplate='Sales: %{y:,.0f}<extra></extra>'))
                fig.add_hline(y=avg_cm, line_color='#60A5FA', line_dash='dot', line_width=1,
                    annotation_text=f'Avg CM {avg_cm:,.0f}', annotation_position='top left',
                    annotation_font=dict(size=9, color='#60A5FA'))
                fig.add_hline(y=avg_sales, line_color='#A78BFA', line_dash='dot', line_width=1,
                    annotation_text=f'Avg Sales {avg_sales:,.0f}', annotation_position='bottom right',
                    annotation_font=dict(size=9, color='#A78BFA'))
            fig.update_layout(**theme, height=380,
                title=dict(text='Material Throughput (ton)', font=dict(size=15, color='#F1F5F9', family='Inter, Arial, sans-serif')),
                xaxis=dict(**axis_style, tickformat='%d %b', title='Date'),
                yaxis=dict(**axis_style, title='Ton', zeroline=False),
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5,
                    font=dict(size=11, color='#E2E8F0'), bgcolor='rgba(30,41,59,0.8)',
                    bordercolor='rgba(148,163,184,0.2)', borderwidth=1))
            return fig

        def make_fig8(theme):
            fig = go.Figure()
            if flow is not None and len(flow) > 0:
                fig.add_trace(go.Scatter(x=flow['Date'], y=flow['Overall Efficiency (%)'],
                    mode='lines+markers', name='Overall Eff.',
                    line=dict(color='#34D399', width=3), marker=dict(size=8, color='#34D399',
                    line=dict(color='#ffffff', width=1)),
                    fill='tozeroy', fillcolor='rgba(52,211,153,0.1)',
                    hovertemplate='Eff: %{y:.1f}%<extra></extra>'))
                fig.add_trace(go.Scatter(x=flow['Date'], y=flow['CH Efficiency (%)'],
                    mode='lines+markers', name='CH Eff.',
                    line=dict(color='#60A5FA', width=2, dash='dot'), marker=dict(size=4, color='#60A5FA'),
                    hovertemplate='CH Eff: %{y:.1f}%<extra></extra>'))
                fig.add_hline(y=100, line_color='#F59E0B', line_dash='dash', line_width=1.5,
                    annotation_text='Target 100%', annotation_position='top right',
                    annotation_font=dict(size=10, color='#F59E0B'))
                avg_eff_val = flow['Overall Efficiency (%)'].mean()
                fig.add_hline(y=avg_eff_val, line_color='#34D399', line_dash='dot', line_width=1,
                    annotation_text=f'Avg {avg_eff_val:.1f}%', annotation_position='bottom left',
                    annotation_font=dict(size=9, color='#34D399'))
                eff_min = min(50, flow['Overall Efficiency (%)'].min() - 5)
                eff_max = max(110, flow['Overall Efficiency (%)'].max() + 5)
            else:
                eff_min, eff_max = 50, 110
            fig.update_layout(**theme, height=380,
                title=dict(text='Overall Efficiency (%)', font=dict(size=15, color='#F1F5F9', family='Inter, Arial, sans-serif')),
                xaxis=dict(**axis_style, tickformat='%d %b', title='Date'),
                yaxis=dict(**axis_style, title='Efficiency (%)', zeroline=False, range=[eff_min, eff_max]),
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5,
                    font=dict(size=11, color='#E2E8F0'), bgcolor='rgba(30,41,59,0.8)',
                    bordercolor='rgba(148,163,184,0.2)', borderwidth=1))
            return fig

        # ══════════════════════════════════════════════════════════
        # 3 EXPORT COLUMNS
        # ══════════════════════════════════════════════════════════
        exp1, exp2, exp3 = st.columns(3)

        # ═══════ EXCEL EXPORT ═══════
        with exp1:
            st.markdown("""
            <div style="background:rgba(22,101,52,0.08);padding:10px 12px;border-radius:8px;
                border-left:3px solid #16A34A;margin-bottom:8px;">
                <div style="font-size:0.82rem;font-weight:700;color:#e5e7eb;">📊 Professional Excel Report</div>
                <div style="font-size:0.72rem;color:#94A3B8;margin-top:2px;">7 sheets • Executive Dashboard • KPP Branding • Conditional formatting</div>
            </div>
            """, unsafe_allow_html=True)

            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
            from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
            from openpyxl.utils import get_column_letter

            excel_buffer = io.BytesIO()

            # ── KPP Mining Brand Color Palette ──
            C = {
                'kpp_dk': '166534', 'kpp_md': '16A34A', 'kpp_lt': 'DCFCE7', 'kpp_bg': 'F0FDF4',
                'gold_dk': '92400E', 'gold_md': 'D97706', 'gold_lt': 'FEF3C7',
                'gray1': '1F2937', 'gray2': '6B7280', 'gray3': 'F3F4F6', 'gray4': 'F9FAFB',
                'green_txt': '059669', 'green_bg': 'D1FAE5',
                'amber_txt': 'D97706', 'amber_bg': 'FEF3C7',
                'red_txt': 'DC2626', 'red_bg': 'FEE2E2',
                'white': 'FFFFFF', 'bdr': 'D1D5DB',
            }

            # ── Reusable styles ──
            hf = Font(bold=True, color=C['white'], size=10)
            hfl = PatternFill(start_color=C['kpp_dk'], end_color=C['kpp_dk'], fill_type='solid')
            bd = Border(
                left=Side(style='thin', color=C['bdr']), right=Side(style='thin', color=C['bdr']),
                top=Side(style='thin', color=C['bdr']), bottom=Side(style='thin', color=C['bdr']))
            al_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
            al_l = Alignment(horizontal='left', vertical='center', wrap_text=True)
            af = PatternFill(start_color=C['kpp_bg'], end_color=C['kpp_bg'], fill_type='solid')

            sfill = {
                'Normal': PatternFill(start_color=C['green_bg'], end_color=C['green_bg'], fill_type='solid'),
                'Good': PatternFill(start_color=C['green_bg'], end_color=C['green_bg'], fill_type='solid'),
                'Caution': PatternFill(start_color=C['amber_bg'], end_color=C['amber_bg'], fill_type='solid'),
                'Warning': PatternFill(start_color=C['amber_bg'], end_color=C['amber_bg'], fill_type='solid'),
                'Critical': PatternFill(start_color=C['red_bg'], end_color=C['red_bg'], fill_type='solid'),
            }
            sfont = {
                'Normal': C['green_txt'], 'Good': C['green_txt'],
                'Caution': C['amber_txt'], 'Warning': C['amber_txt'],
                'Critical': C['red_txt'],
            }

            # ── Helper: style data sheets ──
            def style_data_sheet(ws, title_text, hdr_row, data_start_row, merge_cols,
                                 status_col_idx=None, freeze_cell='A2',
                                 num_fmt_cols=None, pct_fmt_cols=None, date_fmt_cols=None):
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_cols)
                ws['A1'] = title_text
                ws['A1'].font = Font(size=14, bold=True, color=C['white'])
                ws['A1'].fill = PatternFill(start_color=C['kpp_dk'], end_color=C['kpp_dk'], fill_type='solid')
                ws['A1'].alignment = al_c
                ws.row_dimensions[1].height = 30

                for col in range(1, merge_cols + 1):
                    cell = ws.cell(row=hdr_row, column=col)
                    cell.font = hf
                    cell.fill = PatternFill(start_color=C['kpp_md'], end_color=C['kpp_md'], fill_type='solid')
                    cell.alignment = al_c
                    cell.border = bd
                ws.row_dimensions[hdr_row].height = 24

                for ri, row in enumerate(ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, max_col=merge_cols)):
                    for cell in row:
                        cell.border = bd
                        cell.alignment = al_c
                        if ri % 2 == 1:
                            cell.fill = af
                        if status_col_idx and cell.column == status_col_idx:
                            sv = str(cell.value) if cell.value else ''
                            if sv in sfill:
                                cell.fill = sfill[sv]
                                cell.font = Font(bold=True, size=9, color=sfont.get(sv, C['gray1']))
                        if num_fmt_cols and cell.column in num_fmt_cols:
                            cell.number_format = '#,##0'
                        if pct_fmt_cols and cell.column in pct_fmt_cols:
                            cell.number_format = '0.00'
                        if date_fmt_cols and cell.column in date_fmt_cols:
                            cell.number_format = 'YYYY-MM-DD'

                for ci in range(1, merge_cols + 1):
                    ml = 0
                    for row in ws.iter_rows(min_col=ci, max_col=ci):
                        for cell in row:
                            try:
                                if cell.value:
                                    ml = max(ml, len(str(cell.value)[:40]))
                            except:
                                pass
                    ws.column_dimensions[get_column_letter(ci)].width = max(min(ml + 3, 22), 10)

                ws.freeze_panes = freeze_cell

            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:

                # ════════════════════════════════════════
                # SHEET 1: EXECUTIVE DASHBOARD
                # ════════════════════════════════════════
                ed = []
                ed.append(['KPP MINING — DATA ANALYSIS REPORT', '', '', '', '', '', ''])
                ed.append([f'PT Kalimantan Prima Persada | Generated: {datetime.now().strftime("%d %b %Y, %H:%M")}', '', '', '', '', '', ''])
                ed.append(['', '', '', '', '', '', ''])
                ed.append(['KEY PERFORMANCE INDICATORS', '', '', '', '', '', ''])
                ed.append(['Overall Perf', 'Total Periods', 'Critical Alerts', 'Avg Efficiency', 'OB Avg Dev', 'CH Avg Dev', 'CM Avg Dev'])
                ed.append([
                    round(perf, 1), total_normal + total_caution + total_critical,
                    total_critical, round(avg_eff, 1),
                    round(ob_avg_dev, 2), round(ch_avg_dev, 2), round(cm_avg_dev, 2)
                ])
                ed.append(['', '', '', '', '', '', ''])
                ed.append(['OVERBURDEN (OB)', 'Value', 'Status', '', 'COAL HAULING (CH)', 'Value', 'Status'])
                ed.append(['Total Periods', len(df_ob), '', '', 'Total Periods', len(df_ch), ''])
                ed.append(['Avg Dev (%)', round(ob_avg_dev, 2), 'Normal' if ob_avg_dev <= 2 else 'Caution' if ob_avg_dev <= 3 else 'Critical', '', 'Avg Dev (%)', round(ch_avg_dev, 2), 'Normal' if ch_avg_dev <= 2 else 'Caution' if ch_avg_dev <= 3 else 'Critical'])
                ed.append(['Max Dev (%)', round(df_ob["Dev_Relatif_Pct"].max(), 2), '', '', 'Max Dev (%)', round(df_ch["Dev_CH_Relatif_Pct"].max(), 2) if len(df_ch) > 0 else '-', ''])
                ed.append(['Std Dev', round(df_ob["Dev_Relatif_Pct"].std(), 2), '', '', 'Std Dev', round(df_ch["Dev_CH_Relatif_Pct"].std(), 2) if len(df_ch) > 0 else '-', ''])
                ed.append(['Normal', int(matrix.iloc[0]['Normal']), '', '', 'Normal', int(matrix.iloc[1]['Normal']), ''])
                ed.append(['Caution', int(matrix.iloc[0]['Caution']), '', '', 'Caution', int(matrix.iloc[1]['Caution']), ''])
                ed.append(['Critical', int(matrix.iloc[0]['Critical']), '', '', 'Critical', int(matrix.iloc[1]['Critical']), ''])
                ed.append(['Total TC (BCM)', round(df_ob["TC"].sum()), '', '', 'Avg TWB', round(df_ch["TWB_CH"].mean()) if len(df_ch) > 0 else '-', ''])
                ed.append(['Total JS (BCM)', round(df_ob["JS"].sum()), '', '', 'Avg WB Target', round(df_ch["CH_WB"].mean()) if len(df_ch) > 0 else '-', ''])
                ed.append(['', '', '', '', '', '', ''])
                ed.append(['COAL MINING (CM)', 'Value', 'Status', '', 'OVERALL SUMMARY', 'Value', ''])
                ed.append(['Total Periods', len(df_cm_data), '', '', 'Total Normal', total_normal, ''])
                ed.append(['Avg Dev (%)', round(cm_avg_dev, 2), 'Normal' if cm_avg_dev <= 2 else 'Caution' if cm_avg_dev <= 3 else 'Critical', '', 'Total Caution', total_caution, ''])
                ed.append(['Max Dev (%)', round(df_cm_data["Dev_CM_Relatif_Pct"].max(), 2) if len(df_cm_data) > 0 else '-', '', '', 'Total Critical', total_critical, ''])
                ed.append(['Std Dev', round(df_cm_data["Dev_CM_Relatif_Pct"].std(), 2) if len(df_cm_data) > 0 else '-', '', '', 'Performance', round(perf, 1), ''])
                ed.append(['Normal', int(matrix.iloc[2]['Normal']), '', '', 'Avg Efficiency', round(avg_eff, 1), ''])
                ed.append(['Caution', int(matrix.iloc[2]['Caution']), '', '', '', '', ''])
                ed.append(['Critical', int(matrix.iloc[2]['Critical']), '', '', '', '', ''])
                ed.append(['Avg TWB', round(df_cm_data["TWB_CM"].mean()) if len(df_cm_data) > 0 else '-', '', '', '', '', ''])
                ed.append(['Avg WB Target', round(df_cm_data["CM_WB"].mean()) if len(df_cm_data) > 0 else '-', '', '', '', '', ''])

                df_exec = pd.DataFrame(ed)
                df_exec.to_excel(writer, sheet_name='Executive Dashboard', index=False, header=False)

                ws_e = writer.sheets['Executive Dashboard']
                # Title bar
                ws_e.merge_cells('A1:G1')
                ws_e['A1'].font = Font(size=16, bold=True, color=C['white'])
                ws_e['A1'].fill = PatternFill(start_color=C['kpp_dk'], end_color=C['kpp_dk'], fill_type='solid')
                ws_e['A1'].alignment = al_c
                ws_e.row_dimensions[1].height = 36
                # Subtitle
                ws_e.merge_cells('A2:G2')
                ws_e['A2'].font = Font(size=9, italic=True, color=C['kpp_dk'])
                ws_e['A2'].fill = PatternFill(start_color=C['kpp_lt'], end_color=C['kpp_lt'], fill_type='solid')
                ws_e['A2'].alignment = al_c
                ws_e['A2'].border = Border(bottom=Side(style='thin', color=C['kpp_md']))
                ws_e.row_dimensions[2].height = 20
                # KPI section header
                ws_e.merge_cells('A4:G4')
                ws_e['A4'].font = Font(size=11, bold=True, color=C['kpp_dk'])
                ws_e['A4'].fill = PatternFill(start_color=C['kpp_lt'], end_color=C['kpp_lt'], fill_type='solid')
                ws_e['A4'].alignment = al_l
                ws_e['A4'].border = Border(bottom=Side(style='medium', color=C['kpp_md']))
                ws_e.row_dimensions[4].height = 24
                # KPI headers
                for col in range(1, 8):
                    cell = ws_e.cell(row=5, column=col)
                    cell.font = Font(size=9, bold=True, color=C['gray1'])
                    cell.fill = PatternFill(start_color=C['gray3'], end_color=C['gray3'], fill_type='solid')
                    cell.alignment = al_c
                    cell.border = bd
                ws_e.row_dimensions[5].height = 26
                # KPI values (large, bold, green)
                for col in range(1, 8):
                    cell = ws_e.cell(row=6, column=col)
                    cell.font = Font(size=14, bold=True, color=C['kpp_dk'])
                    cell.alignment = al_c
                    cell.border = bd
                    if col == 3 and total_critical > 0:
                        cell.font = Font(size=14, bold=True, color=C['red_txt'])
                ws_e.row_dimensions[6].height = 32
                # Section sub-headers (OB/CH at row 8, CM/Overall at row 19)
                for rn in [8, 19]:
                    for col in range(1, 8):
                        cell = ws_e.cell(row=rn, column=col)
                        if cell.value and str(cell.value).strip():
                            cell.font = Font(size=10, bold=True, color=C['white'])
                            cell.fill = PatternFill(start_color=C['kpp_md'], end_color=C['kpp_md'], fill_type='solid')
                            cell.alignment = al_c
                            cell.border = bd
                    ws_e.row_dimensions[rn].height = 22
                # Data rows
                for rn in list(range(9, 18)) + list(range(20, 29)):
                    for col in range(1, 8):
                        cell = ws_e.cell(row=rn, column=col)
                        cell.font = Font(size=9, color=C['gray1'])
                        cell.alignment = al_l if col in [1, 5] else al_c
                        cell.border = bd
                        if rn % 2 == 0:
                            cell.fill = PatternFill(start_color=C['kpp_bg'], end_color=C['kpp_bg'], fill_type='solid')
                        if col in [3, 7]:
                            sv = str(cell.value) if cell.value else ''
                            if sv in sfill:
                                cell.fill = sfill[sv]
                                cell.font = Font(size=9, bold=True, color=sfont.get(sv, C['gray1']))
                        if col in [2, 6]:
                            try:
                                v = cell.value
                                if isinstance(v, (int, float)) and abs(v) > 999:
                                    cell.number_format = '#,##0'
                            except:
                                pass
                for letter, w in [('A',20),('B',15),('C',12),('D',2),('E',20),('F',15),('G',12)]:
                    ws_e.column_dimensions[letter].width = w
                ws_e.freeze_panes = 'A7'

                # ════════════════════════════════════════
                # SHEET 2: OB ANALYSIS
                # ════════════════════════════════════════
                df_ob_exp = df_ob[['Bulan','TC','JS','Dev_Absolut','Dev_Relatif_Pct','Status']].copy()
                df_ob_exp['Dev_Relatif_Pct'] = df_ob_exp['Dev_Relatif_Pct'].round(2)
                df_ob_exp['Dev_Absolut'] = df_ob_exp['Dev_Absolut'].round(0)
                df_ob_exp.columns = ['Month', 'TC (BCM)', 'JS (BCM)', 'Dev (Absolute)', 'Dev (%)', 'Status']
                df_ob_exp.to_excel(writer, sheet_name='OB Analysis', index=False, startrow=2)
                ws_ob = writer.sheets['OB Analysis']
                ws_ob.merge_cells('A2:F2')
                ws_ob['A2'] = f'Periods: {len(df_ob)} | Avg Dev: {ob_avg_dev:.2f}% | Normal: {int(matrix.iloc[0]["Normal"])} | Caution: {int(matrix.iloc[0]["Caution"])} | Critical: {int(matrix.iloc[0]["Critical"])}'
                ws_ob['A2'].font = Font(size=8, italic=True, color=C['kpp_dk'])
                ws_ob['A2'].fill = PatternFill(start_color=C['kpp_lt'], end_color=C['kpp_lt'], fill_type='solid')
                ws_ob['A2'].alignment = al_c
                style_data_sheet(ws_ob, 'OVERBURDEN (OB) ANALYSIS', 3, 4, 6,
                                 status_col_idx=6, freeze_cell='A4',
                                 num_fmt_cols=[2, 3, 4], pct_fmt_cols=[5])

                # ════════════════════════════════════════
                # SHEET 3: CH ANALYSIS
                # ════════════════════════════════════════
                df_ch_exp = df_ch_cm.dropna(subset=['CH_WB','TWB_CH']).copy()
                ch_cols = [c for c in ['Date','Port_Darat','Port_Laut','Port_Total','CH_WB','TWB_CH','Dev_CH_Relatif_Pct','Status_CH'] if c in df_ch_exp.columns]
                df_ch_exp = df_ch_exp[ch_cols].copy()
                if 'Dev_CH_Relatif_Pct' in df_ch_exp.columns:
                    df_ch_exp['Dev_CH_Relatif_Pct'] = df_ch_exp['Dev_CH_Relatif_Pct'].round(2)
                if 'Date' in df_ch_exp.columns:
                    df_ch_exp['Date'] = pd.to_datetime(df_ch_exp['Date']).dt.strftime('%Y-%m-%d')
                col_map_ch = {'Date':'Date','Port_Darat':'Port Darat','Port_Laut':'Port Laut','Port_Total':'Port Total','CH_WB':'WB Target','TWB_CH':'TWB Actual','Dev_CH_Relatif_Pct':'Dev (%)','Status_CH':'Status'}
                df_ch_exp.columns = [col_map_ch.get(c, c) for c in ch_cols]
                ncols_ch = len(df_ch_exp.columns)
                df_ch_exp.to_excel(writer, sheet_name='CH Analysis', index=False, startrow=2)
                ws_ch = writer.sheets['CH Analysis']
                style_data_sheet(ws_ch, 'COAL HAULING (CH) ANALYSIS', 3, 4, ncols_ch,
                                 status_col_idx=ncols_ch, freeze_cell='A4',
                                 num_fmt_cols=[2,3,4,5,6], pct_fmt_cols=[ncols_ch - 1])

                # ════════════════════════════════════════
                # SHEET 4: CM ANALYSIS
                # ════════════════════════════════════════
                df_cm_exp = df_ch_cm.dropna(subset=['CM_WB','TWB_CM']).copy()
                cm_cols = [c for c in ['Date','CPP_Raw','CPP_Product','CPP_Total','Sales','CM_WB','TWB_CM','Dev_CM_Relatif_Pct','Status_CM'] if c in df_cm_exp.columns]
                df_cm_exp = df_cm_exp[cm_cols].copy()
                if 'Dev_CM_Relatif_Pct' in df_cm_exp.columns:
                    df_cm_exp['Dev_CM_Relatif_Pct'] = df_cm_exp['Dev_CM_Relatif_Pct'].round(2)
                if 'Date' in df_cm_exp.columns:
                    df_cm_exp['Date'] = pd.to_datetime(df_cm_exp['Date']).dt.strftime('%Y-%m-%d')
                col_map_cm = {'Date':'Date','CPP_Raw':'CPP Raw','CPP_Product':'CPP Product','CPP_Total':'CPP Total','Sales':'Sales','CM_WB':'WB Target','TWB_CM':'TWB Actual','Dev_CM_Relatif_Pct':'Dev (%)','Status_CM':'Status'}
                df_cm_exp.columns = [col_map_cm.get(c, c) for c in cm_cols]
                ncols_cm = len(df_cm_exp.columns)
                df_cm_exp.to_excel(writer, sheet_name='CM Analysis', index=False, startrow=2)
                ws_cm = writer.sheets['CM Analysis']
                style_data_sheet(ws_cm, 'COAL MINING (CM) ANALYSIS', 3, 4, ncols_cm,
                                 status_col_idx=ncols_cm, freeze_cell='A4',
                                 num_fmt_cols=[2,3,4,5,6,7], pct_fmt_cols=[ncols_cm - 1])

                # ════════════════════════════════════════
                # SHEET 5: MATERIAL FLOW (conditional)
                # ════════════════════════════════════════
                if flow is not None and len(flow) > 0:
                    df_fl = flow.dropna(subset=['Sales']).copy()
                    df_fl = df_fl[df_fl['CM TWB'] > 0]
                    fl_cols = [c for c in ['Date','CM TWB','CH TWB','Sales','CM Loss','CH Loss',
                               'CH Efficiency (%)','Sales Efficiency (%)','Overall Efficiency (%)'] if c in df_fl.columns]
                    df_fl = df_fl[fl_cols].copy()
                    if 'Date' in df_fl.columns:
                        df_fl['Date'] = pd.to_datetime(df_fl['Date']).dt.strftime('%Y-%m-%d')
                    for ec in ['CH Efficiency (%)','Sales Efficiency (%)','Overall Efficiency (%)']:
                        if ec in df_fl.columns:
                            df_fl[ec] = df_fl[ec].round(1)
                    for lc in ['CM Loss','CH Loss']:
                        if lc in df_fl.columns:
                            df_fl[lc] = df_fl[lc].round(0)
                    ncols_fl = len(df_fl.columns)
                    df_fl.to_excel(writer, sheet_name='Material Flow', index=False, startrow=2)
                    ws_fl = writer.sheets['Material Flow']
                    ws_fl.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols_fl)
                    eff_mean = flow.dropna(subset=['Sales'])['Overall Efficiency (%)'].mean() if 'Overall Efficiency (%)' in flow.columns else 0
                    ws_fl['A2'] = f'Avg Overall Efficiency: {eff_mean:.1f}%'
                    ws_fl['A2'].font = Font(size=8, italic=True, color=C['kpp_dk'])
                    ws_fl['A2'].fill = PatternFill(start_color=C['kpp_lt'], end_color=C['kpp_lt'], fill_type='solid')
                    ws_fl['A2'].alignment = al_c
                    style_data_sheet(ws_fl, 'MATERIAL FLOW ANALYSIS', 3, 4, ncols_fl,
                                     freeze_cell='A4', num_fmt_cols=[2,3,4,5,6], pct_fmt_cols=[7,8,9])

                # ════════════════════════════════════════
                # SHEET 6: PERFORMANCE MATRIX
                # ════════════════════════════════════════
                df_mx = matrix.copy()
                df_mx.columns = ['Stage', 'Normal', 'Caution', 'Critical', 'Total', 'Avg Dev (%)']
                df_mx['Avg Dev (%)'] = df_mx['Avg Dev (%)'].round(2)
                total_row = pd.DataFrame([{
                    'Stage': 'TOTAL',
                    'Normal': df_mx['Normal'].sum(),
                    'Caution': df_mx['Caution'].sum(),
                    'Critical': df_mx['Critical'].sum(),
                    'Total': df_mx['Total'].sum(),
                    'Avg Dev (%)': round(df_mx['Avg Dev (%)'].mean(), 2)
                }])
                df_mx = pd.concat([df_mx, total_row], ignore_index=True)
                df_mx.to_excel(writer, sheet_name='Performance Matrix', index=False, startrow=2)
                ws_mx = writer.sheets['Performance Matrix']
                style_data_sheet(ws_mx, 'PERFORMANCE MATRIX', 3, 4, 6, freeze_cell='A4', pct_fmt_cols=[6])
                # Bold total row
                last_r = ws_mx.max_row
                for col in range(1, 7):
                    cell = ws_mx.cell(row=last_r, column=col)
                    cell.font = Font(bold=True, size=10, color=C['kpp_dk'])
                    cell.fill = PatternFill(start_color=C['kpp_lt'], end_color=C['kpp_lt'], fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin', color=C['bdr']), right=Side(style='thin', color=C['bdr']),
                        top=Side(style='medium', color=C['kpp_dk']), bottom=Side(style='medium', color=C['kpp_dk']))

                # ════════════════════════════════════════
                # SHEET 7: CRITICAL ALERTS
                # ════════════════════════════════════════
                alerts = []
                for _, r in df_ob[df_ob['Status'].isin(['Caution','Critical'])].iterrows():
                    alerts.append({
                        'Priority': 'HIGH' if r['Status']=='Critical' else 'MEDIUM',
                        'Stage': 'OB', 'Period': r['Bulan'],
                        'Dev (%)': round(r['Dev_Relatif_Pct'], 2), 'Status': r['Status']})
                if len(df_ch) > 0:
                    for _, r in df_ch[df_ch['Status_CH'].isin(['Caution','Critical'])].iterrows():
                        alerts.append({
                            'Priority': 'HIGH' if r['Status_CH']=='Critical' else 'MEDIUM',
                            'Stage': 'CH', 'Period': str(r['Date'])[:10],
                            'Dev (%)': round(r['Dev_CH_Relatif_Pct'], 2), 'Status': r['Status_CH']})
                if len(df_cm_data) > 0:
                    for _, r in df_cm_data[df_cm_data['Status_CM'].isin(['Caution','Critical'])].iterrows():
                        alerts.append({
                            'Priority': 'HIGH' if r['Status_CM']=='Critical' else 'MEDIUM',
                            'Stage': 'CM', 'Period': str(r['Date'])[:10],
                            'Dev (%)': round(r['Dev_CM_Relatif_Pct'], 2), 'Status': r['Status_CM']})
                if alerts:
                    df_al = pd.DataFrame(alerts)
                    sort_map = {'HIGH': 0, 'MEDIUM': 1}
                    df_al['_sort'] = df_al['Priority'].map(sort_map)
                    df_al = df_al.sort_values(['_sort', 'Dev (%)'], ascending=[True, False]).drop('_sort', axis=1)
                    df_al.to_excel(writer, sheet_name='Critical Alerts', index=False, startrow=2)
                    ws_al = writer.sheets['Critical Alerts']
                    ws_al.merge_cells('A1:E1')
                    ws_al['A1'] = '⚠ CRITICAL ALERT LOG'
                    ws_al['A1'].font = Font(size=14, bold=True, color=C['white'])
                    ws_al['A1'].fill = PatternFill(start_color=C['red_txt'], end_color=C['red_txt'], fill_type='solid')
                    ws_al['A1'].alignment = al_c
                    ws_al.row_dimensions[1].height = 30
                    for col in range(1, 6):
                        cell = ws_al.cell(row=3, column=col)
                        cell.font = Font(bold=True, size=9, color=C['white'])
                        cell.fill = PatternFill(start_color=C['gold_md'], end_color=C['gold_md'], fill_type='solid')
                        cell.alignment = al_c
                        cell.border = bd
                    for ri, row in enumerate(ws_al.iter_rows(min_row=4, max_row=ws_al.max_row, max_col=5)):
                        for cell in row:
                            cell.border = bd
                            cell.alignment = al_c
                            if ri % 2 == 1:
                                cell.fill = af
                            if cell.column == 1:
                                pv = str(cell.value) if cell.value else ''
                                if pv == 'HIGH':
                                    cell.fill = PatternFill(start_color=C['red_bg'], end_color=C['red_bg'], fill_type='solid')
                                    cell.font = Font(bold=True, size=9, color=C['red_txt'])
                                elif pv == 'MEDIUM':
                                    cell.fill = PatternFill(start_color=C['amber_bg'], end_color=C['amber_bg'], fill_type='solid')
                                    cell.font = Font(bold=True, size=9, color=C['amber_txt'])
                            if cell.column == 5:
                                sv = str(cell.value) if cell.value else ''
                                if sv in sfill:
                                    cell.fill = sfill[sv]
                                    cell.font = Font(bold=True, size=9, color=sfont.get(sv, C['gray1']))
                    for ci in range(1, 6):
                        ml = 0
                        for row in ws_al.iter_rows(min_col=ci, max_col=ci):
                            for cell in row:
                                try:
                                    if cell.value:
                                        ml = max(ml, len(str(cell.value)[:30]))
                                except:
                                    pass
                        ws_al.column_dimensions[get_column_letter(ci)].width = max(min(ml + 3, 22), 10)
                    ws_al.freeze_panes = 'A4'

            # ════════════════════════════════════════
            # POST-PROCESSING: Conditional Formatting
            # ════════════════════════════════════════
            excel_buffer.seek(0)
            wb = openpyxl.load_workbook(excel_buffer)

            try:
                if 'OB Analysis' in wb.sheetnames:
                    ws = wb['OB Analysis']
                    ws.conditional_formatting.add(
                        f'E4:E{ws.max_row}',
                        ColorScaleRule(
                            start_type='num', start_value=-5, start_color=C['red_txt'],
                            mid_type='num', mid_value=0, mid_color='FFFFFF',
                            end_type='num', end_value=5, end_color=C['green_txt']))
            except:
                pass

            try:
                if 'Material Flow' in wb.sheetnames:
                    ws = wb['Material Flow']
                    for cl in ['G', 'H', 'I']:
                        ws.conditional_formatting.add(
                            f'{cl}4:{cl}{ws.max_row}',
                            ColorScaleRule(
                                start_type='num', start_value=85, start_color=C['red_txt'],
                                mid_type='num', mid_value=100, mid_color='FFFFFF',
                                end_type='num', end_value=115, end_color=C['green_txt']))
            except:
                pass

            try:
                if 'Performance Matrix' in wb.sheetnames:
                    ws = wb['Performance Matrix']
                    mr = ws.max_row - 1
                    ws.conditional_formatting.add(f'B4:B{mr}', DataBarRule(start_type='num', start_value=0, end_type='num', end_value=30, color=C['green_txt']))
                    ws.conditional_formatting.add(f'C4:C{mr}', DataBarRule(start_type='num', start_value=0, end_type='num', end_value=30, color=C['gold_md']))
                    ws.conditional_formatting.add(f'D4:D{mr}', DataBarRule(start_type='num', start_value=0, end_type='num', end_value=30, color=C['red_txt']))
            except:
                pass

            buf_final = io.BytesIO()
            wb.save(buf_final)
            buf_final.seek(0)

            st.download_button(
                label="Download Excel Report",
                data=buf_final.getvalue(),
                file_name=f"KPP_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key='dl_excel', use_container_width=True)

        # ═══════ HTML EXPORT ═══════
        with exp2:
            st.markdown("""
            <div style="background:rgba(34,197,94,0.06);padding:10px 12px;border-radius:8px;
                border-left:3px solid #22C55E;margin-bottom:8px;">
                <div style="font-size:0.82rem;font-weight:700;color:#e5e7eb;"> HTML Report</div>
                <div style="font-size:0.72rem;color:#94A3B8;margin-top:2px;">Interactive charts with hover/zoom</div>
            </div>
            """, unsafe_allow_html=True)

            figs_html = [make_fig1(chart_theme_st), make_fig2(chart_theme_st), make_fig3(chart_theme_st),
                         make_fig4(chart_theme_st), make_fig5(chart_theme_st), make_fig6(chart_theme_st),
                         make_fig7(chart_theme_st), make_fig8(chart_theme_st)]
            f_divs = [f.to_html(full_html=False, include_plotlyjs=False) for f in figs_html]

            ch_avg_twb = df_ch['TWB_CH'].mean() if len(df_ch) > 0 else 0
            cm_avg_twb = df_cm_data['TWB_CM'].mean() if len(df_cm_data) > 0 else 0
            ch_avg_wb = df_ch['CH_WB'].mean() if len(df_ch) > 0 else 0
            cm_avg_wb = df_cm_data['CM_WB'].mean() if len(df_cm_data) > 0 else 0
            ob_max_dev = df_ob['Dev_Relatif_Pct'].abs().max()
            ob_tc_total = df_ob['TC'].sum()
            ob_js_total = df_ob['JS'].sum()
            ob_normal_n = int(matrix.iloc[0]['Normal'])
            ob_caution_n = int(matrix.iloc[0]['Caution'])
            ob_critical_n = int(matrix.iloc[0]['Critical'])

            stats_a = [
                ('Avg Deviation', f'{ob_avg_dev:.2f}%', '#F59E0B' if ob_avg_dev > 2 else '#22C55E'),
                ('Max Deviation', f'{ob_max_dev:.2f}%', '#EF4444' if ob_max_dev > 3 else '#F59E0B'),
                ('Total TC', f'{ob_tc_total:,.0f} BCM', '#60A5FA'),
                ('Total JS', f'{ob_js_total:,.0f} BCM', '#34D399'),
                ('Normal', str(ob_normal_n), '#22C55E'),
                ('Caution', str(ob_caution_n), '#F59E0B'),
                ('Critical', str(ob_critical_n), '#EF4444'),
            ]
            stats_b = [
                ('CH Avg TWB', f'{ch_avg_twb:,.0f} ton', '#60A5FA'),
                ('CH Avg WB', f'{ch_avg_wb:,.0f} ton', '#34D399'),
                ('CH Avg Dev', f'{ch_avg_dev:.2f}%', '#F59E0B' if ch_avg_dev > 2 else '#22C55E'),
                ('CM Avg TWB', f'{cm_avg_twb:,.0f} ton', '#60A5FA'),
                ('CM Avg WB', f'{cm_avg_wb:,.0f} ton', '#34D399'),
                ('CM Avg Dev', f'{cm_avg_dev:.2f}%', '#F59E0B' if cm_avg_dev > 2 else '#22C55E'),
            ]
            stats_c = [
                ('CH Normal', str(int(matrix.iloc[1]['Normal'])), '#22C55E'),
                ('CH Caution', str(int(matrix.iloc[1]['Caution'])), '#F59E0B'),
                ('CH Critical', str(int(matrix.iloc[1]['Critical'])), '#EF4444'),
                ('CM Normal', str(int(matrix.iloc[2]['Normal'])), '#22C55E'),
                ('CM Caution', str(int(matrix.iloc[2]['Caution'])), '#F59E0B'),
                ('CM Critical', str(int(matrix.iloc[2]['Critical'])), '#EF4444'),
            ]
            stats_d = [
                ('Avg CM TWB', f'{avg_cm:,.0f} ton', '#60A5FA'),
                ('Avg CH TWB', f'{avg_ch:,.0f} ton', '#34D399'),
                ('Avg Sales', f'{avg_sales:,.0f} ton', '#A78BFA'),
                ('CM Loss', f'{cm_loss:,.0f} ton', '#F59E0B'),
                ('CH Eff', f'{ch_eff:.1f}%', '#34D399'),
                ('Overall Eff', f'{avg_eff:.1f}%', '#34D399'),
            ]
            all_stats = [stats_a, stats_b, stats_c, stats_d]

            html_report = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>KPP AUTOMATIC Deviation Analytics Report {report_date}</title>
<script src="https://cdn.plot.ly/plotly-2.35.0.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Inter,sans-serif;background:#0F172A;color:#E2E8F0;min-height:100vh}}
.rc{{max-width:1200px;margin:0 auto;padding:2rem 2.5rem}}
.rh{{text-align:center;padding:2.5rem 2rem;margin-bottom:2rem;background:linear-gradient(135deg,#0F172A,#1E293B);border:1px solid rgba(148,163,184,0.15);border-radius:16px}}
.rh h1{{font-size:1.8rem;font-weight:900;color:#22C55E;margin-bottom:0.3rem}}
.rh .sub{{color:#94A3B8;font-size:0.95rem}}
.rh .dt{{color:#64748B;font-size:0.85rem;margin-top:0.3rem}}
.sec{{margin-bottom:2rem}}
.sh{{display:flex;align-items:center;gap:0.8rem;padding:0.8rem 1.2rem;margin-bottom:1rem;background:linear-gradient(90deg,rgba(34,197,94,0.1),transparent);border-left:4px solid #22C55E;border-radius:0 8px 8px 0}}
.sh h2{{font-size:1rem;font-weight:700;color:#F1F5F9}}
.kr{{display:grid;grid-template-columns:repeat(4,1fr);gap:1rem;margin-bottom:0.5rem}}
.kc{{background:#1E293B;border:1px solid rgba(148,163,184,0.12);border-radius:12px;padding:1.2rem 1rem;text-align:center;position:relative;overflow:hidden}}
.kc::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:var(--ac)}}
.kc .lb{{font-size:0.72rem;color:#64748B;text-transform:uppercase;letter-spacing:1px;font-weight:600}}
.kc .vl{{font-size:1.6rem;font-weight:900;color:#F1F5F9;margin:0.4rem 0 0.2rem}}
.kc .st{{font-size:0.8rem;font-weight:600}}
.kc .dt{{font-size:0.72rem;color:#64748B;margin-top:0.3rem}}
.cg{{display:grid;grid-template-columns:1fr 1fr;gap:1rem}}
.cc{{background:#1E293B;border:1px solid rgba(148,163,184,0.1);border-radius:12px;padding:0.8rem;overflow:hidden}}
.fc{{display:flex;align-items:center;justify-content:center;gap:0.5rem;padding:1.5rem 1rem;background:#1E293B;border:1px solid rgba(148,163,184,0.1);border-radius:12px;margin-bottom:1rem;flex-wrap:wrap}}
.fn{{background:#0F172A;border:1px solid rgba(148,163,184,0.15);border-radius:10px;padding:1rem 1.5rem;text-align:center;min-width:140px}}
.fn .fl{{font-size:0.7rem;color:#64748B;text-transform:uppercase}}
.fn .fv{{font-size:1.3rem;font-weight:800;margin:0.3rem 0}}
.fn .fs{{font-size:0.7rem;color:#64748B}}
.fa{{display:flex;flex-direction:column;align-items:center;color:#475569}}
.fa .ar{{font-size:1.5rem}}
.fa .ls{{font-size:0.65rem;color:#F59E0B}}
.fa .ef{{font-size:0.65rem;color:#94A3B8}}
.mt{{width:100%;border-collapse:collapse;background:#1E293B;border-radius:12px;overflow:hidden;border:1px solid rgba(148,163,184,0.1)}}
.mt th{{background:#166534;color:#fff;padding:0.7rem 1rem;font-size:0.8rem;font-weight:700;text-transform:uppercase;text-align:center}}
.mt td{{padding:0.6rem 1rem;text-align:center;font-size:0.85rem;border-bottom:1px solid rgba(148,163,184,0.08)}}
.mt tr:last-child td{{border-bottom:none}}
.mt .rt{{background:rgba(34,197,94,0.08);font-weight:700}}
.cg2{{color:#22C55E;font-weight:700}}
.cy{{color:#F59E0B;font-weight:700}}
.cr{{color:#EF4444;font-weight:700}}
.lr{{display:flex;gap:1.5rem;justify-content:center;padding:0.6rem;font-size:0.75rem;color:#94A3B8}}
.lr span{{display:flex;align-items:center;gap:0.3rem}}
.ld{{width:10px;height:10px;border-radius:3px;display:inline-block}}
.rf{{text-align:center;padding:1.5rem;color:#475569;font-size:0.75rem;border-top:1px solid rgba(148,163,184,0.1);margin-top:2rem}}
@media print{{body{{background:#0F172A !important}}-webkit-print-color-adjust:exact}}
</style>
</head>
<body>
<div class="rc">
<div class="rh">
<h1>AUTOMATIC DEVIATION ANALYTICS REPORT</h1>
<div class="sub">PT Kalimantan Prima Persada</div>
<div class="dt">Generated: {report_date}</div>
</div>

<div class="sec">
<div class="sh"><h2> Executive KPI Summary</h2></div>
<div class="kr">
<div class="kc" style="--ac:{ob_clr}"><div class="lb">OVERBURDEN (OB)</div><div class="vl">{ob_avg_dev:.2f}%</div><div class="st" style="color:{ob_clr}">{ob_lbl2}</div><div class="dt">{int(matrix.iloc[0]['Total'])} Periods</div></div>
<div class="kc" style="--ac:{ch_clr}"><div class="lb">COAL HAULING (CH)</div><div class="vl">{ch_avg_dev:.2f}%</div><div class="st" style="color:{ch_clr}">{ch_lbl2}</div><div class="dt">{int(matrix.iloc[1]['Total'])} Periods</div></div>
<div class="kc" style="--ac:{cm_clr}"><div class="lb">COAL MINING (CM)</div><div class="vl">{cm_avg_dev:.2f}%</div><div class="st" style="color:{cm_clr}">{cm_lbl2}</div><div class="dt">{int(matrix.iloc[2]['Total'])} Periods</div></div>
<div class="kc" style="--ac:{'#22C55E' if perf >= 70 else '#F59E0B' if perf >= 50 else '#EF4444'}"><div class="lb">OVERALL PERFORMANCE</div><div class="vl">{perf:.1f}%</div><div class="st" style="color:{'#22C55E' if perf >= 70 else '#F59E0B' if perf >= 50 else '#EF4444'}">{'Good' if perf >= 70 else 'Fair' if perf >= 50 else 'Poor'}</div><div class="dt">Avg Eff: {avg_eff:.1f}%</div></div>
</div>
</div>

<div class="sec">
<div class="sh"><h2>⛰️ Section A — Overburden (OB) Monthly Analysis</h2></div>
<div class="cg"><div class="cc">{f_divs[0]}</div><div class="cc">{f_divs[1]}</div></div>
<div class="fc">"""

            for lbl, val, clr in stats_a:
                html_report += f'<div class="fn"><div class="fl">{lbl}</div><div class="fv" style="color:{clr}">{val}</div></div>'

            html_report += f"""</div>
</div>

<div class="sec">
<div class="sh"><h2>🚛 Section B — Coal Hauling (CH) & Coal Mining (CM)</h2></div>
<div class="cg"><div class="cc">{f_divs[2]}</div><div class="cc">{f_divs[3]}</div></div>
<div class="fc">"""

            for lbl, val, clr in stats_b:
                html_report += f'<div class="fn"><div class="fl">{lbl}</div><div class="fv" style="color:{clr}">{val}</div></div>'

            html_report += f"""</div>
</div>

<div class="sec">
<div class="sh"><h2>📈 Section C — Deviation Pattern Analysis</h2></div>
<div class="cg"><div class="cc">{f_divs[4]}</div><div class="cc">{f_divs[5]}</div></div>
<div class="fc">"""

            for lbl, val, clr in stats_c:
                html_report += f'<div class="fn"><div class="fl">{lbl}</div><div class="fv" style="color:{clr}">{val}</div></div>'

            html_report += f"""</div>
</div>

<div class="sec">
<div class="sh"><h2>🔄 Section D — Material Throughput Flow</h2></div>
<div class="cg"><div class="cc">{f_divs[6]}</div><div class="cc">{f_divs[7]}</div></div>
<div class="fc">"""

            for lbl, val, clr in stats_d:
                html_report += f'<div class="fn"><div class="fl">{lbl}</div><div class="fv" style="color:{clr}">{val}</div></div>'

            html_report += f"""</div>
</div>

<div class="sec">
<div class="sh"><h2>📋 Performance Matrix</h2></div>
<table class="mt">
<thead><tr><th>Stage</th><th>Avg |Dev%|</th><th>Normal</th><th>Caution</th><th>Critical</th><th>Total</th><th>Performance</th></tr></thead>
<tbody>
<tr><td><strong>Overburden (OB)</strong></td><td>{ob_avg_dev:.2f}%</td><td class="cg2">{int(matrix.iloc[0]['Normal'])}</td><td class="cy">{int(matrix.iloc[0]['Caution'])}</td><td class="cr">{int(matrix.iloc[0]['Critical'])}</td><td>{int(matrix.iloc[0]['Total'])}</td><td>{int(matrix.iloc[0]['Normal'])/max(int(matrix.iloc[0]['Total']),1)*100:.0f}%</td></tr>
<tr><td><strong>Coal Hauling (CH)</strong></td><td>{ch_avg_dev:.2f}%</td><td class="cg2">{int(matrix.iloc[1]['Normal'])}</td><td class="cy">{int(matrix.iloc[1]['Caution'])}</td><td class="cr">{int(matrix.iloc[1]['Critical'])}</td><td>{int(matrix.iloc[1]['Total'])}</td><td>{int(matrix.iloc[1]['Normal'])/max(int(matrix.iloc[1]['Total']),1)*100:.0f}%</td></tr>
<tr><td><strong>Coal Mining (CM)</strong></td><td>{cm_avg_dev:.2f}%</td><td class="cg2">{int(matrix.iloc[2]['Normal'])}</td><td class="cy">{int(matrix.iloc[2]['Caution'])}</td><td class="cr">{int(matrix.iloc[2]['Critical'])}</td><td>{int(matrix.iloc[2]['Total'])}</td><td>{int(matrix.iloc[2]['Normal'])/max(int(matrix.iloc[2]['Total']),1)*100:.0f}%</td></tr>
<tr class="rt"><td><strong>OVERALL</strong></td><td></td><td class="cg2"><strong>{total_normal}</strong></td><td class="cy"><strong>{total_caution}</strong></td><td class="cr"><strong>{total_critical}</strong></td><td><strong>{total_all}</strong></td><td><strong>{perf:.1f}%</strong></td></tr>
</tbody></table>
</div>

<div class="rf">PT Kalimantan Prima Persada — Mining Volume Deviation Monitoring System<br>Report generated: {report_date} · </div>
</div>
</body></html>"""

            html_bytes = html_report.encode('utf-8')
            st.download_button(
                label="Download HTML",
                data=html_bytes,
                file_name=f"KPP_Visual_Report_{datetime.now().strftime('%Y%m%d')}.html",
                mime="text/html",
                key='dl_html', use_container_width=True)


        with exp3:
            # ═══════ PNG EXPORT v5 — KPP Branding + Legend Fix ═══════
            st.markdown("""
            <div style="background:rgba(22,101,52,0.08);padding:10px 12px;border-radius:8px;
                border-left:3px solid #16A34A;margin-bottom:8px;">
                <div style="font-size:0.82rem;font-weight:700;color:#e5e7eb;"> PNG Report</div>
                <div style="font-size:0.72rem;color:#94A3B8;margin-top:2px;">
                    High-resolution visual report • KPP branding</div>
            </div>
            """, unsafe_allow_html=True)

            try:
                import kaleido
                kaleido_ok = True
            except ImportError:
                kaleido_ok = False
                st.warning("Library 'kaleido' tidak terinstall.")

            if kaleido_ok:
                if st.button(" Generate PNG", key='gen_png',
                             use_container_width=True):
                    with st.spinner("Generating PNG report..."):
                        try:
                            from PIL import Image as PILImage
                            from PIL import ImageDraw, ImageFont

                            # ── KPP Corporate Palette ──
                            KPP_GREEN     = (22, 101, 52)
                            KPP_LIGHT     = (34, 197, 94)
                            KPP_DARK      = (15, 70, 38)
                            KPP_SUBTLE    = (18, 32, 25)
                            BGCOLOR       = (13, 20, 33)
                            CARDBG        = (24, 35, 50)
                            CARD_ALT      = (19, 28, 42)
                            BLUE          = (59, 130, 246)
                            WHITE         = (241, 245, 249)
                            TEXTMAIN      = (226, 232, 240)
                            GRAY          = (148, 163, 184)
                            DARKGRAY      = (100, 116, 139)
                            YELLOW        = (245, 158, 11)
                            RED           = (239, 68, 68)
                            BORDER        = (35, 48, 65)
                            SEPARATOR     = (40, 55, 72)
                            PURPLE        = (167, 139, 250)

                            W = 2400
                            PAD = 48
                            INNER = 32
                            CW = (W - 2*PAD - INNER) // 2
                            CH = 520
                            GAP = 24
                            CGAP = 16
                            HH = 160
                            SH = 52
                            KH = 150
                            STH = 70
                            FH = 70
                            MRH = 46
                            TH = (HH + GAP + SH + KH + GAP
                                  + (SH + CH + STH + GAP) * 4
                                  + SH + MRH * 5 + GAP + FH
                                  + PAD * 2 + 120)

                            canvas = PILImage.new('RGB', (W, TH), BGCOLOR)
                            draw = ImageDraw.Draw(canvas)

                            try:
                                ft = ImageFont.truetype("arial.ttf", 38)
                                fs = ImageFont.truetype("arial.ttf", 17)
                                fsc = ImageFont.truetype("arial.ttf", 18)
                                fkv = ImageFont.truetype("arial.ttf", 38)
                                fkl = ImageFont.truetype("arial.ttf", 12)
                                fks = ImageFont.truetype("arial.ttf", 14)
                                fsm = ImageFont.truetype("arial.ttf", 14)
                                fpv = ImageFont.truetype("arial.ttf", 22)
                                fpl = ImageFont.truetype("arial.ttf", 11)
                                ffo = ImageFont.truetype("arial.ttf", 13)
                                fmx = ImageFont.truetype("arial.ttf", 17)
                                fmh = ImageFont.truetype("arial.ttf", 15)
                                flg = ImageFont.truetype("arial.ttf", 13)
                            except:
                                ft = ImageFont.load_default()
                                fs = fsc = fkv = fkl = fks = ft
                                fsm = fpv = fpl = ffo = fmx = ft
                                fmh = flg = ft

                            y = PAD

                            # ═══ HEADER ═══
                            draw.rounded_rectangle(
                                [PAD, y, W-PAD, y+HH],
                                radius=14, fill=CARDBG, outline=BORDER)
                            for gx in range(PAD+1, W-PAD-1):
                                t = (gx - PAD) / (W - 2*PAD)
                                rc = int(22 + t * 12)
                                gc = int(101 - t * 31)
                                bc = int(52 + t * 0)
                                draw.line([(gx, y), (gx, y+4)],
                                          fill=(rc, gc, bc))

                            logo_ok = False
                            try:
                                lr = PILImage.open(LOGO_PATH).convert("RGBA")
                                ow, oh = lr.size
                                mlh = HH - 50
                                sc = min(mlh / oh, mlh / ow)
                                nw = int(ow * sc)
                                nh = int(oh * sc)
                                lr = lr.resize((nw, nh), PILImage.LANCZOS)
                                lx = PAD + 30
                                ly = y + (HH - nh) // 2
                                canvas.paste(lr, (lx, ly), lr)
                                tx = lx + nw + 28
                                logo_ok = True
                            except Exception:
                                tx = PAD + 40

                            anc = 'lm' if logo_ok else 'mm'
                            txp = tx if logo_ok else W // 2
                            draw.text(
                                (txp, y + 42),
                                "AUTOMATIC DEVIATION ANALYTICS REPORT",
                                fill=KPP_LIGHT, font=ft, anchor=anc)
                            draw.text(
                                (txp, y + 82),
                                "PT Kalimantan Prima Persada"
                                " — TWB Dashboard",
                                fill=GRAY, font=fs, anchor=anc)
                            draw.text(
                                (txp, y + 108),
                                f"Generated: {report_date}",
                                fill=DARKGRAY, font=fsm, anchor=anc)
                            y += HH + GAP

                            # ── Helpers ──
                            def sec_hdr(cy, txt):
                                draw.rounded_rectangle(
                                    [PAD, cy, W-PAD, cy+SH],
                                    radius=10, fill=KPP_SUBTLE)
                                draw.rounded_rectangle(
                                    [PAD, cy+6, PAD+5, cy+SH-6],
                                    radius=2, fill=KPP_LIGHT)
                                draw.text(
                                    (PAD+24, cy + SH//2), txt,
                                    fill=WHITE, font=fsc, anchor='lm')
                                return cy + SH + 8

                            def stat_pill(sx, sy, sw, sh, sl, sv, sc):
                                hc = sc.lstrip('#')
                                rgb = tuple(
                                    int(hc[j:j+2], 16) for j in (0,2,4))
                                bg = (rgb[0]//10+13, rgb[1]//10+18,
                                      rgb[2]//10+30)
                                draw.rounded_rectangle(
                                    [sx, sy, sx+sw, sy+sh],
                                    radius=8, fill=bg, outline=BORDER)
                                draw.text(
                                    (sx+sw//2, sy+14), sv,
                                    fill=rgb, font=fpv, anchor='mm')
                                draw.text(
                                    (sx+sw//2, sy+40), sl,
                                    fill=DARKGRAY, font=fpl, anchor='mm')

                            def draw_legend(cy, items):
                                total_w = sum(
                                    len(n)*8 + 30 for _, n in items) + 20
                                lx = PAD + (W - 2*PAD - total_w) // 2
                                draw.rounded_rectangle(
                                    [lx-10, cy, lx+total_w+10, cy+24],
                                    radius=6, fill=CARDBG, outline=BORDER)
                                for clr, name in items:
                                    draw.rounded_rectangle(
                                        [lx, cy+7, lx+12, cy+17],
                                        radius=2, fill=clr)
                                    lx += 16
                                    draw.text(
                                        (lx, cy+12), name,
                                        fill=TEXTMAIN, font=flg,
                                        anchor='lm')
                                    lx += len(name)*8 + 14
                                return cy + 30

                            # ═══ KPI SUMMARY ═══
                            y = sec_hdr(y, "EXECUTIVE KPI SUMMARY")
                            kpi = [
                                ("OVERBURDEN (OB)",
                                 f"{ob_avg_dev:.2f}%", ob_lbl2, ob_clr,
                                 f"{int(matrix.iloc[0]['Total'])} Periods"),
                                ("COAL HAULING (CH)",
                                 f"{ch_avg_dev:.2f}%", ch_lbl2, ch_clr,
                                 f"{int(matrix.iloc[1]['Total'])} Periods"),
                                ("COAL MINING (CM)",
                                 f"{cm_avg_dev:.2f}%", cm_lbl2, cm_clr,
                                 f"{int(matrix.iloc[2]['Total'])} Periods"),
                                ("OVERALL PERFORMANCE",
                                 f"{perf:.1f}%",
                                 "Good" if perf >= 70
                                 else "Fair" if perf >= 50 else "Poor",
                                 '#22C55E' if perf >= 70
                                 else '#F59E0B' if perf >= 50
                                 else '#EF4444',
                                 f"Avg Eff: {avg_eff:.1f}%"),
                            ]
                            kw = (W - 2*PAD - 3*CGAP) // 4
                            for i, (lb, vl, st_txt, cl, dt) in enumerate(
                                    kpi):
                                kx = PAD + i*(kw + CGAP)
                                draw.rounded_rectangle(
                                    [kx, y, kx+kw, y+KH],
                                    radius=12, fill=CARDBG, outline=BORDER)
                                hx = cl.lstrip('#')
                                ac = tuple(
                                    int(hx[j:j+2], 16) for j in (0,2,4))
                                draw.rounded_rectangle(
                                    [kx+12, y, kx+kw-12, y+4],
                                    radius=2, fill=ac)
                                draw.text(
                                    (kx+kw//2, y+26), lb,
                                    fill=DARKGRAY, font=fkl, anchor='mm')
                                draw.text(
                                    (kx+kw//2, y+68), vl,
                                    fill=WHITE, font=fkv, anchor='mm')
                                bw = len(st_txt)*9 + 24
                                bx = kx + (kw - bw)//2
                                bbg = (ac[0]//5+13, ac[1]//5+18,
                                       ac[2]//5+28)
                                draw.rounded_rectangle(
                                    [bx, y+92, bx+bw, y+112],
                                    radius=10, fill=bbg, outline=ac)
                                draw.text(
                                    (kx+kw//2, y+102), st_txt,
                                    fill=ac, font=fks, anchor='mm')
                                draw.text(
                                    (kx+kw//2, y+132), dt,
                                    fill=DARKGRAY, font=fkl, anchor='mm')
                            y += KH + GAP

                            # ═══ CHARTS ═══
                            ctp = dict(
                                font=dict(
                                    family='Inter, Arial, sans-serif',
                                    color='#E2E8F0', size=12),
                                plot_bgcolor='#182332',
                                paper_bgcolor='#182332',
                                margin=dict(l=65, r=40, t=70, b=50),
                                hovermode='x unified')
                            fp = [
                                make_fig1(ctp), make_fig2(ctp),
                                make_fig3(ctp), make_fig4(ctp),
                                make_fig5(ctp), make_fig6(ctp),
                                make_fig7(ctp), make_fig8(ctp)]
                            for fi in [6, 7]:
                                fp[fi].update_layout(
                                    showlegend=False,
                                    margin=dict(l=65, r=40, t=70, b=50))

                            sec_d_legends = [
                                [((59,130,246), "CM TWB"),
                                 ((34,197,94), "CH TWB"),
                                 ((167,139,250), "Sales")],
                                [((34,197,94), "Overall Eff."),
                                 ((59,130,246), "CH Eff.")],
                            ]

                            secs = [
                                ("OVERBURDEN (OB)"
                                 " MONTHLY ANALYSIS",
                                 [0, 1], stats_a, None),
                                ("COAL HAULING (CH)"
                                 " & COAL MINING (CM)",
                                 [2, 3], stats_b, None),
                                ("DEVIATION PATTERN"
                                 " ANALYSIS",
                                 [4, 5], stats_c, None),
                                ("MATERIAL FLOW"
                                 " ANALYSIS",
                                 [6, 7], stats_d, sec_d_legends),
                            ]

                            for stitle, fidx, sstats, legends in secs:
                                y = sec_hdr(y, stitle)

                                if legends:
                                    for ci, lg in enumerate(legends):
                                        lx_start = (PAD + ci*(CW + INNER)
                                                     + CW//2)
                                        tw = sum(
                                            len(n)*8+30 for _,n in lg)+20
                                        rx = lx_start - tw//2 - 10
                                        draw.rounded_rectangle(
                                            [rx, y, rx+tw+20, y+24],
                                            radius=6, fill=CARDBG,
                                            outline=BORDER)
                                        px = rx + 10
                                        for clr, nm in lg:
                                            draw.rounded_rectangle(
                                                [px, y+7, px+12, y+17],
                                                radius=2, fill=clr)
                                            px += 16
                                            draw.text(
                                                (px, y+12), nm,
                                                fill=TEXTMAIN, font=flg,
                                                anchor='lm')
                                            px += len(nm)*8 + 14
                                    y += 30

                                for ci, fi in enumerate(fidx):
                                    fig = fp[fi]
                                    fig.update_layout(
                                        width=CW, height=CH)
                                    ib = fig.to_image(
                                        format='png', scale=2,
                                        engine='kaleido')
                                    ci_img = PILImage.open(
                                        io.BytesIO(ib))
                                    ci_img = ci_img.resize(
                                        (CW, CH), PILImage.LANCZOS)
                                    cx = PAD + ci*(CW + INNER)
                                    draw.rounded_rectangle(
                                        [cx-2, y-2,
                                         cx+CW+2, y+CH+2],
                                        radius=14, fill=BORDER)
                                    draw.rounded_rectangle(
                                        [cx, y, cx+CW, y+CH],
                                        radius=12, fill=CARDBG)
                                    canvas.paste(ci_img, (cx, y))
                                y += CH + 10

                                ns = len(sstats)
                                pg = 10
                                pw = (W - 2*PAD - (ns-1)*pg) // ns
                                for si, (sl, sv, sc) in enumerate(
                                        sstats):
                                    sx = PAD + si*(pw + pg)
                                    stat_pill(
                                        sx, y, pw, STH-8, sl, sv, sc)
                                y += STH + GAP

                            # ═══ PERFORMANCE MATRIX ═══
                            y = sec_hdr(y, "PERFORMANCE MATRIX")
                            cols = [
                                'Stage', 'Avg Dev(%)', 'Normal',
                                'Caution', 'Critical', 'Total',
                                'Performance']
                            cw2 = (W - 2*PAD) // len(cols)

                            draw.rounded_rectangle(
                                [PAD, y, W-PAD, y+MRH],
                                radius=8, fill=KPP_GREEN)
                            for ci, ct in enumerate(cols):
                                draw.text(
                                    (PAD + ci*cw2 + cw2//2,
                                     y + MRH//2),
                                    ct, fill=WHITE,
                                    font=fmh, anchor='mm')
                            y += MRH

                            rows = [
                                ('Overburden (OB)',
                                 f'{ob_avg_dev:.2f}%',
                                 str(int(matrix.iloc[0]['Normal'])),
                                 str(int(matrix.iloc[0]['Caution'])),
                                 str(int(matrix.iloc[0]['Critical'])),
                                 str(int(matrix.iloc[0]['Total'])),
                                 f'{int(matrix.iloc[0]["Normal"])/max(int(matrix.iloc[0]["Total"]),1)*100:.0f}%'),
                                ('Coal Hauling (CH)',
                                 f'{ch_avg_dev:.2f}%',
                                 str(int(matrix.iloc[1]['Normal'])),
                                 str(int(matrix.iloc[1]['Caution'])),
                                 str(int(matrix.iloc[1]['Critical'])),
                                 str(int(matrix.iloc[1]['Total'])),
                                 f'{int(matrix.iloc[1]["Normal"])/max(int(matrix.iloc[1]["Total"]),1)*100:.0f}%'),
                                ('Coal Mining (CM)',
                                 f'{cm_avg_dev:.2f}%',
                                 str(int(matrix.iloc[2]['Normal'])),
                                 str(int(matrix.iloc[2]['Caution'])),
                                 str(int(matrix.iloc[2]['Critical'])),
                                 str(int(matrix.iloc[2]['Total'])),
                                 f'{int(matrix.iloc[2]["Normal"])/max(int(matrix.iloc[2]["Total"]),1)*100:.0f}%'),
                            ]
                            cc = [TEXTMAIN, YELLOW, KPP_LIGHT,
                                  YELLOW, RED, GRAY, BLUE]
                            for ri, rd in enumerate(rows):
                                rbg = CARDBG if ri % 2 == 0 else CARD_ALT
                                draw.rectangle(
                                    [PAD, y, W-PAD, y+MRH],
                                    fill=rbg, outline=BORDER)
                                for ci, v in enumerate(rd):
                                    draw.text(
                                        (PAD + ci*cw2 + cw2//2,
                                         y + MRH//2),
                                        v, fill=cc[ci],
                                        font=fmx, anchor='mm')
                                y += MRH

                            draw.rounded_rectangle(
                                [PAD, y, W-PAD, y+MRH],
                                radius=0, fill=KPP_SUBTLE,
                                outline=BORDER)
                            ov = ['OVERALL', '—',
                                  str(total_normal),
                                  str(total_caution),
                                  str(total_critical),
                                  str(total_all),
                                  f'{perf:.1f}%']
                            oc = [KPP_LIGHT, DARKGRAY, KPP_LIGHT,
                                  YELLOW, RED, TEXTMAIN, BLUE]
                            for ci, v in enumerate(ov):
                                draw.text(
                                    (PAD + ci*cw2 + cw2//2,
                                     y + MRH//2),
                                    v, fill=oc[ci],
                                    font=fmx, anchor='mm')
                            y += MRH + GAP

                            # ═══ FOOTER ═══
                            draw.line(
                                [(PAD+60, y+8), (W-PAD-60, y+8)],
                                fill=SEPARATOR, width=1)
                            draw.text(
                                (W//2, y + 32),
                                "PT Kalimantan Prima Persada"
                                " — Automatic Deviation Calculation"
                                " Dashboard System",
                                fill=DARKGRAY, font=ffo, anchor='mm')
                            draw.text(
                                (W//2, y + 52),
                                f"Generated: {report_date}"
                                " — ",
                                fill=(70, 82, 100),
                                font=ffo, anchor='mm')
                            y += FH

                            canvas = canvas.crop((0, 0, W, y + PAD))
                            buf = io.BytesIO()
                            canvas.save(
                                buf, format='PNG',
                                quality=95, optimize=True)
                            buf.seek(0)

                            st.success(
                                f"✅ PNG generated"
                                f" — {canvas.size[0]}×{canvas.size[1]}px")
                            st.image(
                                canvas,
                                caption="Preview — Full Resolution",
                                use_container_width=True)
                            st.download_button(
                                label=" Download PNG Report",
                                data=buf.getvalue(),
                                file_name=(
                                    f"KPP_Report_"
                                    f"{datetime.now().strftime('%Y%m%d')}"
                                    f".png"),
                                mime="image/png",
                                key='dl_png',
                                use_container_width=True)

                        except Exception as e:
                            st.error(f"Error: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())
                
if __name__ == "__main__":
    main()
